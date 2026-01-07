from fastapi import FastAPI, APIRouter, HTTPException, File, UploadFile, BackgroundTasks, Request, Response, Body, Query
from fastapi.responses import FileResponse
from fastapi.middleware.trustedhost import TrustedHostMiddleware
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict
import uuid
from datetime import datetime, timezone
from csv_upload import parse_shopify_orders_csv
import pandas as pd
import io
import httpx
from shopify_sync import ShopifyOrderSync
from shopify_sync_async import ShopifyAsyncSync
from datetime import timedelta
from tcs_tracking import TCSTracker
from scheduler import get_scheduler
import asyncio
from whatsapp_service import whatsapp_service
from whatsapp_webhook import router as whatsapp_webhook_router
from whatsapp_marketing import whatsapp_marketing, MARKETING_TEMPLATES
from dynamic_pricing import DynamicPricingEngine
from meta_whatsapp_service import MetaWhatsAppService
from whatsapp_crm_routes import whatsapp_router
from tracking_routes import tracking_router
from finance_reconciliation import get_finance_reconciliation
from routes.facebook import facebook_router, set_database as set_facebook_db
from routes.finance import finance_router, set_database as set_finance_db
from routes.pricing import pricing_router, set_dependencies as set_pricing_deps
from routes.tcs import tcs_router, set_database as set_tcs_db
from routes.customers import customers_router, set_dependencies as set_customers_deps
from routes.clearance import clearance_router, set_dependencies as set_clearance_deps
from routes.users import users_router, set_dependencies as set_users_deps
from routes.tenants import tenants_router, set_dependencies as set_tenants_deps
from routes.meta_ads import meta_ads_router, set_dependencies as set_meta_ads_deps
from routes.whatsapp_embedded import whatsapp_embedded_router, set_database as set_whatsapp_embedded_db
from routes.subscriptions import subscriptions_router
from routes.shopify_webhooks import shopify_webhooks_router
from routes.api_keys import api_keys_router, set_database as set_api_keys_db
from routes.shopify_oauth import shopify_oauth_router, set_database as set_shopify_oauth_db
from routes.lead_ads import lead_ads_router, set_database as set_lead_ads_db
from routes.super_admin import super_admin_router, set_database as set_super_admin_db
from routes.dwz56 import router as dwz56_router
from routes.alibaba_1688 import router as alibaba_1688_router
from routes.fulfillment import router as fulfillment_router
from routes.product_scraper import router as product_scraper_router
from routes.tmapi_buyer import router as tmapi_buyer_router
from routes.storefront import router as storefront_router, set_database as set_storefront_db
from routes.storefront_cms import router as storefront_cms_router, set_database as set_storefront_cms_db
from routes.warehouse import router as warehouse_router, set_database as set_warehouse_db
from routes.shopify_sync import router as shopify_sync_router, set_dependencies as set_shopify_sync_deps
from routes.whatsapp_api import router as whatsapp_api_router, set_dependencies as set_whatsapp_api_deps
from routes.marketing import router as marketing_router, set_database as set_marketing_db
from routes.settings import router as settings_router, set_database as set_settings_db
from services.fulfillment_webhooks import router as fulfillment_webhooks_router
from services.whatsapp_notifications import router as whatsapp_notifications_router
from services.fulfillment_pipeline_service import router as fulfillment_pipeline_router
from services.shopify_fulfillment_sync import router as shopify_fulfillment_sync_router
from services.dwz56_sync_service import router as dwz56_sync_router
from services.email_notification_service import router as email_notification_router
from services.sync_service import router as sync_service_router, set_dependencies as set_sync_dependencies
from inventory_clearance_engine import InventoryClearanceEngine


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Initialize pricing engine with db
pricing_engine = DynamicPricingEngine(db)

# Initialize clearance engine
clearance_engine = InventoryClearanceEngine(db)

# Shopify sync job tracking (in-memory)
shopify_sync_jobs = {}

# Initialize modular routers with database and dependencies
set_facebook_db(db)
set_finance_db(db)
set_tcs_db(db)
set_customers_deps(db)
set_pricing_deps(db, pricing_engine)
set_clearance_deps(db, clearance_engine)
set_users_deps(db)
set_tenants_deps(db)
set_meta_ads_deps(db)
set_whatsapp_embedded_db(db)
set_api_keys_db(db)
set_shopify_oauth_db(db)
set_lead_ads_db(db)
set_super_admin_db(db)
set_storefront_db(db)
set_storefront_cms_db(db)
set_warehouse_db(db)
set_sync_dependencies(db)
set_shopify_sync_deps(db)
set_marketing_db(db)
set_settings_db(db)

# Set WhatsApp API dependencies
try:
    set_whatsapp_api_deps(db, whatsapp_service, MARKETING_TEMPLATES if 'MARKETING_TEMPLATES' in dir() else None)
except Exception as e:
    logger.warning(f"Could not set WhatsApp API dependencies: {e}")

# ==================== Initialize Default Users ====================
async def init_default_users():
    """Create default admin user if not exists"""
    import hashlib
    
    try:
        def hash_password(password: str) -> str:
            return hashlib.sha256(password.encode()).hexdigest()
        
        # Check if admin user exists
        admin_user = await db.users.find_one({"username": "admin"})
        if not admin_user:
            default_users = [
                {
                    "id": "admin-default-user",
                    "username": "admin",
                    "password": hash_password("admin"),
                    "full_name": "Administrator",
                    "role": "admin",
                    "status": "active",
                    "stores": [],
                    "created_at": datetime.now(timezone.utc).isoformat()
                },
                {
                    "id": "manager-default-user",
                    "username": "manager1",
                    "password": hash_password("password"),
                    "full_name": "Manager",
                    "role": "manager",
                    "status": "active",
                    "stores": [],
                    "created_at": datetime.now(timezone.utc).isoformat()
                },
                {
                    "id": "viewer-default-user",
                    "username": "viewer1",
                    "password": hash_password("password"),
                    "full_name": "Viewer",
                    "role": "viewer",
                    "status": "active",
                    "stores": [],
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
            ]
            
            for user in default_users:
                await db.users.update_one(
                    {"username": user["username"]},
                    {"$setOnInsert": user},
                    upsert=True
                )
            print("✅ Default users created: admin, manager1, viewer1")
        else:
            print("✅ Default users already exist")
    except Exception as e:
        print(f"⚠️ Could not initialize default users: {str(e)}")
        print("⚠️ Users may need to be created manually or database permissions need to be checked")

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")


# ==================== PERFORMANCE: In-Memory Caches ====================
# These caches store frequently accessed data to avoid repeated DB queries

# Inventory cache - refreshed every 5 minutes
_inventory_cache = {
    "data": {},  # SKU -> {cost, profit}
    "stock_skus": set(),  # Set of SKUs in stock
    "last_updated": None,
    "ttl_seconds": 300  # 5 minutes
}

async def get_inventory_cache():
    """Get cached inventory data, refresh if stale"""
    global _inventory_cache
    now = datetime.now(timezone.utc)
    
    # Check if cache is valid
    if (_inventory_cache["last_updated"] and 
        (now - _inventory_cache["last_updated"]).total_seconds() < _inventory_cache["ttl_seconds"]):
        return _inventory_cache
    
    # Refresh cache
    try:
        # Fetch inventory for cost/profit calculations
        inventory_items = await db.inventory_v2.find(
            {}, 
            {"_id": 0, "sku": 1, "cost": 1, "profit": 1, "quantity": 1}
        ).to_list(10000)
        
        inventory_map = {}
        stock_skus = set()
        
        for item in inventory_items:
            sku = item.get("sku", "").upper()
            if sku:
                inventory_map[sku] = {
                    "cost": item.get("cost", 0) or 0,
                    "profit": item.get("profit", 0) or 0
                }
                if (item.get("quantity", 0) or 0) > 0:
                    stock_skus.add(sku)
        
        # Also check old stock collection
        stock_items = await db.stock.find({}, {"_id": 0, "sku": 1}).to_list(10000)
        for item in stock_items:
            sku = item.get("sku", "").upper()
            if sku:
                stock_skus.add(sku)
        
        _inventory_cache = {
            "data": inventory_map,
            "stock_skus": stock_skus,
            "last_updated": now,
            "ttl_seconds": 300
        }
        
        logger.info(f"✅ Inventory cache refreshed: {len(inventory_map)} items, {len(stock_skus)} in stock")
    except Exception as e:
        logger.error(f"❌ Error refreshing inventory cache: {str(e)}")
    
    return _inventory_cache

async def invalidate_inventory_cache():
    """Force refresh of inventory cache"""
    global _inventory_cache
    _inventory_cache["last_updated"] = None
    return await get_inventory_cache()

# ==================== END PERFORMANCE CACHES ====================

# Startup and shutdown events for scheduler
@app.on_event("startup")
async def startup_event():
    """Start background scheduler on server startup"""
    logger.info("🚀 Starting server...")
    
    # Initialize default users FIRST
    await init_default_users()
    
    # Fix stores with missing shopify_domain
    async def fix_store_domains_on_startup():
        try:
            logger.info("🔧 Checking stores for missing domains...")
            stores = await db.stores.find({"$or": [
                {"shopify_domain": None},
                {"shopify_domain": {"$exists": False}}
            ]}).to_list(None)
            
            updated = 0
            for store in stores:
                shop_url = store.get('shop_url')
                if shop_url:
                    await db.stores.update_one(
                        {"_id": store["_id"]},
                        {"$set": {"shopify_domain": shop_url}}
                    )
                    updated += 1
                    logger.info(f"  Fixed domain for {store.get('store_name')}: {shop_url}")
            
            if updated > 0:
                logger.info(f"✅ Fixed {updated} store domains")
            else:
                logger.info("✅ All store domains are set correctly")
        except Exception as e:
            logger.error(f"❌ Error fixing store domains: {str(e)}")
    
    await fix_store_domains_on_startup()
    
    scheduler = get_scheduler()
    scheduler.start()
    logger.info("✅ Background scheduler initialized")
    
    # Create database indexes for performance
    async def create_indexes():
        try:
            logger.info("📊 Creating database indexes for performance...")
            
            # Customers collection indexes
            await db.customers.create_index("customer_id", unique=True, background=True)
            await db.customers.create_index("order_number", background=True)
            await db.customers.create_index("order_number_int", background=True)
            await db.customers.create_index("store_name", background=True)
            await db.customers.create_index("fulfillment_status", background=True)
            await db.customers.create_index("delivery_status", background=True)
            await db.customers.create_index("payment_status", background=True)
            await db.customers.create_index("last_order_date", background=True)
            await db.customers.create_index("created_at", background=True)
            await db.customers.create_index("tracking_number", background=True)
            
            # Compound indexes for common queries
            await db.customers.create_index([
                ("fulfillment_status", 1),
                ("created_at", -1)
            ], background=True)
            await db.customers.create_index([
                ("store_name", 1),
                ("order_number_int", -1)
            ], background=True)
            
            # Inventory indexes
            await db.inventory_v2.create_index("sku", unique=True, background=True)
            await db.inventory_v2.create_index("store_name", background=True)
            await db.inventory_v2.create_index("quantity", background=True)
            
            # Stock collection index
            await db.stock.create_index("sku", background=True)
            
            logger.info("✅ Database indexes created successfully")
        except Exception as e:
            logger.warning(f"⚠️ Index creation warning (may already exist): {str(e)}")
    
    # Warm up inventory cache
    async def warmup_inventory_cache():
        try:
            logger.info("🔥 Warming up inventory cache...")
            await get_inventory_cache()
            logger.info("✅ Inventory cache warmed up")
        except Exception as e:
            logger.error(f"❌ Error warming up inventory cache: {str(e)}")
    
    # Start dynamic pricing cache warmup in background
    async def warmup_pricing_cache():
        try:
            from dynamic_pricing_engine import dynamic_pricing_engine
            logger.info("🔥 Warming up dynamic pricing cache...")
            
            # Check if cache exists
            cached = await db.dynamic_pricing_cache.find_one({"type": "analysis_report"})
            if not cached:
                logger.info("📊 Running initial pricing analysis (60-day lookback)...")
                result = await dynamic_pricing_engine.analyze_product_velocity(db, days_lookback=60)
                if result.get('success'):
                    await db.dynamic_pricing_cache.update_one(
                        {"type": "analysis_report"},
                        {"$set": {
                            "type": "analysis_report",
                            "data": result,
                            "last_updated": datetime.now(timezone.utc).isoformat()
                        }},
                        upsert=True
                    )
                    logger.info("✅ Pricing cache initialized")
            else:
                logger.info("✅ Pricing cache already exists")
        except Exception as e:
            logger.error(f"❌ Error warming up pricing cache: {str(e)}")
    
    # Run startup tasks in background
    asyncio.create_task(create_indexes())
    asyncio.create_task(warmup_inventory_cache())
    asyncio.create_task(warmup_pricing_cache())


@app.on_event("shutdown")
async def shutdown_event():
    """Stop background scheduler on server shutdown"""
    logger.info("🛑 Shutting down server...")
    scheduler = get_scheduler()
    scheduler.stop()
    logger.info("✅ Background scheduler stopped")


# Define Models
class Customer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    customer_id: str
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    country_code: Optional[str] = None
    shoe_sizes: List[str] = []
    order_skus: List[str] = []  # SKUs from customer orders
    order_count: int = 0
    order_number: Optional[str] = None  # Shopify order number
    last_order_date: Optional[str] = None
    total_spent: float = 0.0
    store_name: Optional[str] = None
    messaged: bool = False
    last_messaged_at: Optional[str] = None
    message_count: int = 0
    messaged_by: Optional[str] = None
    converted: Optional[bool] = None
    conversion_notes: Optional[str] = None
    sale_amount: Optional[float] = None
    stock_status: Optional[str] = None  # "in_stock", "out_of_stock", "partial"
    fulfillment_status: Optional[str] = None  # "unfulfilled", "fulfilled", "partially_fulfilled"
    fulfilled_at: Optional[str] = None  # Date when order was fulfilled
    tracking_number: Optional[str] = None
    tracking_company: Optional[str] = None
    tracking_url: Optional[str] = None
    delivery_status: Optional[str] = None  # TCS delivery status: DELIVERED, OUT_FOR_DELIVERY, IN_TRANSIT, etc.
    delivery_location: Optional[str] = None  # Current TCS facility location
    delivery_updated_at: Optional[str] = None  # Last TCS status update time
    payment_status: Optional[str] = None  # Shopify payment status: paid, pending, refunded, partially_refunded, voided
    payment_method: Optional[str] = None  # Payment method used
    cod_payment_status: Optional[str] = None  # TCS COD payment status: PENDING, COLLECTED, REMITTED, HELD, REVERSED
    cod_amount: Optional[float] = None  # COD amount to be collected
    amount_paid: Optional[float] = None  # Amount paid by customer (from TCS Payment API)
    payment_balance: Optional[float] = None  # Remaining balance after deducting delivery charges
    delivery_charges: Optional[float] = None  # TCS delivery charges
    parcel_weight: Optional[float] = None  # Parcel weight in kg
    booking_date: Optional[str] = None  # TCS booking date
    delivery_date: Optional[str] = None  # TCS delivery date
    payment_date: Optional[str] = None  # COD payment date/time (when payment was collected)
    collection_date: Optional[str] = None  # COD collection date
    remittance_date: Optional[str] = None  # COD remittance date
    remittance_amount: Optional[float] = None  # Amount remitted to merchant
    cod_collection_date: Optional[str] = None  # Date when COD was collected by TCS
    cod_remittance_date: Optional[str] = None  # Date when COD was remitted to merchant
    cod_remittance_amount: Optional[float] = None  # Amount remitted to merchant
    abandoned_checkout: Optional[bool] = None  # Whether customer has abandoned checkouts
    abandoned_checkout_value: Optional[float] = None  # Value of abandoned items
    cost: Optional[float] = None  # Total cost from inventory (sum of all SKU costs)
    profit: Optional[float] = None  # Total profit (sum of all SKU profits)


class Agent(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    password: str
    full_name: str
    role: str = "agent"
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class AgentLogin(BaseModel):
    username: str
    password: str


class AgentCreate(BaseModel):
    username: str
    password: str
    full_name: str


class Store(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    store_name: str
    shop_url: str
    shopify_domain: Optional[str] = None
    shopify_token: Optional[str] = None
    last_synced_at: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class StoreCreate(BaseModel):
    store_name: str
    shop_url: str
    shopify_domain: Optional[str] = None
    shopify_token: Optional[str] = None


class InventoryItemCreate(BaseModel):
    sku: str
    title: str
    quantity: int = 0
    store_name: str = "tnvcollectionpk"
    price: Optional[float] = None
    compare_at_price: Optional[float] = None
    status: str = "active"


class InventoryItemUpdate(BaseModel):
    title: Optional[str] = None
    quantity: Optional[int] = None
    price: Optional[float] = None
    compare_at_price: Optional[float] = None
    status: Optional[str] = None


class DeliveryStatusUpdate(BaseModel):
    status: str
    notes: Optional[str] = None
    tracking_number: Optional[str] = None


class WhatsAppRequest(BaseModel):
    phone: str
    country_code: Optional[str] = None


class BulkMessageRequest(BaseModel):
    customer_ids: List[str]
    message_template: Optional[str] = None
    delay_seconds: int = 5  # Delay between messages to avoid ban


@api_router.get("/scheduler/status")
async def get_scheduler_status():
    """
    Get background scheduler status and next run times
    """
    scheduler = get_scheduler()
    status = scheduler.get_status()
    return status


@api_router.post("/scheduler/trigger-shopify-sync")
async def trigger_manual_shopify_sync():
    """
    Manually trigger Shopify sync (in addition to automatic hourly sync)
    """
    try:
        scheduler = get_scheduler()
        scheduler.sync_shopify_orders()
        return {
            "success": True,
            "message": "Shopify sync triggered manually (running in background)"
        }
    except Exception as e:
        logger.error(f"Error triggering manual sync: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/scheduler/trigger-tcs-sync")
async def trigger_manual_tcs_sync():
    """
    Manually trigger TCS delivery status sync
    """
    try:
        scheduler = get_scheduler()
        scheduler.sync_tcs_deliveries()
        return {
            "success": True,
            "message": "TCS delivery sync triggered manually (running in background)"
        }
    except Exception as e:
        logger.error(f"Error triggering manual TCS sync: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

    customer_ids: List[str]
    message_template: Optional[str] = None
    delay_seconds: int = 5  # Delay between messages to avoid ban


# Add your routes to the router instead of directly to app
@api_router.get("/")
async def root():
    return {"message": "Customer Manager API"}


@api_router.post("/agents/register")
async def register_agent(agent: AgentCreate):
    """
    Register a new agent
    """
    try:
        # Check if username exists
        existing = await db.agents.find_one({"username": agent.username})
        if existing:
            raise HTTPException(status_code=400, detail="Username already exists")
        
        # Create agent with hashed password (simple hash for demo)
        import hashlib
        hashed_password = hashlib.sha256(agent.password.encode()).hexdigest()
        
        agent_obj = Agent(
            username=agent.username,
            password=hashed_password,
            full_name=agent.full_name
        )
        
        doc = agent_obj.model_dump()
        await db.agents.insert_one(doc)
        
        # Return without password
        return {
            "id": agent_obj.id,
            "username": agent_obj.username,
            "full_name": agent_obj.full_name,
            "role": agent_obj.role
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error registering agent: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to register agent: {str(e)}")


@api_router.post("/agents/login")
async def login_agent(credentials: AgentLogin):
    """
    Agent login with bcrypt password verification
    """
    try:
        import bcrypt
        import hashlib
        
        def verify_password(password: str, hashed: str) -> bool:
            """Verify password against bcrypt or SHA256 hash"""
            try:
                return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))
            except Exception:
                # Fallback for old SHA256 hashes
                sha256_hash = hashlib.sha256(password.encode()).hexdigest()
                return sha256_hash == hashed
        
        agent = await db.agents.find_one({"username": credentials.username})
        
        if not agent:
            raise HTTPException(status_code=401, detail="Invalid username or password")
        
        # Verify password (supports both bcrypt and legacy SHA256)
        if not verify_password(credentials.password, agent["password"]):
            raise HTTPException(status_code=401, detail="Invalid username or password")
        
        # Update last login
        await db.agents.update_one(
            {"username": credentials.username},
            {"$set": {"last_login": datetime.now(timezone.utc).isoformat()}}
        )
        
        return {
            "success": True,
            "agent": {
                "id": agent["id"],
                "username": agent["username"],
                "full_name": agent["full_name"],
                "role": agent["role"]
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error logging in: {str(e)}")
        raise HTTPException(status_code=500, detail="Login failed")


class PasswordChange(BaseModel):
    username: str
    current_password: str
    new_password: str


@api_router.post("/setup/create-demo-user")
async def create_demo_user_endpoint():
    """
    ONE-TIME SETUP: Create demo user for Meta app review
    This endpoint can only be called once - it will fail if demo user already exists
    """
    try:
        import hashlib
        
        # Demo user credentials
        username = "demo_reviewer"
        password = "MetaReview2024!"
        
        # Check if demo user already exists
        existing_user = await db.agents.find_one({"username": username})
        
        if existing_user:
            return {
                "success": True,
                "message": "Demo user already exists",
                "username": username,
                "note": "Demo user was already created"
            }
        
        # Hash password
        hashed_password = hashlib.sha256(password.encode()).hexdigest()
        
        # Create demo user
        demo_user = {
            "id": str(uuid.uuid4()),
            "username": username,
            "password": hashed_password,
            "full_name": "Meta Reviewer",
            "role": "demo",
            "email": "reviewer@meta.com",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "last_login": None,
            "active": True,
            "permissions": [
                "whatsapp_inbox",
                "whatsapp_templates",
                "whatsapp_campaigns",
                "whatsapp_analytics"
            ]
        }
        
        # Insert into database
        result = await db.agents.insert_one(demo_user)
        
        logger.info("✅ Demo user created successfully for Meta app review")
        
        return {
            "success": True,
            "message": "Demo user created successfully!",
            "username": username,
            "password": password,
            "role": "demo",
            "note": "Demo user can only access WhatsApp CRM features"
        }
        
    except Exception as e:
        logger.error(f"Error creating demo user: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to create demo user: {str(e)}")


    username: str
    current_password: str
    new_password: str


@api_router.post("/agents/change-password")
async def change_agent_password(password_data: PasswordChange):
    """
    Change agent password with bcrypt encryption
    """
    try:
        import bcrypt
        import hashlib
        
        def verify_password(password: str, hashed: str) -> bool:
            """Verify password against bcrypt or SHA256 hash"""
            try:
                return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))
            except Exception:
                # Fallback for old SHA256 hashes
                sha256_hash = hashlib.sha256(password.encode()).hexdigest()
                return sha256_hash == hashed
        
        def hash_password(password: str) -> str:
            """Hash password using bcrypt"""
            salt = bcrypt.gensalt(rounds=12)
            return bcrypt.hashpw(password.encode('utf-8'), salt).decode('utf-8')
        
        # Find agent
        agent = await db.agents.find_one({"username": password_data.username})
        
        if not agent:
            raise HTTPException(status_code=404, detail="User not found")
        
        # Verify current password
        if not verify_password(password_data.current_password, agent["password"]):
            raise HTTPException(status_code=401, detail="Current password is incorrect")
        
        # Validate new password
        if len(password_data.new_password) < 6:
            raise HTTPException(status_code=400, detail="New password must be at least 6 characters")
        
        # Update password with bcrypt encryption
        new_hashed = hash_password(password_data.new_password)
        result = await db.agents.update_one(
            {"username": password_data.username},
            {"$set": {
                "password": new_hashed,
                "password_encrypted": True,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Agent not found")
        
        logger.info(f"Password changed for agent: {password_data.username}")
        
        return {
            "success": True,
            "message": "Password changed successfully. Your new password is securely encrypted."
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error changing password: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to change password")


# ==================== Sync Single Order (Utility) ====================
@api_router.post("/sync-single-order/{store_name}/{shopify_order_id}")
async def sync_single_order(store_name: str, shopify_order_id: str):
    """
    Manually sync a single order from Shopify by its ID
    """
    try:
        # Get store credentials
        store = await db.stores.find_one({"store_name": store_name}, {"_id": 0})
        if not store:
            raise HTTPException(status_code=404, detail=f"Store {store_name} not found")
        
        shopify_domain = store.get("shopify_domain")
        shopify_token = store.get("shopify_token")
        
        if not shopify_domain or not shopify_token:
            raise HTTPException(status_code=400, detail="Store missing Shopify credentials")
        
        import httpx
        
        # Fetch order from Shopify
        async with httpx.AsyncClient() as client:
            resp = await client.get(
                f"https://{shopify_domain}/admin/api/2024-01/orders/{shopify_order_id}.json",
                headers={"X-Shopify-Access-Token": shopify_token},
                timeout=30
            )
            
            if resp.status_code != 200:
                raise HTTPException(status_code=resp.status_code, detail=f"Shopify API error: {resp.text}")
            
            order = resp.json().get("order", {})
        
        if not order:
            raise HTTPException(status_code=404, detail="Order not found in Shopify")
        
        # Parse order data
        customer = order.get("customer", {})
        shipping = order.get("shipping_address", {}) or {}
        
        line_items = []
        order_skus = []
        sizes = []
        
        for item in order.get("line_items", []):
            # Skip removed items (fulfillable_quantity = 0 means item was removed/refunded)
            if item.get("fulfillable_quantity", 1) == 0 and item.get("fulfillment_status") is None:
                continue
                
            sku = item.get("sku", "") or ""
            line_items.append({
                "product_id": item.get("product_id"),
                "variant_id": item.get("variant_id"),
                "sku": sku,
                "name": item.get("name"),
                "quantity": item.get("quantity"),
                "price": float(item.get("price", 0))
            })
            if sku:
                order_skus.append(sku.upper())
            name = item.get("name", "")
            if "/" in name:
                parts = name.split("/")
                size = parts[-1].strip() if len(parts) > 1 else "Unknown"
                sizes.append(size)
        
        tracking_number = None
        tracking_company = "TCS Pakistan"
        fulfillments = order.get("fulfillments", [])
        if fulfillments:
            tracking_number = fulfillments[0].get("tracking_number")
            tracking_company = fulfillments[0].get("tracking_company") or "TCS Pakistan"
        
        order_doc = {
            "id": str(uuid.uuid4()),
            "customer_id": f"shopify_{customer.get('id', order['id'])}",
            "first_name": shipping.get("first_name") or customer.get("first_name") or "",
            "last_name": shipping.get("last_name") or customer.get("last_name") or "",
            "email": customer.get("email", ""),
            "phone": shipping.get("phone") or customer.get("phone") or "",
            "country_code": shipping.get("country_code", ""),
            "store_name": store_name,
            "order_skus": order_skus,
            "shoe_sizes": sizes if sizes else ["Unknown"],
            "line_items": line_items,
            "order_count": 1,
            "order_number": str(order.get("order_number", "")),
            "shopify_order_id": str(order.get("id")),
            "last_order_date": order.get("created_at"),
            "fulfilled_at": order.get("fulfilled_at"),
            "total_spent": float(order.get("total_price", 0)),
            "fulfillment_status": order.get("fulfillment_status") or "unfulfilled",
            "payment_status": order.get("financial_status"),
            "tracking_number": tracking_number,
            "tracking_company": tracking_company,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        # Upsert to database
        result = await db.customers.update_one(
            {"shopify_order_id": str(order.get("id")), "store_name": store_name},
            {"$set": order_doc},
            upsert=True
        )
        
        return {
            "success": True,
            "message": f"Order #{order.get('order_number')} synced successfully",
            "order_number": order.get("order_number"),
            "customer": f"{order_doc['first_name']} {order_doc['last_name']}",
            "upserted": result.upserted_id is not None,
            "modified": result.modified_count > 0
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error syncing single order: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/prepare-migration")
async def prepare_migration():
    """
    Prepare database for migration by dropping problematic indexes
    """
    try:
        dropped_indexes = []
        
        # Drop customer_id unique index that causes conflicts
        try:
            await db.customers.drop_index("customer_id_1")
            dropped_indexes.append("customers.customer_id_1")
            logger.info("Dropped customers.customer_id_1 index")
        except Exception as e:
            logger.info(f"Index customers.customer_id_1 may not exist: {e}")
        
        # Clear customers collection
        delete_result = await db.customers.delete_many({})
        
        return {
            "success": True,
            "message": "Database prepared for migration",
            "dropped_indexes": dropped_indexes,
            "deleted_customers": delete_result.deleted_count
        }
    except Exception as e:
        logger.error(f"Error preparing migration: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/init-admin")
async def initialize_admin():
    """
    One-time initialization endpoint to create default admin user
    Only works if no admin exists yet
    """
    try:
        import hashlib
        
        # Check if any admin already exists
        existing_admin = await db.agents.find_one({"username": "admin"})
        if existing_admin:
            return {
                "success": False,
                "message": "Admin user already exists. Use /api/reset-admin to reset password."
            }
        
        # Create default admin
        admin_password = "admin"
        hashed_password = hashlib.sha256(admin_password.encode()).hexdigest()
        
        admin_user = {
            "id": "admin-default-user",
            "username": "admin",
            "password": hashed_password,
            "full_name": "Administrator",
            "role": "admin",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.agents.insert_one(admin_user)
        
        logger.info("✅ Default admin user initialized")
        
        return {
            "success": True,
            "message": "Admin user created successfully",
            "credentials": {
                "username": "admin",
                "password": "admin"
            }
        }
    except Exception as e:
        logger.error(f"Error initializing admin: {str(e)}")
        raise HTTPException(status_code=500, detail="Admin initialization failed")


@api_router.post("/reset-admin")
async def reset_admin_password():
    """
    Reset admin password to default (admin/admin)
    Creates admin if doesn't exist, resets password if exists
    """
    try:
        import hashlib
        
        admin_password = "admin"
        hashed_password = hashlib.sha256(admin_password.encode()).hexdigest()
        
        # Check if admin exists
        existing_admin = await db.agents.find_one({"username": "admin"})
        
        if existing_admin:
            # Reset password
            await db.agents.update_one(
                {"username": "admin"},
                {"$set": {"password": hashed_password}}
            )
            logger.info("✅ Admin password reset to default")
            return {
                "success": True,
                "message": "Admin password reset successfully",
                "credentials": {
                    "username": "admin",
                    "password": "admin"
                }
            }
        else:
            # Create new admin
            admin_user = {
                "id": "admin-default-user",
                "username": "admin",
                "password": hashed_password,
                "full_name": "Administrator",
                "role": "admin",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.agents.insert_one(admin_user)
            logger.info("✅ Admin user created with default password")
            return {
                "success": True,
                "message": "Admin user created successfully",
                "credentials": {
                    "username": "admin",
                    "password": "admin"
                }
            }
    except Exception as e:
        logger.error(f"Error resetting admin: {str(e)}")
        raise HTTPException(status_code=500, detail="Admin reset failed")


@api_router.post("/agents/signup")
async def signup_agent(agent_data: AgentCreate):
    """
    Agent registration/signup with bcrypt encrypted password
    Your password is securely encrypted using industry-standard bcrypt hashing.
    """
    try:
        import bcrypt
        
        def hash_password(password: str) -> str:
            """Hash password using bcrypt"""
            salt = bcrypt.gensalt(rounds=12)
            return bcrypt.hashpw(password.encode('utf-8'), salt).decode('utf-8')
        
        # Check if username already exists
        existing_agent = await db.agents.find_one({"username": agent_data.username})
        if existing_agent:
            raise HTTPException(status_code=400, detail="Username already exists")
        
        # Hash password with bcrypt (secure encryption)
        hashed_password = hash_password(agent_data.password)
        
        # Create new agent
        new_agent = {
            "id": str(uuid.uuid4()),
            "username": agent_data.username,
            "password": hashed_password,
            "full_name": agent_data.full_name,
            "role": "agent",  # Default role - only WhatsApp CRM + Shopify access
            "created_at": datetime.now(timezone.utc).isoformat(),
            "password_encrypted": True  # Flag to indicate bcrypt encryption
        }
        
        await db.agents.insert_one(new_agent)
        
        logger.info(f"✅ New agent registered with encrypted password: {agent_data.username}")
        
        return {
            "success": True,
            "message": "Account created successfully. Your password is securely encrypted.",
            "agent": {
                "id": new_agent["id"],
                "username": new_agent["username"],
                "full_name": new_agent["full_name"],
                "role": new_agent["role"]
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error registering agent: {str(e)}")
        raise HTTPException(status_code=500, detail="Registration failed")


@api_router.get("/agents")
async def get_agents():
    """
    Get all agents
    """
    agents = await db.agents.find({}, {"_id": 0, "password": 0}).to_list(100)
    return agents


@api_router.post("/upload-stock")
async def upload_stock_file(file: UploadFile = File(...), store_name: str = "tnvcollectionpk"):
    """
    Upload stock Excel file and store SKUs for a specific store
    """
    try:
        # Read Excel file
        content = await file.read()
        
        logger.info(f"Received stock file: {file.filename}, size: {len(content)} bytes for store: {store_name}")
        
        # Parse Excel using pandas
        df = pd.read_excel(io.BytesIO(content))
        
        logger.info(f"Excel columns: {df.columns.tolist()}")
        
        # Extract SKUs (assuming column is named 'SKU')
        if 'SKU' not in df.columns:
            raise HTTPException(status_code=400, detail="Excel file must have a 'SKU' column")
        
        # Get all SKUs and filter out empty values
        skus = df['SKU'].dropna().astype(str).tolist()
        skus = [sku.strip() for sku in skus if sku.strip()]
        
        logger.info(f"Extracted {len(skus)} SKUs from Excel")
        
        # Clear existing stock for this store
        await db.stock.delete_many({"store_name": store_name})
        
        # Insert new stock items
        stock_items = []
        for sku in skus:
            stock_items.append({
                "id": str(uuid.uuid4()),
                "sku": sku.upper(),  # Normalize to uppercase for matching
                "store_name": store_name,
                "created_at": datetime.now(timezone.utc).isoformat()
            })
        
        if stock_items:
            await db.stock.insert_many(stock_items)
        
        # Get unique SKU count
        unique_skus = len(set(sku.upper() for sku in skus))
        
        return {
            "success": True,
            "message": f"Stock updated for {store_name}",
            "total_items": len(skus),
            "unique_skus": unique_skus,
            "store_name": store_name
        }
        
    except Exception as e:
        logger.error(f"Error uploading stock file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to process stock file: {str(e)}")


@api_router.get("/stock/{store_name}")
async def get_store_stock(store_name: str):
    """
    Get all stock SKUs for a specific store
    """
    stock_items = await db.stock.find({"store_name": store_name}, {"_id": 0}).to_list(10000)
    skus = [item["sku"] for item in stock_items]
    return {
        "store_name": store_name,
        "total_skus": len(skus),
        "unique_skus": len(set(skus)),
        "skus": list(set(skus))  # Return unique SKUs
    }


@api_router.post("/shopify/configure")
async def configure_shopify_store(
    store_name: str,
    shopify_domain: str,
    shopify_token: str
):
    """
    Configure Shopify credentials for a store
    """
    try:
        # Update store with Shopify credentials
        result = await db.stores.update_one(
            {"store_name": store_name},
            {"$set": {
                "shopify_domain": shopify_domain,
                "shopify_token": shopify_token,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Store not found")
        
        # Test connection
        sync = ShopifyOrderSync(shopify_domain, shopify_token)
        shop_info = sync.get_shop_info()
        
        if not shop_info:
            raise HTTPException(status_code=400, detail="Failed to connect to Shopify. Please check credentials.")
        
        return {
            "success": True,
            "message": f"Shopify configured for {store_name}",
            "shop_info": shop_info
        }
    except Exception as e:
        logger.error(f"Error configuring Shopify: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/shopify/sync/{store_name}")
async def sync_shopify_orders(store_name: str, days_back: int = 3650, full_sync: bool = False):
    """
    Manually sync orders from Shopify for a specific store
    
    Args:
        store_name: Store to sync
        days_back: Number of days to look back (default 3650 = ~10 years from 2015)
        full_sync: If True, syncs ALL orders with pagination (default False)
    """
    try:
        # Get store credentials
        store = await db.stores.find_one({"store_name": store_name}, {"_id": 0})
        
        if not store:
            raise HTTPException(status_code=404, detail="Store not found")
        
        if not store.get('shopify_domain') or not store.get('shopify_token'):
            raise HTTPException(status_code=400, detail="Shopify not configured for this store")
        
        # Fetch orders from Shopify
        sync = ShopifyOrderSync(store['shopify_domain'], store['shopify_token'])
        
        if full_sync:
            # Sync all orders (no date filter, with pagination)
            created_after = None
            logger.info(f"Starting FULL sync for {store_name} - fetching ALL orders with pagination")
            orders = sync.fetch_orders(limit=250, status="any", created_after=created_after, fetch_all=True)
            
            # Also fetch cancelled orders separately
            cancelled_orders = sync.fetch_orders(limit=250, status="cancelled", created_after=created_after, fetch_all=True)
            logger.info(f"Fetched {len(cancelled_orders)} cancelled orders separately")
            orders.extend(cancelled_orders)
        else:
            # Fetch orders from last X days
            created_after = (datetime.now(timezone.utc) - timedelta(days=days_back)).isoformat()
            logger.info(f"Syncing {store_name} orders from last {days_back} days")
            # Use fetch_all=True when syncing more than 30 days to get all historical orders
            fetch_all_orders = days_back > 30
            # Fetch all statuses including cancelled orders
            orders = sync.fetch_orders(limit=250, status="any", created_after=created_after, fetch_all=fetch_all_orders)
            
            # Also fetch cancelled orders separately (Shopify API quirk - "any" doesn't always include cancelled)
            cancelled_orders = sync.fetch_orders(limit=250, status="cancelled", created_after=created_after, fetch_all=fetch_all_orders)
            logger.info(f"Fetched {len(cancelled_orders)} cancelled orders separately")
            orders.extend(cancelled_orders)
        
        if not orders:
            return {
                "success": True,
                "message": "No new orders to sync",
                "orders_synced": 0
            }
        
        # Process and merge orders into customers
        customers_updated = 0
        customers_created = 0
        
        for order_data in orders:
            customer_id = order_data['customer_id']
            
            # Check if customer exists
            existing = await db.customers.find_one({"customer_id": customer_id, "store_name": store_name}, {"_id": 0})
            
            # Extract SKUs from line items (use current line_items only, not merged)
            order_skus = [item['sku'].upper() for item in order_data['line_items'] if item['sku']]
            
            # Extract sizes from line item names
            sizes = []
            for item in order_data['line_items']:
                # Try to extract size from product name
                name = item['name']
                if '/' in name:
                    parts = name.split('/')
                    size = parts[-1].strip() if len(parts) > 1 else 'Unknown'
                    sizes.append(size)
            
            if existing:
                # Update existing customer - REPLACE line_items and SKUs with current order data
                # Don't merge SKUs as order may have been edited/changed
                await db.customers.update_one(
                    {"customer_id": customer_id, "store_name": store_name},
                    {"$set": {
                        "first_name": order_data['first_name'] or existing.get('first_name'),
                        "last_name": order_data['last_name'] or existing.get('last_name'),
                        "email": order_data['email'] or existing.get('email'),
                        "phone": order_data['phone'] or existing.get('phone'),
                        "country_code": order_data['country_code'] or existing.get('country_code'),
                        "order_skus": order_skus,  # Replace, don't merge
                        "shoe_sizes": sizes if sizes else existing.get('shoe_sizes', ['Unknown']),  # Replace, don't merge
                        "line_items": order_data.get('line_items', []),
                        "order_number": str(order_data['order_number']),
                        "shopify_order_id": order_data.get('shopify_order_id'),
                        "last_order_date": order_data['order_date'],
                        "fulfillment_status": order_data['fulfillment_status'],
                        "fulfilled_at": order_data.get('fulfilled_at'),
                        "payment_status": order_data.get('payment_status'),
                        "payment_method": order_data.get('payment_method'),
                        "tracking_number": order_data['tracking_info']['tracking_number'] if order_data['tracking_info'] else None,
                        "tracking_company": order_data['tracking_info']['tracking_company'] if order_data['tracking_info'] else 'TCS Pakistan',
                        "tracking_url": order_data['tracking_info']['tracking_url'] if order_data['tracking_info'] else None,
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }}
                )
                customers_updated += 1
            else:
                # Create new customer
                new_customer = {
                    "id": str(uuid.uuid4()),
                    "customer_id": customer_id,
                    "first_name": order_data['first_name'],
                    "last_name": order_data['last_name'],
                    "email": order_data['email'],
                    "phone": order_data['phone'],
                    "country_code": order_data['country_code'],
                    "store_name": store_name,
                    "order_skus": order_skus,
                    "shoe_sizes": sizes if sizes else ['Unknown'],
                    "line_items": order_data.get('line_items', []),
                    "order_count": 1,
                    "order_number": str(order_data['order_number']),
                    "shopify_order_id": order_data.get('shopify_order_id'),
                    "last_order_date": order_data['order_date'],
                    "fulfilled_at": order_data.get('fulfilled_at'),
                    "total_spent": order_data['total_price'],
                    "fulfillment_status": order_data['fulfillment_status'],
                    "payment_status": order_data.get('payment_status'),
                    "payment_method": order_data.get('payment_method'),
                    "tracking_number": order_data['tracking_info']['tracking_number'] if order_data['tracking_info'] else None,
                    "tracking_company": order_data['tracking_info']['tracking_company'] if order_data['tracking_info'] else 'TCS Pakistan',
                    "tracking_url": order_data['tracking_info']['tracking_url'] if order_data['tracking_info'] else None,
                    "messaged": False,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.customers.insert_one(new_customer)
                customers_created += 1
            
            # Track order for dynamic pricing
            try:
                await pricing_engine.track_new_order(order_data)
            except Exception as pricing_error:
                logger.warning(f"Failed to track order for pricing: {pricing_error}")
        
        # Update last sync time
        await db.stores.update_one(
            {"store_name": store_name},
            {"$set": {"last_synced_at": datetime.now(timezone.utc).isoformat()}}
        )
        
        logger.info(f"✅ Pricing: Tracked {len(orders)} orders for dynamic pricing")
        
        return {
            "success": True,
            "message": f"Synced {len(orders)} orders from Shopify",
            "orders_synced": len(orders),
            "customers_created": customers_created,
            "customers_updated": customers_updated
        }
        
    except Exception as e:
        logger.error(f"Error syncing Shopify orders: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== NON-BLOCKING BACKGROUND SYNC ====================

@api_router.post("/shopify/sync-background/{store_name}")
async def start_shopify_sync_background(store_name: str, days_back: int = 30, background_tasks: BackgroundTasks = None):
    """
    Start a non-blocking background Shopify sync.
    Returns immediately with a job ID to track progress.
    """
    import uuid as uuid_module
    
    # Get store credentials
    store = await db.stores.find_one({"store_name": store_name}, {"_id": 0})
    
    if not store:
        raise HTTPException(status_code=404, detail="Store not found")
    
    if not store.get('shopify_domain') or not store.get('shopify_token'):
        raise HTTPException(status_code=400, detail="Shopify not configured for this store")
    
    # Create job
    job_id = str(uuid_module.uuid4())[:8]
    shopify_sync_jobs[job_id] = {
        "status": "started",
        "store_name": store_name,
        "progress": 0,
        "orders_fetched": 0,
        "orders_processed": 0,
        "errors": [],
        "started_at": datetime.now(timezone.utc).isoformat(),
    }
    
    # Start background task
    background_tasks.add_task(
        run_shopify_sync_background,
        job_id,
        store_name,
        store['shopify_domain'],
        store['shopify_token'],
        days_back
    )
    
    return {
        "success": True,
        "job_id": job_id,
        "message": f"Sync started for {store_name}",
        "status_url": f"/api/shopify/sync-status/{job_id}"
    }


@api_router.get("/shopify/sync-status/{job_id}")
async def get_shopify_sync_status(job_id: str):
    """Get the status of a background sync job"""
    if job_id not in shopify_sync_jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    return {
        "success": True,
        "job": shopify_sync_jobs[job_id]
    }


@api_router.post("/shopify/sync-orders/{store_name}")
async def sync_specific_orders(store_name: str, order_numbers: List[str] = Query(...)):
    """
    Sync specific orders by their order numbers.
    Useful for fetching missing orders.
    
    Usage: POST /api/shopify/sync-orders/tnvcollection?order_numbers=29160&order_numbers=29156
    """
    store = await db.stores.find_one({"store_name": store_name}, {"_id": 0})
    
    if not store:
        raise HTTPException(status_code=404, detail="Store not found")
    
    if not store.get('shopify_domain') or not store.get('shopify_token'):
        raise HTTPException(status_code=400, detail="Shopify not configured for this store")
    
    shopify_domain = store['shopify_domain']
    shopify_token = store['shopify_token']
    
    results = []
    
    for order_number in order_numbers:
        try:
            # Fetch order from Shopify by order number
            url = f"https://{shopify_domain}/admin/api/2024-01/orders.json"
            headers = {"X-Shopify-Access-Token": shopify_token}
            params = {"name": order_number, "status": "any"}
            
            async with httpx.AsyncClient(timeout=30.0) as client:
                response = await client.get(url, headers=headers, params=params)
                response.raise_for_status()
                data = response.json()
                
                orders = data.get("orders", [])
                
                if not orders:
                    results.append({"order": order_number, "status": "not_found"})
                    continue
                
                order = orders[0]
                
                # Process and save the order
                customer = order.get("customer", {})
                shipping = order.get("shipping_address", {}) or order.get("billing_address", {})
                
                customer_id = f"shopify_{customer.get('id', order['id'])}"
                
                line_items = []
                order_skus = []
                sizes = []
                
                for item in order.get("line_items", []):
                    # Skip removed items (fulfillable_quantity = 0 means item was removed/refunded)
                    if item.get("fulfillable_quantity", 1) == 0 and item.get("fulfillment_status") is None:
                        continue
                    
                    sku = item.get("sku", "") or ""
                    line_items.append({
                        "product_id": item.get("product_id"),
                        "variant_id": item.get("variant_id"),
                        "sku": sku,
                        "name": item.get("name"),
                        "quantity": item.get("quantity"),
                        "price": float(item.get("price", 0))
                    })
                    if sku:
                        order_skus.append(sku.upper())
                    
                    # Extract size from name
                    name = item.get("name", "")
                    if "/" in name:
                        parts = name.split("/")
                        size = parts[-1].strip() if len(parts) > 1 else "Unknown"
                        sizes.append(size)
                
                # Get tracking info
                tracking_number = None
                tracking_company = "TCS Pakistan"
                fulfillments = order.get("fulfillments", [])
                if fulfillments:
                    tracking_number = fulfillments[0].get("tracking_number")
                    tracking_company = fulfillments[0].get("tracking_company") or "TCS Pakistan"
                
                # Build customer/order document
                order_doc = {
                    "customer_id": customer_id,
                    "first_name": shipping.get("first_name") or customer.get("first_name"),
                    "last_name": shipping.get("last_name") or customer.get("last_name"),
                    "email": customer.get("email", ""),
                    "phone": shipping.get("phone") or customer.get("phone") or "",
                    "country_code": shipping.get("country_code", ""),
                    "store_name": store_name,
                    "order_skus": order_skus,
                    "shoe_sizes": sizes if sizes else ["Unknown"],
                    "line_items": line_items,
                    "order_count": 1,
                    "order_number": str(order.get("order_number", "")),
                    "shopify_order_id": str(order.get("id")),
                    "last_order_date": order.get("created_at"),
                    "fulfilled_at": order.get("fulfilled_at"),
                    "total_spent": float(order.get("total_price", 0)),
                    "fulfillment_status": order.get("fulfillment_status") or "unfulfilled",
                    "payment_status": order.get("financial_status"),
                    "tracking_number": tracking_number,
                    "tracking_company": tracking_company,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
                
                # Upsert to database
                existing = await db.customers.find_one(
                    {"customer_id": customer_id, "store_name": store_name}
                )
                
                if existing:
                    await db.customers.update_one(
                        {"customer_id": customer_id, "store_name": store_name},
                        {"$set": order_doc}
                    )
                    results.append({"order": order_number, "status": "updated"})
                else:
                    order_doc["id"] = str(uuid.uuid4())
                    order_doc["created_at"] = datetime.now(timezone.utc).isoformat()
                    order_doc["messaged"] = False
                    await db.customers.insert_one(order_doc)
                    results.append({"order": order_number, "status": "created"})
                    
        except Exception as e:
            logger.error(f"Error syncing order {order_number}: {e}")
            results.append({"order": order_number, "status": "error", "error": str(e)})
    
    return {
        "success": True,
        "results": results,
        "synced": len([r for r in results if r["status"] in ["created", "updated"]]),
        "failed": len([r for r in results if r["status"] in ["not_found", "error"]])
    }


async def run_shopify_sync_background(
    job_id: str,
    store_name: str,
    shopify_domain: str,
    shopify_token: str,
    days_back: int
):
    """Background task to sync Shopify orders without blocking.
    For large date ranges (>90 days), processes in monthly chunks to prevent timeouts.
    """
    try:
        shopify_sync_jobs[job_id]["status"] = "fetching"
        shopify_sync_jobs[job_id]["phase"] = "Initializing..."
        
        # Use async sync
        sync = ShopifyAsyncSync(shopify_domain, shopify_token, max_workers=5)
        
        all_orders = []
        
        # For large syncs (>90 days), process in monthly chunks
        if days_back > 90:
            logger.info(f"🚀 [BG Sync {job_id}] Large sync ({days_back} days) - processing in monthly chunks")
            shopify_sync_jobs[job_id]["phase"] = "Fetching in monthly chunks..."
            
            # Calculate number of chunks (30-day periods)
            chunk_size = 30  # days per chunk
            num_chunks = (days_back + chunk_size - 1) // chunk_size  # ceil division
            
            for chunk_idx in range(num_chunks):
                chunk_start_days = chunk_idx * chunk_size
                chunk_end_days = min((chunk_idx + 1) * chunk_size, days_back)
                
                # Calculate date range for this chunk
                end_date = datetime.now(timezone.utc) - timedelta(days=chunk_start_days)
                start_date = datetime.now(timezone.utc) - timedelta(days=chunk_end_days)
                
                created_after = start_date.isoformat()
                created_before = end_date.isoformat() if chunk_start_days > 0 else None
                
                shopify_sync_jobs[job_id]["phase"] = f"Fetching chunk {chunk_idx + 1}/{num_chunks} ({chunk_end_days - chunk_start_days} days)"
                shopify_sync_jobs[job_id]["progress"] = int((chunk_idx / num_chunks) * 30)  # First 30% is fetching
                
                logger.info(f"📦 [BG Sync {job_id}] Fetching chunk {chunk_idx + 1}/{num_chunks}: {start_date.date()} to {end_date.date()}")
                
                try:
                    chunk_orders = await sync.fetch_orders_concurrent(
                        created_after=created_after,
                        status="any",
                        max_batches=5  # Smaller batches per chunk
                    )
                    all_orders.extend(chunk_orders)
                    logger.info(f"📦 [BG Sync {job_id}] Chunk {chunk_idx + 1} fetched {len(chunk_orders)} orders")
                except Exception as chunk_error:
                    logger.warning(f"⚠️ [BG Sync {job_id}] Chunk {chunk_idx + 1} failed: {chunk_error}")
                    shopify_sync_jobs[job_id]["errors"].append(f"Chunk {chunk_idx + 1}: {str(chunk_error)}")
                
                # Small delay between chunks to avoid rate limiting
                await asyncio.sleep(1)
            
            # Deduplicate orders by shopify_order_id
            seen_ids = set()
            unique_orders = []
            for order in all_orders:
                order_id = order.get('shopify_order_id')
                if order_id and order_id not in seen_ids:
                    seen_ids.add(order_id)
                    unique_orders.append(order)
            
            all_orders = unique_orders
            logger.info(f"✅ [BG Sync {job_id}] Total unique orders after deduplication: {len(all_orders)}")
        else:
            # Regular sync for smaller date ranges
            created_after = (datetime.now(timezone.utc) - timedelta(days=days_back)).isoformat()
            logger.info(f"🚀 [BG Sync {job_id}] Starting for {store_name} (last {days_back} days)")
            shopify_sync_jobs[job_id]["phase"] = f"Fetching orders (last {days_back} days)..."
            
            all_orders = await sync.fetch_orders_concurrent(created_after=created_after, status="any", max_batches=10)
        
        orders = all_orders
        shopify_sync_jobs[job_id]["orders_fetched"] = len(orders)
        shopify_sync_jobs[job_id]["status"] = "processing"
        shopify_sync_jobs[job_id]["phase"] = f"Processing {len(orders)} orders..."
        shopify_sync_jobs[job_id]["progress"] = 30  # 30% done after fetching
        
        sync.close()
        
        if not orders:
            shopify_sync_jobs[job_id]["status"] = "completed"
            shopify_sync_jobs[job_id]["message"] = "No new orders to sync"
            return
        
        # Process orders in batches
        total = len(orders)
        customers_updated = 0
        customers_created = 0
        
        for i, order_data in enumerate(orders):
            try:
                customer_id = order_data['customer_id']
                existing = await db.customers.find_one(
                    {"customer_id": customer_id, "store_name": store_name}, 
                    {"_id": 0}
                )
                
                order_skus = [item['sku'].upper() for item in order_data['line_items'] if item.get('sku')]
                
                sizes = []
                for item in order_data['line_items']:
                    name = item.get('name', '')
                    if '/' in name:
                        parts = name.split('/')
                        size = parts[-1].strip() if len(parts) > 1 else 'Unknown'
                        sizes.append(size)
                
                if existing:
                    existing_skus = set(existing.get('order_skus', []))
                    merged_skus = list(existing_skus.union(set(order_skus)))
                    
                    existing_sizes = set(existing.get('shoe_sizes', []))
                    merged_sizes = list(existing_sizes.union(set(sizes)))
                    
                    await db.customers.update_one(
                        {"customer_id": customer_id, "store_name": store_name},
                        {"$set": {
                            "first_name": order_data.get('first_name') or existing.get('first_name'),
                            "last_name": order_data.get('last_name') or existing.get('last_name'),
                            "email": order_data.get('email') or existing.get('email'),
                            "phone": order_data.get('phone') or existing.get('phone'),
                            "country_code": order_data.get('country_code') or existing.get('country_code'),
                            "order_skus": merged_skus,
                            "shoe_sizes": merged_sizes,
                            "line_items": order_data.get('line_items', []),
                            "order_number": str(order_data.get('order_number', '')),
                            "shopify_order_id": order_data.get('shopify_order_id'),
                            "last_order_date": order_data.get('order_date'),
                            "fulfillment_status": order_data.get('fulfillment_status'),
                            "fulfilled_at": order_data.get('fulfilled_at'),
                            "payment_status": order_data.get('payment_status'),
                            "payment_method": order_data.get('payment_method'),
                            "tracking_number": order_data['tracking_info']['tracking_number'] if order_data.get('tracking_info') else None,
                            "tracking_company": order_data['tracking_info']['tracking_company'] if order_data.get('tracking_info') else 'TCS Pakistan',
                            "updated_at": datetime.now(timezone.utc).isoformat()
                        }}
                    )
                    customers_updated += 1
                else:
                    new_customer = {
                        "id": str(uuid.uuid4()),
                        "customer_id": customer_id,
                        "first_name": order_data.get('first_name'),
                        "last_name": order_data.get('last_name'),
                        "email": order_data.get('email'),
                        "phone": order_data.get('phone'),
                        "country_code": order_data.get('country_code'),
                        "store_name": store_name,
                        "order_skus": order_skus,
                        "shoe_sizes": sizes if sizes else ['Unknown'],
                        "line_items": order_data.get('line_items', []),
                        "order_count": 1,
                        "order_number": str(order_data.get('order_number', '')),
                        "shopify_order_id": order_data.get('shopify_order_id'),
                        "last_order_date": order_data.get('order_date'),
                        "fulfilled_at": order_data.get('fulfilled_at'),
                        "total_spent": order_data.get('total_price', 0),
                        "fulfillment_status": order_data.get('fulfillment_status'),
                        "payment_status": order_data.get('payment_status'),
                        "payment_method": order_data.get('payment_method'),
                        "tracking_number": order_data['tracking_info']['tracking_number'] if order_data.get('tracking_info') else None,
                        "tracking_company": order_data['tracking_info']['tracking_company'] if order_data.get('tracking_info') else 'TCS Pakistan',
                        "messaged": False,
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    await db.customers.insert_one(new_customer)
                    customers_created += 1
                
                # Update progress every 10 orders (processing is 30-100% of overall progress)
                if i % 10 == 0:
                    process_progress = int((i + 1) / total * 70)  # 70% for processing (30-100%)
                    shopify_sync_jobs[job_id]["progress"] = 30 + process_progress
                    shopify_sync_jobs[job_id]["orders_processed"] = i + 1
                    shopify_sync_jobs[job_id]["phase"] = f"Processing order {i + 1}/{total}..."
                    
            except Exception as e:
                shopify_sync_jobs[job_id]["errors"].append(f"Order {order_data.get('order_number')}: {str(e)}")
        
        # Update last sync time
        await db.stores.update_one(
            {"store_name": store_name},
            {"$set": {"last_synced_at": datetime.now(timezone.utc).isoformat()}}
        )
        
        shopify_sync_jobs[job_id]["status"] = "completed"
        shopify_sync_jobs[job_id]["progress"] = 100
        shopify_sync_jobs[job_id]["orders_processed"] = total
        shopify_sync_jobs[job_id]["customers_created"] = customers_created
        shopify_sync_jobs[job_id]["customers_updated"] = customers_updated
        shopify_sync_jobs[job_id]["completed_at"] = datetime.now(timezone.utc).isoformat()
        
        logger.info(f"✅ [BG Sync {job_id}] Completed: {total} orders, {customers_created} new, {customers_updated} updated")
        
    except Exception as e:
        shopify_sync_jobs[job_id]["status"] = "failed"
        shopify_sync_jobs[job_id]["error"] = str(e)
        logger.error(f"❌ [BG Sync {job_id}] Failed: {e}")


@api_router.post("/shopify/sync-fast/{store_name}")
async def sync_shopify_orders_fast(store_name: str, days_back: int = 3650):
    """
    PHASE 1: Fast concurrent sync with incremental updates
    10x faster than regular sync - perfect for daily updates
    
    Args:
        store_name: Store to sync
        days_back: Number of days to look back (default 3650 = ~10 years from 2015)
    """
    try:
        # Get store credentials
        store = await db.stores.find_one({"store_name": store_name}, {"_id": 0})
        
        if not store:
            raise HTTPException(status_code=404, detail="Store not found")
        
        if not store.get('shopify_domain') or not store.get('shopify_token'):
            raise HTTPException(status_code=400, detail="Shopify not configured for this store")
        
        # Use async concurrent sync
        sync = ShopifyAsyncSync(store['shopify_domain'], store['shopify_token'], max_workers=10)
        
        # Incremental sync - only fetch recent orders
        created_after = (datetime.now(timezone.utc) - timedelta(days=days_back)).isoformat()
        logger.info(f"🚀 Starting FAST sync for {store_name} (last {days_back} days, concurrent)")
        
        # Fetch orders concurrently (including cancelled)
        orders = await sync.fetch_orders_concurrent(created_after=created_after, status="any", max_batches=20)
        
        # Also fetch cancelled orders separately
        cancelled_orders = await sync.fetch_orders_concurrent(created_after=created_after, status="cancelled", max_batches=20)
        logger.info(f"Fast sync: Fetched {len(cancelled_orders)} cancelled orders separately")
        orders.extend(cancelled_orders)
        
        sync.close()
        
        if not orders:
            return {
                "success": True,
                "message": "No new orders to sync",
                "orders_synced": 0,
                "mode": "fast_concurrent"
            }
        
        # Process orders (same logic as regular sync)
        customers_updated = 0
        customers_created = 0
        
        for order_data in orders:
            customer_id = order_data['customer_id']
            
            existing = await db.customers.find_one({"customer_id": customer_id, "store_name": store_name}, {"_id": 0})
            
            order_skus = [item['sku'].upper() for item in order_data['line_items'] if item['sku']]
            
            sizes = []
            for item in order_data['line_items']:
                name = item['name']
                if '/' in name:
                    parts = name.split('/')
                    size = parts[-1].strip() if len(parts) > 1 else 'Unknown'
                    sizes.append(size)
            
            if existing:
                # Update existing customer
                existing_skus = set(existing.get('order_skus', []))
                merged_skus = list(existing_skus.union(set(order_skus)))
                
                existing_sizes = set(existing.get('shoe_sizes', []))
                merged_sizes = list(existing_sizes.union(set(sizes)))
                
                await db.customers.update_one(
                    {"customer_id": customer_id, "store_name": store_name},
                    {"$set": {
                        "first_name": order_data['first_name'] or existing.get('first_name'),
                        "last_name": order_data['last_name'] or existing.get('last_name'),
                        "email": order_data['email'] or existing.get('email'),
                        "phone": order_data['phone'] or existing.get('phone'),
                        "country_code": order_data['country_code'] or existing.get('country_code'),
                        "order_skus": merged_skus,
                        "shoe_sizes": merged_sizes,
                        "line_items": order_data.get('line_items', []),
                        "order_number": str(order_data['order_number']),
                        "shopify_order_id": order_data.get('shopify_order_id'),
                        "last_order_date": order_data['order_date'],
                        "fulfillment_status": order_data['fulfillment_status'],
                        "fulfilled_at": order_data.get('fulfilled_at'),
                        "payment_status": order_data.get('payment_status'),
                        "payment_method": order_data.get('payment_method'),
                        "tracking_number": order_data['tracking_info']['tracking_number'] if order_data['tracking_info'] else None,
                        "tracking_company": order_data['tracking_info']['tracking_company'] if order_data['tracking_info'] else 'TCS Pakistan',
                        "tracking_url": order_data['tracking_info']['tracking_url'] if order_data['tracking_info'] else None,
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }}
                )
                customers_updated += 1
            else:
                # Create new customer
                new_customer = {
                    "id": str(uuid.uuid4()),
                    "customer_id": customer_id,
                    "first_name": order_data['first_name'],
                    "last_name": order_data['last_name'],
                    "email": order_data['email'],
                    "phone": order_data['phone'],
                    "country_code": order_data['country_code'],
                    "store_name": store_name,
                    "order_skus": order_skus,
                    "shoe_sizes": sizes if sizes else ['Unknown'],
                    "line_items": order_data.get('line_items', []),
                    "order_count": 1,
                    "order_number": str(order_data['order_number']),
                    "shopify_order_id": order_data.get('shopify_order_id'),
                    "last_order_date": order_data['order_date'],
                    "fulfilled_at": order_data.get('fulfilled_at'),
                    "total_spent": order_data['total_price'],
                    "fulfillment_status": order_data['fulfillment_status'],
                    "payment_status": order_data.get('payment_status'),
                    "payment_method": order_data.get('payment_method'),
                    "tracking_number": order_data['tracking_info']['tracking_number'] if order_data['tracking_info'] else None,
                    "tracking_company": order_data['tracking_info']['tracking_company'] if order_data['tracking_info'] else 'TCS Pakistan',
                    "tracking_url": order_data['tracking_info']['tracking_url'] if order_data['tracking_info'] else None,
                    "messaged": False,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.customers.insert_one(new_customer)
                customers_created += 1
        
        # Update last sync time
        await db.stores.update_one(
            {"store_name": store_name},
            {"$set": {"last_synced_at": datetime.now(timezone.utc).isoformat()}}
        )
        
        return {
            "success": True,
            "message": f"Fast sync completed: {len(orders)} orders",
            "orders_synced": len(orders),
            "customers_created": customers_created,
            "customers_updated": customers_updated,
            "mode": "fast_concurrent",
            "days_synced": days_back
        }
        
    except Exception as e:
        logger.error(f"Error in fast sync: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/shopify/sync-abandoned-checkouts/{store_name}")
async def sync_abandoned_checkouts(store_name: str, days_back: int = 30):
    """
    P0: Sync abandoned checkouts from Shopify
    
    Args:
        store_name: Store to sync
        days_back: Number of days to look back (default 30)
    """
    try:
        # Get store credentials
        store = await db.stores.find_one({"store_name": store_name}, {"_id": 0})
        
        if not store:
            raise HTTPException(status_code=404, detail="Store not found")
        
        if not store.get('shopify_domain') or not store.get('shopify_token'):
            raise HTTPException(status_code=400, detail="Shopify not configured for this store")
        
        # Fetch abandoned checkouts
        sync = ShopifyAsyncSync(store['shopify_domain'], store['shopify_token'])
        created_after = (datetime.now(timezone.utc) - timedelta(days=days_back)).isoformat()
        
        logger.info(f"🛒 Syncing abandoned checkouts for {store_name} (last {days_back} days)")
        
        checkouts = await sync.fetch_abandoned_checkouts(created_after=created_after)
        sync.close()
        
        if not checkouts:
            return {
                "success": True,
                "message": "No abandoned checkouts found",
                "checkouts_synced": 0
            }
        
        # Process checkouts
        checkouts_saved = 0
        checkouts_updated = 0
        
        for checkout_data in checkouts:
            customer_id = checkout_data['customer_id']
            
            # Check if customer exists
            existing = await db.customers.find_one({"customer_id": customer_id, "store_name": store_name})
            
            if existing:
                # Update existing customer with abandoned checkout info
                await db.customers.update_one(
                    {"customer_id": customer_id, "store_name": store_name},
                    {"$set": {
                        "abandoned_checkout": True,
                        "abandoned_checkout_value": checkout_data.get('total_price', 0.0),
                        "abandoned_checkout_url": checkout_data.get('abandoned_checkout_url'),
                        "abandoned_at": checkout_data.get('abandoned_at'),
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }}
                )
                checkouts_updated += 1
            else:
                # Create new customer from abandoned checkout
                new_customer = {
                    "id": str(uuid.uuid4()),
                    "customer_id": customer_id,
                    "first_name": checkout_data.get('first_name', ''),
                    "last_name": checkout_data.get('last_name', ''),
                    "email": checkout_data.get('email', ''),
                    "phone": checkout_data.get('phone', ''),
                    "store_name": store_name,
                    "order_count": 0,
                    "shoe_sizes": [],
                    "order_skus": [item['sku'] for item in checkout_data.get('line_items', []) if item.get('sku')],
                    "abandoned_checkout": True,
                    "abandoned_checkout_value": checkout_data.get('total_price', 0.0),
                    "abandoned_checkout_url": checkout_data.get('abandoned_checkout_url'),
                    "abandoned_at": checkout_data.get('abandoned_at'),
                    "messaged": False,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.customers.insert_one(new_customer)
                checkouts_saved += 1
        
        return {
            "success": True,
            "message": f"Synced {len(checkouts)} abandoned checkouts",
            "checkouts_synced": len(checkouts),
            "new_customers": checkouts_saved,
            "existing_updated": checkouts_updated
        }
        
    except Exception as e:
        logger.error(f"Error syncing abandoned checkouts: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ========================================
# SHOPIFY PRODUCTS SYNC
# ========================================

@api_router.post("/shopify/sync-products/{store_name}")
async def sync_shopify_products(store_name: str, background_tasks: BackgroundTasks):
    """
    Sync ALL products from a Shopify store
    This runs in the background due to potentially large number of products
    """
    try:
        # Get store credentials
        store = await db.stores.find_one({"store_name": store_name}, {"_id": 0})
        
        if not store:
            raise HTTPException(status_code=404, detail="Store not found")
        
        if not store.get('shopify_domain') or not store.get('shopify_token'):
            raise HTTPException(status_code=400, detail="Shopify not configured for this store")
        
        # Start background sync
        background_tasks.add_task(
            _sync_products_background,
            store_name,
            store['shopify_domain'],
            store['shopify_token']
        )
        
        return {
            "success": True,
            "message": f"Product sync started for {store_name}. This may take a few minutes.",
            "status": "processing"
        }
        
    except Exception as e:
        logger.error(f"Error starting product sync: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


async def _sync_products_background(store_name: str, domain: str, token: str):
    """Background task to sync products"""
    try:
        logger.info(f"🛍️ Starting product sync for {store_name}...")
        
        sync = ShopifyOrderSync(domain, token)
        products = sync.fetch_products(fetch_all=True)
        
        if not products:
            logger.info(f"No products found for {store_name}")
            await db.product_sync_status.update_one(
                {"store_name": store_name},
                {"$set": {
                    "store_name": store_name,
                    "status": "completed",
                    "products_synced": 0,
                    "last_sync": datetime.now(timezone.utc).isoformat(),
                    "error": None
                }},
                upsert=True
            )
            return
        
        # Save products to database
        synced = 0
        for product in products:
            product['store_name'] = store_name
            product['synced_at'] = datetime.now(timezone.utc).isoformat()
            
            await db.shopify_products.update_one(
                {"shopify_product_id": product['shopify_product_id'], "store_name": store_name},
                {"$set": product},
                upsert=True
            )
            synced += 1
        
        # Update sync status
        await db.product_sync_status.update_one(
            {"store_name": store_name},
            {"$set": {
                "store_name": store_name,
                "status": "completed",
                "products_synced": synced,
                "last_sync": datetime.now(timezone.utc).isoformat(),
                "error": None
            }},
            upsert=True
        )
        
        logger.info(f"✅ Product sync completed for {store_name}: {synced} products")
        
    except Exception as e:
        logger.error(f"❌ Product sync failed for {store_name}: {str(e)}")
        await db.product_sync_status.update_one(
            {"store_name": store_name},
            {"$set": {
                "store_name": store_name,
                "status": "error",
                "error": str(e),
                "last_sync": datetime.now(timezone.utc).isoformat()
            }},
            upsert=True
        )


@api_router.post("/shopify/sync-products-all")
async def sync_all_store_products(background_tasks: BackgroundTasks):
    """
    Sync products from ALL configured Shopify stores
    """
    try:
        stores = await db.stores.find(
            {"shopify_domain": {"$ne": None, "$exists": True}},
            {"_id": 0, "store_name": 1, "shopify_domain": 1, "shopify_token": 1}
        ).to_list(20)
        
        if not stores:
            return {"success": False, "message": "No Shopify stores configured"}
        
        started = []
        for store in stores:
            if store.get('shopify_domain') and store.get('shopify_token'):
                background_tasks.add_task(
                    _sync_products_background,
                    store['store_name'],
                    store['shopify_domain'],
                    store['shopify_token']
                )
                started.append(store['store_name'])
        
        return {
            "success": True,
            "message": f"Product sync started for {len(started)} stores",
            "stores": started
        }
        
    except Exception as e:
        logger.error(f"Error starting all-store product sync: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/shopify/products")
async def get_shopify_products(
    store_name: str = None,
    search: str = None,
    link_status: str = None,
    page: int = 1,
    page_size: int = 50
):
    """Get synced Shopify products from database"""
    try:
        query = {}
        if store_name and store_name != 'all':
            query["store_name"] = store_name
        if search:
            query["$or"] = [
                {"title": {"$regex": search, "$options": "i"}},
                {"handle": {"$regex": search, "$options": "i"}},
                {"vendor": {"$regex": search, "$options": "i"}},
            ]
        if link_status == 'linked':
            query["linked_1688_product_id"] = {"$exists": True, "$ne": None}
        elif link_status == 'unlinked':
            query["$or"] = [
                {"linked_1688_product_id": {"$exists": False}},
                {"linked_1688_product_id": None}
            ]
        
        skip = (page - 1) * page_size
        
        products = await db.shopify_products.find(
            query,
            {"_id": 0}
        ).sort("updated_at", -1).skip(skip).limit(page_size).to_list(page_size)
        
        total = await db.shopify_products.count_documents(query)
        
        return {
            "success": True,
            "products": products,
            "total": total,
            "page": page,
            "page_size": page_size
        }
        
    except Exception as e:
        logger.error(f"Error fetching products: {str(e)}")
        return {"success": True, "products": [], "total": 0, "page": page, "page_size": page_size}


@api_router.get("/shopify/products/sync-status")
async def get_product_sync_status(store_name: str = None):
    """Get product sync status for stores"""
    try:
        query = {}
        if store_name:
            query["store_name"] = store_name
        
        statuses = await db.product_sync_status.find(query, {"_id": 0}).to_list(20)
        
        # Get product counts per store
        stores_with_counts = []
        for status in statuses:
            count = await db.shopify_products.count_documents({"store_name": status["store_name"]})
            status["product_count"] = count
            stores_with_counts.append(status)
        
        # Total products
        total_products = await db.shopify_products.count_documents({})
        
        return {
            "success": True,
            "statuses": stores_with_counts,
            "total_products": total_products
        }
        
    except Exception as e:
        logger.error(f"Error fetching sync status: {str(e)}")
        return {"success": True, "statuses": [], "total_products": 0}


@api_router.get("/shopify/products/{product_id}")
async def get_single_shopify_product(product_id: str, store_name: str = None):
    """Get a single Shopify product by ID - used by storefront"""
    try:
        query = {}
        if store_name:
            query["store_name"] = store_name
        
        # Try to find by shopify_product_id (main field)
        query_with_id = {**query, "shopify_product_id": product_id}
        product = await db.shopify_products.find_one(query_with_id, {"_id": 0})
        
        # If not found, try converting to int
        if not product:
            try:
                query_with_id = {**query, "shopify_product_id": int(product_id)}
                product = await db.shopify_products.find_one(query_with_id, {"_id": 0})
            except ValueError:
                pass
        
        # Also try with 'id' field for backwards compatibility
        if not product:
            query_with_id = {**query, "id": product_id}
            product = await db.shopify_products.find_one(query_with_id, {"_id": 0})
            if not product:
                try:
                    query_with_id = {**query, "id": int(product_id)}
                    product = await db.shopify_products.find_one(query_with_id, {"_id": 0})
                except ValueError:
                    pass
        
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        
        # Add 'id' field for frontend compatibility if missing
        if 'id' not in product and 'shopify_product_id' in product:
            product['id'] = product['shopify_product_id']
        
        return {
            "success": True,
            "product": product
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching product {product_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/shopify/products/{product_id}/link-1688")
async def link_product_to_1688(product_id: str, data: dict = Body(...)):
    """
    Link a Shopify product to a 1688 product
    """
    try:
        alibaba_product_id = data.get("alibaba_product_id")
        store_name = data.get("store_name")
        
        if not alibaba_product_id:
            raise HTTPException(status_code=400, detail="alibaba_product_id is required")
        
        # Clean the product ID (remove any non-numeric characters)
        alibaba_product_id = ''.join(filter(str.isdigit, str(alibaba_product_id)))
        
        if not alibaba_product_id:
            raise HTTPException(status_code=400, detail="Invalid 1688 product ID")
        
        # Update the Shopify product with the 1688 link
        query = {"shopify_product_id": product_id}
        if store_name:
            query["store_name"] = store_name
        
        result = await db.shopify_products.update_one(
            query,
            {"$set": {
                "linked_1688_product_id": alibaba_product_id,
                "linked_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Product not found")
        
        # Also add to the 1688 product catalog if not exists
        existing_catalog = await db.product_catalog_1688.find_one(
            {"source_id": alibaba_product_id}
        )
        
        if not existing_catalog:
            # Get the Shopify product info
            shopify_product = await db.shopify_products.find_one(query, {"_id": 0})
            if shopify_product:
                await db.product_catalog_1688.insert_one({
                    "source_id": alibaba_product_id,
                    "title": shopify_product.get("title"),
                    "shopify_product_id": product_id,
                    "store_name": store_name,
                    "linked_from_shopify": True,
                    "created_at": datetime.now(timezone.utc).isoformat()
                })
        
        return {
            "success": True,
            "message": f"Product linked to 1688 ID: {alibaba_product_id}",
            "alibaba_product_id": alibaba_product_id
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error linking product: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/shopify/products/{product_id}/unlink-1688")
async def unlink_product_from_1688(product_id: str, data: dict = Body(...)):
    """
    Remove 1688 link from a Shopify product
    """
    try:
        store_name = data.get("store_name")
        
        query = {"shopify_product_id": product_id}
        if store_name:
            query["store_name"] = store_name
        
        result = await db.shopify_products.update_one(
            query,
            {"$unset": {
                "linked_1688_product_id": "",
                "linked_at": ""
            }}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Product not found")
        
        return {
            "success": True,
            "message": "1688 link removed"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error unlinking product: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ========================================
# BULK AUTO-LINK BY IMAGE SEARCH
# ========================================

# Track bulk link jobs in memory
bulk_link_jobs = {}

@api_router.post("/shopify/products/bulk-auto-link")
async def bulk_auto_link_products(
    store_name: str = Body(...),
    limit: int = Body(100, description="Max products to process"),
    background_tasks: BackgroundTasks = None
):
    """
    Bulk auto-link Shopify products to 1688 using image search.
    Processes unlinked products from the specified store.
    """
    import uuid as uuid_module
    
    # Create job ID
    job_id = str(uuid_module.uuid4())[:8]
    
    bulk_link_jobs[job_id] = {
        "status": "started",
        "store_name": store_name,
        "total": 0,
        "processed": 0,
        "linked": 0,
        "failed": 0,
        "skipped": 0,
        "results": [],
        "started_at": datetime.now(timezone.utc).isoformat(),
    }
    
    # Start background task
    background_tasks.add_task(
        run_bulk_auto_link,
        job_id,
        store_name,
        limit
    )
    
    return {
        "success": True,
        "job_id": job_id,
        "message": f"Bulk auto-link started for {store_name}",
        "status_url": f"/api/shopify/products/bulk-auto-link/status/{job_id}"
    }


@api_router.get("/shopify/products/bulk-auto-link/status/{job_id}")
async def get_bulk_link_status(job_id: str):
    """Get status of bulk auto-link job"""
    if job_id not in bulk_link_jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    return {
        "success": True,
        "job": bulk_link_jobs[job_id]
    }


async def run_bulk_auto_link(job_id: str, store_name: str, limit: int):
    """Background task to auto-link products by image search"""
    from services.image_search_service import search_products_by_image
    import asyncio
    
    try:
        bulk_link_jobs[job_id]["status"] = "fetching_products"
        
        # Get unlinked products from the store
        query = {
            "store_name": store_name,
            "$or": [
                {"linked_1688_product_id": {"$exists": False}},
                {"linked_1688_product_id": None},
                {"linked_1688_product_id": ""}
            ]
        }
        
        # Filter to only products with images
        products = await db.shopify_products.find(
            query,
            {"_id": 0, "shopify_product_id": 1, "title": 1, "image_url": 1, "images": 1, "handle": 1}
        ).limit(limit).to_list(limit)
        
        bulk_link_jobs[job_id]["total"] = len(products)
        bulk_link_jobs[job_id]["status"] = "processing"
        
        logger.info(f"[Bulk Link {job_id}] Found {len(products)} unlinked products for {store_name}")
        
        for i, product in enumerate(products):
            try:
                shopify_id = product.get("shopify_product_id")
                title = product.get("title", "Unknown")
                
                # Get image URL
                image_url = product.get("image_url")
                if not image_url and product.get("images"):
                    images = product.get("images", [])
                    if images and len(images) > 0:
                        if isinstance(images[0], dict):
                            image_url = images[0].get("src")
                        else:
                            image_url = images[0]
                
                bulk_link_jobs[job_id]["processed"] = i + 1
                bulk_link_jobs[job_id]["current_product"] = title[:50]
                
                if not image_url:
                    bulk_link_jobs[job_id]["skipped"] += 1
                    bulk_link_jobs[job_id]["results"].append({
                        "shopify_id": shopify_id,
                        "title": title[:50],
                        "status": "skipped",
                        "reason": "No image"
                    })
                    continue
                
                # Search for similar products on 1688
                search_result = await search_products_by_image(image_url, limit=5)
                
                if not search_result.get("success") or not search_result.get("products"):
                    bulk_link_jobs[job_id]["failed"] += 1
                    bulk_link_jobs[job_id]["results"].append({
                        "shopify_id": shopify_id,
                        "title": title[:50],
                        "status": "failed",
                        "reason": search_result.get("error", "No matches found")
                    })
                    # Small delay to avoid rate limiting
                    await asyncio.sleep(1)
                    continue
                
                # Get the best match
                best_match = search_result["products"][0]
                alibaba_product_id = best_match.get("product_id")
                
                if not alibaba_product_id:
                    bulk_link_jobs[job_id]["failed"] += 1
                    continue
                
                # Link the product
                await db.shopify_products.update_one(
                    {"shopify_product_id": shopify_id, "store_name": store_name},
                    {"$set": {
                        "linked_1688_product_id": alibaba_product_id,
                        "linked_1688_title": best_match.get("title", ""),
                        "linked_1688_price": best_match.get("price", ""),
                        "linked_1688_image": best_match.get("image", ""),
                        "linked_1688_url": best_match.get("url", ""),
                        "linked_at": datetime.now(timezone.utc).isoformat(),
                        "linked_by": "auto_image_search"
                    }}
                )
                
                bulk_link_jobs[job_id]["linked"] += 1
                bulk_link_jobs[job_id]["results"].append({
                    "shopify_id": shopify_id,
                    "title": title[:50],
                    "status": "linked",
                    "alibaba_id": alibaba_product_id,
                    "alibaba_title": best_match.get("title", "")[:50]
                })
                
                logger.info(f"[Bulk Link {job_id}] Linked {title[:30]}... -> {alibaba_product_id}")
                
                # Rate limiting delay (TMAPI image search costs credits)
                await asyncio.sleep(2)
                
            except Exception as e:
                logger.error(f"[Bulk Link {job_id}] Error processing product: {e}")
                bulk_link_jobs[job_id]["failed"] += 1
                bulk_link_jobs[job_id]["results"].append({
                    "shopify_id": product.get("shopify_product_id"),
                    "title": product.get("title", "")[:50],
                    "status": "error",
                    "reason": str(e)
                })
        
        bulk_link_jobs[job_id]["status"] = "completed"
        bulk_link_jobs[job_id]["completed_at"] = datetime.now(timezone.utc).isoformat()
        
        logger.info(f"[Bulk Link {job_id}] Completed: {bulk_link_jobs[job_id]['linked']} linked, {bulk_link_jobs[job_id]['failed']} failed, {bulk_link_jobs[job_id]['skipped']} skipped")
        
    except Exception as e:
        logger.error(f"[Bulk Link {job_id}] Fatal error: {e}")
        bulk_link_jobs[job_id]["status"] = "error"
        bulk_link_jobs[job_id]["error"] = str(e)


# ========================================
# DRAFT ORDERS ENDPOINTS
# ========================================

@api_router.get("/shopify/draft-orders")
async def get_draft_orders(store_name: str = None):
    """Get draft orders from database"""
    try:
        query = {}
        if store_name and store_name != 'all':
            query["store_name"] = store_name
        
        drafts = await db.draft_orders.find(query, {"_id": 0}).sort("created_at", -1).limit(100).to_list(100)
        
        total_value = sum(float(d.get('total_price', 0) or 0) for d in drafts)
        
        return {
            "success": True,
            "drafts": drafts,
            "total": len(drafts),
            "total_value": total_value
        }
    except Exception as e:
        logger.error(f"Error fetching draft orders: {str(e)}")
        return {"success": True, "drafts": [], "total": 0, "total_value": 0}


@api_router.post("/shopify/sync-drafts/{store_name}")
async def sync_draft_orders(store_name: str):
    """
    Sync draft orders from Shopify
    """
    try:
        # Get store credentials
        store = await db.stores.find_one({"store_name": store_name}, {"_id": 0})
        
        if not store:
            raise HTTPException(status_code=404, detail="Store not found")
        
        if not store.get('shopify_domain') or not store.get('shopify_token'):
            raise HTTPException(status_code=400, detail="Shopify not configured for this store")
        
        domain = store['shopify_domain']
        token = store['shopify_token']
        
        # Fetch draft orders from Shopify API
        import httpx
        
        headers = {
            "X-Shopify-Access-Token": token,
            "Content-Type": "application/json"
        }
        
        url = f"https://{domain}/admin/api/2024-01/draft_orders.json?status=open&limit=250"
        
        async with httpx.AsyncClient() as client:
            response = await client.get(url, headers=headers, timeout=30.0)
            
            if response.status_code != 200:
                logger.error(f"Shopify API error: {response.status_code} - {response.text}")
                return {
                    "success": False,
                    "message": f"Shopify API error: {response.status_code}",
                    "drafts_synced": 0
                }
            
            data = response.json()
            draft_orders = data.get('draft_orders', [])
        
        if not draft_orders:
            return {
                "success": True,
                "message": "No draft orders found",
                "drafts_synced": 0
            }
        
        # Save draft orders to database
        drafts_saved = 0
        for draft in draft_orders:
            draft_doc = {
                "id": str(draft['id']),
                "name": draft.get('name', f"D{draft['id']}"),
                "store_name": store_name,
                "customer": draft.get('customer', {}),
                "email": draft.get('email'),
                "line_items": draft.get('line_items', []),
                "subtotal_price": draft.get('subtotal_price'),
                "total_price": draft.get('total_price'),
                "total_tax": draft.get('total_tax'),
                "currency": draft.get('currency', 'INR'),
                "status": draft.get('status', 'open'),
                "invoice_url": draft.get('invoice_url'),
                "created_at": draft.get('created_at'),
                "updated_at": draft.get('updated_at'),
                "synced_at": datetime.now(timezone.utc).isoformat()
            }
            
            await db.draft_orders.update_one(
                {"id": draft_doc["id"], "store_name": store_name},
                {"$set": draft_doc},
                upsert=True
            )
            drafts_saved += 1
        
        logger.info(f"✅ Synced {drafts_saved} draft orders for {store_name}")
        
        return {
            "success": True,
            "message": f"Synced {drafts_saved} draft orders",
            "drafts_synced": drafts_saved
        }
        
    except Exception as e:
        logger.error(f"Error syncing draft orders: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))



# ========================================
# DTDC CONFIGURATION ENDPOINTS
# ========================================

class DTDCConfigRequest(BaseModel):
    username: Optional[str] = None
    password: Optional[str] = None
    api_url: Optional[str] = "https://customer.dtdc.in/api"

@api_router.get("/dtdc/credentials")
async def get_dtdc_credentials():
    """Get DTDC API configuration status"""
    try:
        config = await db.dtdc_config.find_one({"service": "dtdc_india"}, {"_id": 0})
        
        if config:
            return {
                "configured": True,
                "username": config.get("username", "")[:5] + "..." if config.get("username") else None,
                "api_url": config.get("api_url", "https://customer.dtdc.in/api")
            }
        else:
            return {"configured": False}
    except Exception as e:
        logger.error(f"Error fetching DTDC credentials: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/dtdc/configure")
async def configure_dtdc_api(config_data: dict):
    """Configure DTDC API credentials for tnvcollection and ashmiaa stores"""
    try:
        username = config_data.get("username", "").strip()
        password = config_data.get("password", "").strip()
        api_url = config_data.get("api_url", "https://customer.dtdc.in/api").strip()
        
        if not username or not password:
            raise HTTPException(status_code=400, detail="Username and password are required")
        
        dtdc_config = {
            "service": "dtdc_india",
            "username": username,
            "password": password,
            "api_url": api_url,
            "configured_at": datetime.now(timezone.utc).isoformat()
        }
        
        # Save to database (upsert)
        await db.dtdc_config.update_one(
            {"service": "dtdc_india"},
            {"$set": dtdc_config},
            upsert=True
        )
        
        logger.info(f"✅ DTDC API configured successfully")
        
        return {
            "success": True,
            "message": "DTDC API configured successfully",
            "api_url": api_url
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error configuring DTDC API: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))



@api_router.post("/dtdc/sync-all-tracking")
async def sync_all_dtdc_tracking():
    """
    Sync DTDC tracking data from Shopify fulfillments to local database
    """
    try:
        from shopify_tracking_sync import shopify_tracking_sync
        
        logger.info("Starting DTDC tracking sync from Shopify...")
        
        # Sync tracking data from Shopify fulfillments
        result = await shopify_tracking_sync.sync_tracking_for_orders(db)
        
        if result.get('success'):
            logger.info(f"✅ Successfully synced tracking for {result.get('updated', 0)}/{result.get('total', 0)} orders")
            return {
                "success": True,
                "message": f"Successfully synced tracking data for {result.get('updated', 0)} orders",
                "updated": result.get('updated', 0),
                "total": result.get('total', 0),
                "errors": result.get('errors')
            }
        else:
            raise HTTPException(status_code=500, detail=result.get('error', 'Unknown error'))
            
    except Exception as e:
        logger.error(f"❌ Error syncing DTDC tracking: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/dynamic-pricing/analyze")
async def analyze_product_velocity(days_lookback: int = 60):
    """
    Analyze product sales velocity and categorize as A/B/C
    - Category A: Fast-moving (top 20% - 10+ orders in last 60 days)
    - Category B: Medium-moving (next 30% - 3-9 orders in last 60 days)
    - Category C: Slow-moving (bottom 50% - 0-2 orders in last 60 days)
    
    Price multipliers are calculated based on 7-day rolling orders.
    This will update the cached report used by the dashboard.
    """
    try:
        from dynamic_pricing_engine import dynamic_pricing_engine
        
        logger.info(f"🔍 Starting product velocity analysis (last {days_lookback} days)...")
        
        result = await dynamic_pricing_engine.analyze_product_velocity(db, days_lookback)
        
        if result.get('success'):
            categories = result.get('categories', {})
            logger.info(f"✅ Analysis complete:")
            logger.info(f"   Category A (Fast): {len(categories.get('A', []))} products")
            logger.info(f"   Category B (Medium): {len(categories.get('B', []))} products")
            logger.info(f"   Category C (Slow): {len(categories.get('C', []))} products")
            
            # Update cached report
            await db.dynamic_pricing_cache.update_one(
                {"type": "analysis_report"},
                {"$set": {
                    "type": "analysis_report",
                    "data": result,
                    "last_updated": datetime.now(timezone.utc).isoformat()
                }},
                upsert=True
            )
            logger.info("✅ Cached report updated")
            
            # Also update pricing_rules collection with the new categories
            logger.info("📊 Syncing analysis to pricing_rules collection...")
            updated_count = 0
            for cat, products in categories.items():
                for product in products:
                    sku = product.get('sku')
                    if sku:
                        await db.pricing_rules.update_one(
                            {"sku": sku},
                            {"$set": {
                                "sku": sku,
                                "category": cat,
                                "historical_orders": product.get('order_count', 0),
                                "base_price": product.get('current_price', 0),
                                "product_name": product.get('product_name', sku),
                                "velocity_score": product.get('velocity_score', 0),
                                "total_revenue": product.get('total_revenue', 0),
                                "days_since_last_sale": product.get('days_since_last_sale', 0),
                                "classified_at": datetime.now(timezone.utc).isoformat(),
                                "enabled": True
                            }},
                            upsert=True
                        )
                        updated_count += 1
            
            logger.info(f"✅ Updated {updated_count} pricing rules")
            
            return result
        else:
            raise HTTPException(status_code=500, detail=result.get('error', 'Analysis failed'))
            
    except Exception as e:
        logger.error(f"❌ Error analyzing product velocity: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/dynamic-pricing/apply")
async def apply_dynamic_pricing(auto_apply: bool = False, days_lookback: int = 60):
    """
    Apply dynamic pricing based on product velocity analysis (last 60 days)
    - Category A: No discount (fast-moving, 10+ orders)
    - Category B: 10% discount (medium-moving, 3-9 orders)
    - Category C: 20% discount (slow-moving, 0-2 orders)
    
    Price changes are calculated based on 7-day rolling orders.
    Set auto_apply=true to automatically update inventory prices.
    """
    try:
        from dynamic_pricing_engine import dynamic_pricing_engine
        
        # First analyze
        logger.info("🔍 Analyzing product velocity...")
        analysis = await dynamic_pricing_engine.analyze_product_velocity(db, days_lookback)
        
        if not analysis.get('success'):
            raise HTTPException(status_code=500, detail=analysis.get('error', 'Analysis failed'))
        
        recommendations = analysis.get('pricing_recommendations', [])
        
        if not recommendations:
            return {
                'success': True,
                'message': 'No pricing recommendations generated',
                'total_products': 0
            }
        
        # Apply pricing
        logger.info(f"💰 Applying pricing recommendations (auto_apply={auto_apply})...")
        result = await dynamic_pricing_engine.apply_pricing(db, recommendations, auto_apply)
        
        # Include analysis summary
        result['analysis_summary'] = {
            'total_products': analysis.get('total_products', 0),
            'category_a_count': len(analysis.get('categories', {}).get('A', [])),
            'category_b_count': len(analysis.get('categories', {}).get('B', [])),
            'category_c_count': len(analysis.get('categories', {}).get('C', [])),
            'analysis_period_days': days_lookback
        }
        
        return result
        
    except Exception as e:
        logger.error(f"❌ Error applying dynamic pricing: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/dynamic-pricing/report")
async def get_pricing_report():
    """
    Get current dynamic pricing report with all categorized products
    Returns cached analysis from last run
    """
    try:
        # Get cached analysis from database
        cached_report = await db.dynamic_pricing_cache.find_one(
            {"type": "analysis_report"},
            {"_id": 0}
        )
        
        if not cached_report:
            # Return loading state - cache is being populated in background
            return {
                "success": True,
                "loading": True,
                "message": "Analysis is running in background. Please refresh in a moment.",
                "total_products": 0,
                "categories": {"A": [], "B": [], "C": []}
            }
        
        # Return cached data
        data = cached_report.get('data')
        data['loading'] = False
        data['last_updated'] = cached_report.get('last_updated')
        return data
        
    except Exception as e:
        logger.error(f"❌ Error getting pricing report: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


class ShopifySyncRequest(BaseModel):
    discounts: Dict[str, float] = {"A": 0, "B": 10, "C": 20}


@api_router.get("/dynamic-pricing/sync-status")
async def get_sync_status():
    """
    Get the status of the last Shopify sync
    """
    try:
        sync_result = await db.dynamic_pricing_cache.find_one(
            {"type": "last_sync_result"},
            {"_id": 0}
        )
        
        if not sync_result:
            return {
                "success": True,
                "has_sync": False,
                "message": "No sync has been run yet"
            }
        
        return {
            "success": True,
            "has_sync": True,
            "updated_count": sync_result.get('updated_count', 0),
            "total_products": sync_result.get('total_products', 0),
            "discounts_applied": sync_result.get('discounts_applied', {}),
            "completed_at": sync_result.get('completed_at'),
            "errors_count": len(sync_result.get('errors', []))
        }
        
    except Exception as e:
        logger.error(f"❌ Error getting sync status: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/dynamic-pricing/sync-to-shopify")
async def sync_pricing_to_shopify(request: ShopifySyncRequest):
    """
    Sync dynamic pricing to Shopify with custom discount percentages
    Processes synchronously with progress updates
    """
    try:
        from dynamic_pricing_engine import dynamic_pricing_engine
        
        # Get total count for response
        cached_report = await db.dynamic_pricing_cache.find_one(
            {"type": "analysis_report"},
            {"_id": 0}
        )
        
        total_products = 0
        categories_data = {}
        if cached_report:
            analysis = cached_report.get('data', {})
            categories_data = analysis.get('categories', {})
            total_products = sum(len(products) for products in categories_data.values())
        
        if total_products == 0:
            return {
                'success': False,
                'message': 'No products found. Please run "Re-Analyze Products" first.',
                'total_products': 0
            }
        
        # Store sync request in database for tracking
        await db.dynamic_pricing_cache.update_one(
            {"type": "sync_request"},
            {"$set": {
                "type": "sync_request",
                "discounts": request.discounts,
                "status": "processing",
                "requested_at": datetime.now(timezone.utc).isoformat(),
                "total_products": total_products
            }},
            upsert=True
        )
        
        logger.info("🔄 Starting Shopify price sync...")
        
        updated_count = 0
        failed_count = 0
        
        # Process each category with its discount
        for category, discount in request.discounts.items():
            products = categories_data.get(category, [])
            logger.info(f"📦 Processing Category {category}: {len(products)} products with {discount}% discount")
            
            # Process up to 100 products per category
            for product in products[:100]:
                try:
                    sku = product.get('sku')
                    current_price = product.get('current_price', 0)
                    
                    if current_price <= 0:
                        continue
                    
                    # Calculate discounted price
                    new_price = round(current_price * (1 - discount / 100), 2)
                    
                    # Update in local inventory_v2
                    result = await db.inventory_v2.update_one(
                        {"sku": sku},
                        {"$set": {
                            "dynamic_price": new_price,
                            "original_price": current_price,
                            "discount_percent": discount,
                            "velocity_category": category,
                            "price_updated_at": datetime.now(timezone.utc).isoformat()
                        }},
                        upsert=True
                    )
                    updated_count += 1
                    
                except Exception as e:
                    logger.error(f"Error updating {sku}: {e}")
                    failed_count += 1
        
        # Update sync status and save to last_sync_result for status endpoint
        sync_result = {
            "status": "completed",
            "completed_at": datetime.now(timezone.utc).isoformat(),
            "updated_count": updated_count,
            "failed_count": failed_count,
            "total_products": total_products,
            "discounts_applied": request.discounts
        }
        
        await db.dynamic_pricing_cache.update_one(
            {"type": "sync_request"},
            {"$set": sync_result},
            upsert=True
        )
        
        await db.dynamic_pricing_cache.update_one(
            {"type": "last_sync_result"},
            {"$set": {**sync_result, "type": "last_sync_result"}},
            upsert=True
        )
        
        logger.info(f"✅ Sync completed: {updated_count} updated, {failed_count} failed")
        
        return {
            'success': True,
            'message': f'Successfully updated {updated_count} products with dynamic pricing!',
            'updated_count': updated_count,
            'failed_count': failed_count,
            'total_products': total_products,
            'discounts_applied': request.discounts
        }
        
    except Exception as e:
        logger.error(f"❌ Error in sync: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/orders/{order_id}/manual-delivery-status")
async def update_manual_delivery_status(order_id: str, status: str, location: str = None, updated_by: str = None):
    """
    Manually update delivery status by employee
    """
    try:
        valid_statuses = ['PENDING', 'IN_TRANSIT', 'OUT_FOR_DELIVERY', 'DELIVERED', 'RETURN_IN_PROCESS', 'RETURNED', 'UNKNOWN']
        
        if status not in valid_statuses:
            raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {', '.join(valid_statuses)}")
        
        # Find customer
        customer = await db.customers.find_one({'order_number': order_id}, {'_id': 0})
        
        if not customer:
            raise HTTPException(status_code=404, detail="Order not found")
        
        # Update delivery status
        update_data = {
            'delivery_status': status,
            'delivery_updated_at': datetime.now(timezone.utc).isoformat(),
            'manually_updated': True,
            'manually_updated_by': updated_by or 'admin',
            'manually_updated_at': datetime.now(timezone.utc).isoformat()
        }
        
        if location:
            update_data['delivery_location'] = location
        
        # If status is RETURN_IN_PROCESS, add return fields
        if status == 'RETURN_IN_PROCESS':
            update_data['return_status'] = 'in_transit'
            update_data['return_received'] = False
            update_data['return_updated_at'] = datetime.now(timezone.utc).isoformat()
        
        # If status is RETURNED, mark as received
        if status == 'RETURNED':
            update_data['return_status'] = 'received'
            update_data['return_received'] = True
            update_data['return_updated_at'] = datetime.now(timezone.utc).isoformat()
        
        result = await db.customers.update_one(
            {'customer_id': customer['customer_id'], 'store_name': customer['store_name']},
            {'$set': update_data}
        )
        
        if result.modified_count == 0:
            raise HTTPException(status_code=500, detail="Failed to update order")
        
        logger.info(f"Manual status update by {updated_by}: Order {order_id} → {status}")
        
        return {
            "success": True,
            "message": f"Delivery status updated to {status}",
            "order_number": order_id,
            "new_status": status
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating manual delivery status: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.put("/customers/{order_number}/delivery-status")
async def update_customer_delivery_status(order_number: str, request: dict):
    """
    Update delivery status for a customer order (used by frontend TCS sync)
    """
    try:
        delivery_status = request.get('delivery_status')
        if not delivery_status:
            raise HTTPException(status_code=400, detail="delivery_status is required")
        
        # Build update data
        update_data = {
            "delivery_status": delivery_status,
            "delivery_updated_at": datetime.now(timezone.utc).isoformat(),
            "tcs_last_sync": datetime.now(timezone.utc).isoformat()
        }
        
        # Add weight and TCS charges if provided
        if request.get('tcs_weight'):
            update_data['tcs_weight'] = request.get('tcs_weight')
        if request.get('tcs_charges'):
            update_data['tcs_charges'] = request.get('tcs_charges')
        
        # Find and update by order_number
        result = await db.customers.update_one(
            {"$or": [
                {"order_number": order_number},
                {"order_number": int(order_number) if order_number.isdigit() else order_number}
            ]},
            {"$set": update_data}
        )
        
        if result.modified_count == 0:
            raise HTTPException(status_code=404, detail="Order not found")
        
        return {"success": True, "message": f"Delivery status updated to {delivery_status}"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating delivery status: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/orders/{order_id}/mark-return-received")
async def mark_return_received(order_id: str, received: bool = True, received_by: str = None):
    """
    Mark return as received or not received
    """
    try:
        customer = await db.customers.find_one({'order_number': order_id}, {'_id': 0})
        
        if not customer:
            raise HTTPException(status_code=404, detail="Order not found")
        
        result = await db.customers.update_one(
            {'customer_id': customer['customer_id'], 'store_name': customer['store_name']},
            {'$set': {
                'return_received': received,
                'return_received_by': received_by or 'admin',
                'return_received_at': datetime.now(timezone.utc).isoformat() if received else None,
                'return_status': 'received' if received else 'in_transit',
                'delivery_status': 'RETURNED' if received else 'RETURN_IN_PROCESS'
            }}
        )
        
        logger.info(f"Return marked as {'received' if received else 'not received'} for order {order_id}")
        
        return {
            "success": True,
            "message": f"Return marked as {'received' if received else 'not received'}",
            "order_number": order_id
        }
        
    except Exception as e:
        logger.error(f"Error marking return: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/inventory/v2/overview-stats")
async def get_inventory_overview_stats(
    start_date: str = None,
    end_date: str = None
):
    """Get comprehensive inventory stats with breakdowns by TCS tracking and fulfillment status"""
    try:
        from datetime import datetime, timezone, timedelta
        
        # Aggregate by collection
        collection_pipeline = [
            {"$match": {"collection": {"$exists": True, "$ne": None}}},
            {"$group": {
                "_id": "$collection",
                "count": {"$sum": 1},
                "total_cost": {"$sum": "$cost"}
            }},
            {"$sort": {"count": -1}}
        ]
        
        # Extract size from SKU (get last numeric part, typically shoe size)
        # We'll do this in Python for better control
        all_items_for_agg = await db.inventory_v2.find({}, {"_id": 0, "sku": 1}).to_list(100000)
        
        size_counts = {}
        for item in all_items_for_agg:
            sku = item.get('sku', '')
            parts = sku.split('-')
            # Look for numeric sizes (typically 20-50 for shoes)
            size_found = None
            for part in reversed(parts):  # Start from end
                # Check if it's a number or starts with a number
                import re
                size_match = re.match(r'^(\d+(?:\.\d+)?)', part.strip())
                if size_match:
                    size_num = float(size_match.group(1))
                    if 20 <= size_num <= 50:  # Typical shoe size range
                        size_found = str(int(size_num))
                        break
            
            if size_found:
                size_counts[size_found] = size_counts.get(size_found, 0) + 1
        
        by_size = [{"_id": size, "count": count} for size, count in sorted(size_counts.items(), key=lambda x: float(x[0]) if x[0].replace('.','').isdigit() else 0)]
        
        # Extract color from SKU
        color_pipeline = [
            {"$project": {
                "color": {
                    "$let": {
                        "vars": {
                            "parts": {"$split": ["$sku", "-"]}
                        },
                        "in": {
                            "$cond": {
                                "if": {"$gte": [{"$size": "$$parts"}, 2]},
                                "then": {"$arrayElemAt": ["$$parts", 1]},
                                "else": "Unknown"
                            }
                        }
                    }
                }
            }},
            {"$group": {
                "_id": "$color",
                "count": {"$sum": 1}
            }},
            {"$sort": {"count": -1}}
        ]
        
        # Run aggregations
        by_collection = await db.inventory_v2.aggregate(collection_pipeline).to_list(100)
        by_color = await db.inventory_v2.aggregate(color_pipeline).to_list(100)
        
        # Build date filter query
        date_query = {}
        if start_date or end_date:
            date_filter = {}
            if start_date:
                date_filter["$gte"] = start_date
            if end_date:
                # Add one day to include the end date
                from datetime import datetime, timedelta
                end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                end_dt = end_dt + timedelta(days=1)
                date_filter["$lt"] = end_dt.isoformat()
            date_query["created_at"] = date_filter
        
        # Get inventory items with date filter - ONLY items with valid order_number (sold items)
        # Filter out null, empty string, and "None" string values
        all_items = await db.inventory_v2.find(
            {
                **date_query, 
                "order_number": {"$exists": True, "$ne": None, "$ne": "", "$ne": "None"},
                "$expr": {"$ne": [{"$type": "$order_number"}, "null"]}
            },
            {"_id": 0, "sku": 1, "cost": 1, "order_number": 1, "sale_price": 1, "created_at": 1}
        ).to_list(100000)
        
        # Build order_number to inventory item mapping
        order_to_items = {}
        for item in all_items:
            order_num = str(item.get("order_number", ""))
            if order_num:
                if order_num not in order_to_items:
                    order_to_items[order_num] = []
                order_to_items[order_num].append(item)
        
        # Get all customer orders that match our inventory order numbers
        order_numbers = list(order_to_items.keys())
        
        # Fetch order details from customers collection
        customers_data = await db.customers.find(
            {"order_number": {"$in": [int(o) for o in order_numbers if o.isdigit()] + order_numbers}},
            {
                "_id": 0,
                "order_number": 1,
                "fulfillment_status": 1,
                "tracking_number": 1,
                "delivery_status": 1,
                "total_spent": 1
            }
        ).to_list(100000)
        
        # Create order lookup
        order_lookup = {}
        for cust in customers_data:
            order_num = str(cust.get("order_number", ""))
            order_lookup[order_num] = cust
        
        # Categorize inventory based on order status
        ready_to_ship = []  # Unfulfilled orders (not yet shipped)
        in_transit = []  # Fulfilled with tracking, not delivered
        delivered = []  # Delivered orders
        unknown_status = []  # Orders not found or no status
        
        # Track unique orders for each category
        ready_orders = {}
        transit_orders = {}
        delivered_orders = {}
        unknown_orders = {}
        
        for item in all_items:
            order_num = str(item.get("order_number", ""))
            cost = item.get("cost", 0)
            
            if order_num in order_lookup:
                order = order_lookup[order_num]
                fulfillment = (order.get("fulfillment_status") or "").lower()
                tracking = order.get("tracking_number")
                delivery = (order.get("delivery_status") or "").upper()
                total_spent = order.get("total_spent", 0)
                
                # Category 1: Ready to Ship (unfulfilled)
                if fulfillment in ["unfulfilled", "", "null"] or fulfillment is None:
                    ready_to_ship.append(item)
                    if order_num not in ready_orders:
                        ready_orders[order_num] = total_spent
                
                # Category 2: Delivered
                elif delivery == "DELIVERED":
                    delivered.append(item)
                    if order_num not in delivered_orders:
                        delivered_orders[order_num] = total_spent
                
                # Category 3: In Transit (fulfilled with tracking, not delivered)
                elif tracking and delivery in ["PENDING", "IN_TRANSIT", "BOOKED", "OUT_FOR_DELIVERY", 
                                                "ARRIVAL_AT_DESTINATION", "RETURN_IN_PROCESS", ""]:
                    in_transit.append(item)
                    if order_num not in transit_orders:
                        transit_orders[order_num] = total_spent
                
                # Category 4: Fulfilled but no tracking yet - also ready to ship
                elif fulfillment == "fulfilled" and not tracking:
                    ready_to_ship.append(item)
                    if order_num not in ready_orders:
                        ready_orders[order_num] = total_spent
                
                # Otherwise unknown
                else:
                    unknown_status.append(item)
                    if order_num not in unknown_orders:
                        unknown_orders[order_num] = total_spent
            else:
                # Order not found in customers - unknown
                unknown_status.append(item)
        
        # Use new categorized lists
        can_fulfill_today = ready_to_ship
        in_transit_tracked = in_transit
        delivered_recent = delivered
        unknown_old = unknown_status
        
        fulfill_orders = ready_orders
        transit_orders = transit_orders
        delivered_orders = delivered_orders
        
        # Calculate stats
        def calc_stats(items, orders_map):
            cost = sum(i.get('cost', 0) for i in items)
            sale_value = sum(orders_map.values()) if orders_map else 0
            profit = sale_value - cost if sale_value > 0 else 0
            return {
                "count": len(items),
                "cost": round(cost, 2),
                "sale_value": round(sale_value, 2),
                "profit": round(profit, 2)
            }
        
        # Get all orders that have been matched with inventory (for totals and unknown calc)
        all_matched_orders_for_calc = await db.customers.find({
            "order_skus": {"$exists": True, "$ne": []},
            "order_number": {"$exists": True}
        }, {
            "_id": 0,
            "order_number": 1,
            "total_spent": 1,
            "order_skus": 1
        }).to_list(50000)
        
        # Create SKU set from inventory
        inventory_skus = {item.get('sku', '').upper().strip() for item in all_items}
        
        # Calculate totals from all orders that match our inventory
        total_matched_orders = {}
        for order in all_matched_orders_for_calc:
            order_skus = [sku.upper().strip() for sku in order.get('order_skus', [])]
            # Check if any SKU from this order is in our inventory
            if any(sku in inventory_skus for sku in order_skus):
                order_num = order.get('order_number')
                if order_num not in total_matched_orders:
                    total_matched_orders[order_num] = order.get('total_spent', 0)
        
        # Calculate stats for categorized items
        fulfill_stats = calc_stats(can_fulfill_today, fulfill_orders)
        transit_stats = calc_stats(in_transit_tracked, transit_orders)
        delivered_stats = calc_stats(delivered_recent, delivered_orders)
        
        # For unknown/old, calculate from all remaining matched orders
        unknown_matched_orders = {}
        unknown_skus = {item.get('sku', '').upper().strip() for item in unknown_old}
        for order in all_matched_orders_for_calc:
            order_skus = [sku.upper().strip() for sku in order.get('order_skus', [])]
            if any(sku in unknown_skus for sku in order_skus):
                order_num = order.get('order_number')
                # Exclude if already counted in other categories
                if (order_num not in fulfill_orders and 
                    order_num not in transit_orders and 
                    order_num not in delivered_orders):
                    unknown_matched_orders[order_num] = order.get('total_spent', 0)
        
        unknown_stats = calc_stats(unknown_old, unknown_matched_orders)
        
        # Calculate overall totals
        total_cost = sum(item.get('cost', 0) for item in all_items)
        total_sale_value = sum(total_matched_orders.values())
        total_profit = total_sale_value - total_cost if total_sale_value > 0 else 0
        
        # Calculate cost breakdown by currency (INR for India stores, PKR for Pakistan stores)
        # Get all inventory items with store info
        all_items_with_store = await db.inventory_v2.find(
            {
                **date_query, 
                "order_number": {"$exists": True, "$ne": None, "$ne": "", "$ne": "None"},
                "$expr": {"$ne": [{"$type": "$order_number"}, "null"]}
            },
            {"_id": 0, "cost": 1, "store_name": 1}
        ).to_list(100000)
        
        # INR stores: tnvcollection, ashmiaa (India)
        # PKR stores: tnvcollectionpk (Pakistan)
        inr_stores = ['tnvcollection', 'ashmiaa', 'asmia']
        pkr_stores = ['tnvcollectionpk']
        
        total_cost_inr = 0
        total_cost_pkr = 0
        
        for item in all_items_with_store:
            cost = item.get('cost', 0) or 0
            store = (item.get('store_name') or '').lower()
            
            if store in inr_stores:
                total_cost_inr += cost
            elif store in pkr_stores:
                total_cost_pkr += cost
            else:
                # Default to PKR if unknown store
                total_cost_pkr += cost
        
        # NEW: Calculate sale value from inventory SKU sale prices
        # Get all inventory items with sale_price field
        inventory_with_sale_price = await db.inventory_v2.find(
            date_query,
            {"_id": 0, "sku": 1, "cost": 1, "sale_price": 1}
        ).to_list(100000)
        
        total_inventory_sale_value = 0
        total_inventory_cost = 0
        items_with_sale_price = 0
        
        for inv_item in inventory_with_sale_price:
            cost = inv_item.get('cost', 0)
            sale_price = inv_item.get('sale_price', 0)
            
            total_inventory_cost += cost
            
            # Use sale_price if available, otherwise use cost as fallback
            if sale_price and sale_price > 0:
                total_inventory_sale_value += sale_price
                items_with_sale_price += 1
            else:
                total_inventory_sale_value += cost
        
        inventory_profit = total_inventory_sale_value - total_inventory_cost
        
        return {
            "success": True,
            "stats": {
                "total_items": len(all_items),
                "total_cost": round(total_cost, 2),
                "total_cost_inr": round(total_cost_inr, 2),
                "total_cost_pkr": round(total_cost_pkr, 2),
                "total_sale_value": round(total_sale_value, 2),
                "total_profit": round(total_profit, 2),
                # NEW: Inventory-based sale value (from SKU sale prices)
                "inventory_sale_value": round(total_inventory_sale_value, 2),
                "inventory_profit": round(inventory_profit, 2),
                "items_with_sale_price": items_with_sale_price,
                "can_fulfill_today": fulfill_stats,
                "in_transit_tracked": transit_stats,
                "delivered_recent": delivered_stats,
                "unknown_old": unknown_stats,
                "by_collection": by_collection,
                "by_size": by_size,
                "by_color": by_color
            }
        }
    except Exception as e:
        logger.error(f"Error fetching inventory overview stats: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/inventory/v2/overview-detail/{category}")
async def get_inventory_category_detail(
    category: str,
    start_date: str = None,
    end_date: str = None
):
    """Get detailed breakdown for a specific inventory category"""
    try:
        from datetime import datetime, timezone, timedelta
        
        # Build date filter query
        date_query = {}
        if start_date or end_date:
            date_filter = {}
            if start_date:
                date_filter["$gte"] = start_date
            if end_date:
                from datetime import datetime, timedelta
                end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                end_dt = end_dt + timedelta(days=1)
                date_filter["$lt"] = end_dt.isoformat()
            date_query["created_at"] = date_filter
        
        # Get inventory items with date filter
        all_items = await db.inventory_v2.find(date_query, {
            "_id": 0, 
            "sku": 1,
            "cost": 1,
            "sale_price": 1,
            "order_number": 1,
            "collection": 1,
            "created_at": 1
        }).to_list(10000)
        
        # Get today's date
        today = datetime.now(timezone.utc).date()
        
        # Get unfulfilled orders from today
        unfulfilled_today = await db.customers.find({
            "fulfillment_status": {"$in": ["unfulfilled", "Unfulfilled", "UNFULFILLED", None]},
            "order_skus": {"$exists": True, "$ne": []},
            "created_at": {"$exists": True}
        }, {
            "_id": 0,
            "order_number": 1,
            "order_skus": 1,
            "total_spent": 1,
            "created_at": 1,
            "first_name": 1,
            "last_name": 1
        }).to_list(10000)
        
        # Get orders with tracking
        orders_with_tracking = await db.customers.find({
            "tracking_number": {"$exists": True, "$ne": None, "$ne": ""},
            "order_skus": {"$exists": True, "$ne": []}
        }, {
            "_id": 0,
            "order_number": 1,
            "order_skus": 1,
            "delivery_status": 1,
            "tracking_number": 1,
            "total_spent": 1,
            "created_at": 1,
            "first_name": 1,
            "last_name": 1
        }).to_list(50000)
        
        # Create mappings
        sku_to_unfulfilled_orders = {}
        sku_to_tracked_orders = {}
        
        # Map unfulfilled orders
        for order in unfulfilled_today:
            try:
                order_date = datetime.fromisoformat(order.get("created_at", "")).date()
                if (today - order_date).days <= 1:
                    for sku in order.get("order_skus", []):
                        sku_upper = sku.upper().strip()
                        if sku_upper not in sku_to_unfulfilled_orders:
                            sku_to_unfulfilled_orders[sku_upper] = []
                        sku_to_unfulfilled_orders[sku_upper].append({
                            "order_number": order.get("order_number"),
                            "total_spent": order.get("total_spent", 0),
                            "customer": f"{order.get('first_name', '')} {order.get('last_name', '')}".strip()
                        })
            except:
                continue
        
        # Map tracked orders
        for order in orders_with_tracking:
            delivery_status = order.get("delivery_status", "")
            if delivery_status and delivery_status != "UNKNOWN":
                for sku in order.get("order_skus", []):
                    sku_upper = sku.upper().strip()
                    if sku_upper not in sku_to_tracked_orders:
                        sku_to_tracked_orders[sku_upper] = []
                    sku_to_tracked_orders[sku_upper].append({
                        "order_number": order.get("order_number"),
                        "delivery_status": delivery_status,
                        "tracking_number": order.get("tracking_number"),
                        "total_spent": order.get("total_spent", 0),
                        "customer": f"{order.get('first_name', '')} {order.get('last_name', '')}".strip()
                    })
        
        # Categorize items based on requested category
        detailed_items = []
        
        for item in all_items:
            sku_upper = item.get('sku', '').upper().strip()
            cost = item.get('cost', 0)
            sale_price = item.get('sale_price', 0)
            
            item_detail = {
                "sku": item.get('sku'),
                "cost": cost,
                "sale_price": sale_price,
                "collection": item.get('collection'),
                "orders": []
            }
            
            # Filter by category
            if category == "can_fulfill_today":
                if sku_upper in sku_to_unfulfilled_orders:
                    item_detail["orders"] = sku_to_unfulfilled_orders[sku_upper]
                    detailed_items.append(item_detail)
            
            elif category == "in_transit":
                if sku_upper in sku_to_tracked_orders:
                    for order in sku_to_tracked_orders[sku_upper]:
                        if order["delivery_status"] in ["BOOKED", "IN_TRANSIT", "OUT_FOR_DELIVERY", "ARRIVAL_AT_DESTINATION"]:
                            item_detail["orders"].append(order)
                    if item_detail["orders"]:
                        detailed_items.append(item_detail)
            
            elif category == "delivered":
                if sku_upper in sku_to_tracked_orders:
                    for order in sku_to_tracked_orders[sku_upper]:
                        if order["delivery_status"] == "DELIVERED":
                            item_detail["orders"].append(order)
                    if item_detail["orders"]:
                        detailed_items.append(item_detail)
            
            elif category == "other":
                # Items not in other categories
                if (sku_upper not in sku_to_unfulfilled_orders and 
                    sku_upper not in sku_to_tracked_orders):
                    detailed_items.append(item_detail)
        
        return {
            "success": True,
            "category": category,
            "total_items": len(detailed_items),
            "items": detailed_items[:500]  # Limit to 500 for performance
        }
    
    except Exception as e:
        logger.error(f"Error fetching category detail: {str(e)}")


@api_router.get("/inventory/v2/collection-detail/{collection_name}")
async def get_collection_detail(collection_name: str):
    """Get detailed breakdown for a specific collection"""
    try:
        # Get all items in this collection
        items = await db.inventory_v2.find(
            {"collection": collection_name},
            {"_id": 0, "sku": 1, "cost": 1, "sale_price": 1, "order_number": 1, "collection": 1}
        ).to_list(1000)
        
        # Get related orders
        skus = [item.get('sku', '').upper().strip() for item in items]
        orders = await db.customers.find(
            {"order_skus": {"$in": skus}},
            {"_id": 0, "order_number": 1, "order_skus": 1, "total_spent": 1, "first_name": 1, "last_name": 1, "tracking_number": 1, "delivery_status": 1}
        ).to_list(1000)
        
        # Map SKUs to orders
        sku_to_orders = {}
        for order in orders:
            for sku in order.get('order_skus', []):
                sku_upper = sku.upper().strip()
                if sku_upper not in sku_to_orders:
                    sku_to_orders[sku_upper] = []
                sku_to_orders[sku_upper].append({
                    "order_number": order.get("order_number"),
                    "customer": f"{order.get('first_name', '')} {order.get('last_name', '')}".strip(),
                    "tracking_number": order.get("tracking_number"),
                    "delivery_status": order.get("delivery_status"),
                    "total_spent": order.get("total_spent", 0)
                })
        
        # Add orders to items
        detailed_items = []
        for item in items:
            sku_upper = item.get('sku', '').upper().strip()
            item['orders'] = sku_to_orders.get(sku_upper, [])
            detailed_items.append(item)
        
        total_cost = sum(item.get('cost', 0) for item in items)
        total_sale = sum(item.get('sale_price', item.get('cost', 0)) for item in items)
        
        return {
            "success": True,
            "collection": collection_name,
            "total_items": len(items),
            "total_cost": round(total_cost, 2),
            "total_sale": round(total_sale, 2),
            "items": detailed_items[:500]
        }
    except Exception as e:
        logger.error(f"Error fetching collection detail: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/inventory/v2/size-detail/{size}")
async def get_size_detail(size: str):
    """Get detailed breakdown for a specific size"""
    try:
        # Find items with this size in SKU
        all_items = await db.inventory_v2.find({}, {"_id": 0}).to_list(10000)
        
        # Filter by size
        import re
        filtered_items = []
        for item in all_items:
            sku = item.get('sku', '')
            # Extract size from SKU
            size_match = re.search(r'-(\d+(?:\.\d+)?)\s*$', sku)
            if size_match:
                item_size = size_match.group(1)
                if item_size == size:
                    filtered_items.append(item)
        
        # Get related orders
        skus = [item.get('sku', '').upper().strip() for item in filtered_items]
        orders = await db.customers.find(
            {"order_skus": {"$in": skus}},
            {"_id": 0, "order_number": 1, "order_skus": 1, "total_spent": 1, "first_name": 1, "last_name": 1, "tracking_number": 1, "delivery_status": 1}
        ).to_list(1000)
        
        # Map SKUs to orders
        sku_to_orders = {}
        for order in orders:
            for sku in order.get('order_skus', []):
                sku_upper = sku.upper().strip()
                if sku_upper not in sku_to_orders:
                    sku_to_orders[sku_upper] = []
                sku_to_orders[sku_upper].append({
                    "order_number": order.get("order_number"),
                    "customer": f"{order.get('first_name', '')} {order.get('last_name', '')}".strip(),
                    "tracking_number": order.get("tracking_number"),
                    "delivery_status": order.get("delivery_status")
                })
        
        # Add orders to items
        detailed_items = []
        for item in filtered_items:
            sku_upper = item.get('sku', '').upper().strip()
            item['orders'] = sku_to_orders.get(sku_upper, [])
            detailed_items.append(item)
        
        total_cost = sum(item.get('cost', 0) for item in filtered_items)
        
        return {
            "success": True,
            "size": size,
            "total_items": len(filtered_items),
            "total_cost": round(total_cost, 2),
            "items": detailed_items[:500]
        }
    except Exception as e:
        logger.error(f"Error fetching size detail: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/inventory/v2/all-items-detail")
async def get_all_items_detail(
    start_date: str = None,
    end_date: str = None
):
    """Get detailed breakdown of all inventory items"""
    try:
        from datetime import datetime, timedelta
        
        # Build date filter
        date_query = {}
        if start_date or end_date:
            date_filter = {}
            if start_date:
                date_filter["$gte"] = start_date
            if end_date:
                end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                end_dt = end_dt + timedelta(days=1)
                date_filter["$lt"] = end_dt.isoformat()
            date_query["created_at"] = date_filter
        
        items = await db.inventory_v2.find(
            date_query,
            {"_id": 0, "sku": 1, "cost": 1, "sale_price": 1, "order_number": 1, "collection": 1, "created_at": 1}
        ).to_list(1000)
        
        # Get related orders
        skus = [item.get('sku', '').upper().strip() for item in items]
        orders = await db.customers.find(
            {"order_skus": {"$in": skus}},
            {"_id": 0, "order_number": 1, "order_skus": 1, "total_spent": 1, "first_name": 1, "last_name": 1, "tracking_number": 1, "delivery_status": 1}
        ).to_list(1000)
        
        # Map SKUs to orders
        sku_to_orders = {}
        for order in orders:
            for sku in order.get('order_skus', []):
                sku_upper = sku.upper().strip()
                if sku_upper not in sku_to_orders:
                    sku_to_orders[sku_upper] = []
                sku_to_orders[sku_upper].append({
                    "order_number": order.get("order_number"),
                    "customer": f"{order.get('first_name', '')} {order.get('last_name', '')}".strip(),
                    "tracking_number": order.get("tracking_number"),
                    "delivery_status": order.get("delivery_status"),
                    "total_spent": order.get("total_spent", 0)
                })
        
        # Add orders to items
        detailed_items = []
        for item in items:
            sku_upper = item.get('sku', '').upper().strip()
            item['orders'] = sku_to_orders.get(sku_upper, [])
            detailed_items.append(item)
        
        total_cost = sum(item.get('cost', 0) for item in items)
        total_sale = sum(item.get('sale_price', item.get('cost', 0)) for item in items)
        
        return {
            "success": True,
            "total_items": len(items),
            "total_cost": round(total_cost, 2),
            "total_sale": round(total_sale, 2),
            "items": detailed_items[:1000]
        }
    except Exception as e:
        logger.error(f"Error fetching all items detail: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/inventory/v2/search")
async def search_inventory(q: str):
    """Search inventory by SKU, order number, collection, or customer name"""
    try:
        if not q or len(q) < 2:
            return {
                "success": False,
                "message": "Search query must be at least 2 characters"
            }
        
        search_query = q.strip()
        
        # Search in inventory_v2 collection
        # Create a case-insensitive regex pattern
        regex_pattern = {"$regex": search_query, "$options": "i"}
        
        # Search by SKU or collection
        inventory_items = await db.inventory_v2.find({
            "$or": [
                {"sku": regex_pattern},
                {"collection": regex_pattern},
                {"order_number": regex_pattern}
            ]
        }, {
            "_id": 0,
            "sku": 1,
            "cost": 1,
            "sale_price": 1,
            "order_number": 1,
            "collection": 1
        }).to_list(100)
        
        # Search in customers collection for order numbers or customer names
        customer_orders = await db.customers.find({
            "$or": [
                {"order_number": regex_pattern},
                {"first_name": regex_pattern},
                {"last_name": regex_pattern},
                {"tracking_number": regex_pattern}
            ]
        }, {
            "_id": 0,
            "order_number": 1,
            "order_skus": 1,
            "first_name": 1,
            "last_name": 1,
            "tracking_number": 1,
            "total_spent": 1,
            "delivery_status": 1
        }).to_list(50)
        
        # Create a mapping of SKUs to orders
        sku_to_orders = {}
        for order in customer_orders:
            for sku in order.get('order_skus', []):
                sku_upper = sku.upper().strip()
                if sku_upper not in sku_to_orders:
                    sku_to_orders[sku_upper] = []
                sku_to_orders[sku_upper].append({
                    "order_number": order.get("order_number"),
                    "customer": f"{order.get('first_name', '')} {order.get('last_name', '')}".strip(),
                    "tracking_number": order.get("tracking_number"),
                    "delivery_status": order.get("delivery_status"),
                    "total_spent": order.get("total_spent", 0)
                })
        
        # If searching by order number or customer name, find related SKUs
        if customer_orders:
            order_skus = set()
            for order in customer_orders:
                order_skus.update([sku.upper().strip() for sku in order.get('order_skus', [])])
            
            # Get inventory items for these SKUs
            additional_items = await db.inventory_v2.find({
                "sku": {"$in": list(order_skus)}
            }, {
                "_id": 0,
                "sku": 1,
                "cost": 1,
                "sale_price": 1,
                "order_number": 1,
                "collection": 1
            }).to_list(100)
            
            # Merge with existing inventory items (avoid duplicates)
            existing_skus = {item.get('sku') for item in inventory_items}
            for item in additional_items:
                if item.get('sku') not in existing_skus:
                    inventory_items.append(item)
        
        # Add order information to inventory items
        results = []
        for item in inventory_items:
            sku_upper = item.get('sku', '').upper().strip()
            item['orders'] = sku_to_orders.get(sku_upper, [])
            results.append(item)
        
        # Sort by relevance (exact match first, then partial match)
        def relevance_score(item):
            sku = item.get('sku', '').lower()
            collection = item.get('collection', '').lower()
            query_lower = search_query.lower()
            
            if sku == query_lower:
                return 0  # Exact SKU match
            elif sku.startswith(query_lower):
                return 1  # SKU starts with query
            elif query_lower in sku:
                return 2  # SKU contains query
            elif collection == query_lower:
                return 3  # Exact collection match
            elif query_lower in collection:
                return 4  # Collection contains query
            else:
                return 5  # Other matches
        
        results.sort(key=relevance_score)
        
        return {
            "success": True,
            "query": search_query,
            "total_results": len(results),
            "items": results[:100]  # Limit to 100 results
        }
    
    except Exception as e:
        logger.error(f"Error searching inventory: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


        raise HTTPException(status_code=500, detail=str(e))



@api_router.get("/inventory/v2/stats")
async def get_inventory_stats():
    """Get financial stats for inventory"""
    try:
        # Get all inventory items
        all_items = await db.inventory_v2.find({}, {"_id": 0}).to_list(10000)
        
        # Calculate totals
        total_cost = sum(item.get('cost', 0) for item in all_items)
        total_sale_value = sum(item.get('sale_price', 0) for item in all_items if item.get('sale_price', 0) > 0)
        total_profit = sum(item.get('profit', 0) for item in all_items if item.get('profit', 0) > 0)
        
        # Count by status
        in_stock_count = len([item for item in all_items if item.get('status') == 'in_stock'])
        in_transit_count = len([item for item in all_items if item.get('status') == 'in_transit'])
        delivered_count = len([item for item in all_items if item.get('status') == 'delivered'])
        
        # Get today's orders (items created today)
        from datetime import datetime, timezone
        today = datetime.now(timezone.utc).date()
        today_items = [item for item in all_items if datetime.fromisoformat(item.get('created_at', '')).date() == today]
        today_orders_count = len(today_items)
        
        # Get orders from customers
        customers_with_orders = await db.customers.find(
            {"order_skus": {"$exists": True, "$ne": []}},
            {"_id": 0, "delivery_status": 1}
        ).to_list(50000)
        
        # Count transit orders (not delivered)
        transit_orders = len([c for c in customers_with_orders if c.get('delivery_status') and c.get('delivery_status') not in ['DELIVERED', 'RETURN_IN_PROCESS']])
        delivered_orders = len([c for c in customers_with_orders if c.get('delivery_status') == 'DELIVERED'])
        
        return {
            "success": True,
            "stats": {
                "total_inventory_items": len(all_items),
                "total_cost": round(total_cost, 2),
                "total_sale_value": round(total_sale_value, 2),
                "total_profit": round(total_profit, 2),
                "in_stock_count": in_stock_count,
                "in_transit_count": in_transit_count,
                "delivered_count": delivered_count,
                "today_orders": today_orders_count,
                "transit_orders": transit_orders,
                "delivered_orders": delivered_orders
            }
        }
    except Exception as e:
        logger.error(f"Error fetching inventory stats: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/inventory/v2")
async def get_all_inventory_items(
    store_name: str = None, 
    status: str = None, 
    search: str = None,
    has_price: str = None,
    profitable: str = None,
    sort_by: str = None,
    currency: str = None,
    limit: int = 100
):
    """Get all inventory items with filtering and search"""
    try:
        query = {}
        sort_field = [("created_at", -1)]  # Default sort
        
        if store_name and store_name != "all":
            query["store_name"] = store_name
        if status and status != "all":
            query["status"] = status
        
        # Filter by currency (store region)
        # INR stores: tnvcollection, ashmiaa (India)
        # PKR stores: tnvcollectionpk (Pakistan)
        if currency == "inr":
            query["store_name"] = {"$in": ["tnvcollection", "ashmiaa", "asmia"]}
        elif currency == "pkr":
            query["store_name"] = {"$in": ["tnvcollectionpk"]}
        
        # Filter for items with sale price
        if has_price == "true":
            query["sale_price"] = {"$gt": 0}
        
        # Filter for profitable items
        if profitable == "true":
            query["profit"] = {"$gt": 0}
        
        # Sort options
        if sort_by == "cost_desc":
            sort_field = [("cost", -1)]
        elif sort_by == "cost_asc":
            sort_field = [("cost", 1)]
        elif sort_by == "profit_desc":
            sort_field = [("profit", -1)]
        elif sort_by == "price_desc":
            sort_field = [("sale_price", -1)]
        
        # Add search filter using $and to combine with other filters
        if search and search.strip():
            search_regex = {"$regex": search.strip(), "$options": "i"}
            search_conditions = {
                "$or": [
                    {"sku": search_regex},
                    {"product_name": search_regex},
                    {"collection": search_regex},
                    {"order_number": search_regex}
                ]
            }
            # Combine existing query with search using $and
            if query:
                query = {"$and": [query, search_conditions]}
            else:
                query = search_conditions
        
        items = await db.inventory_v2.find(query, {"_id": 0}).sort(sort_field).limit(limit).to_list(limit)
        return {"success": True, "items": items, "total": len(items)}
    except Exception as e:
        logger.error(f"Error fetching inventory: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/inventory/v2/add")
async def add_inventory_item(item: InventoryItemCreate, added_by: str = None):
    """Add new inventory item with optional Shopify order linking"""
    try:
        # Create inventory item
        new_item = InventoryItem(
            sku=item.sku,
            product_name=item.product_name,
            collection=item.collection,
            order_number=item.order_number,
            cost=item.cost,
            store_name=item.store_name,
            added_by=added_by
        )
        
        # If order number provided, fetch sale price from Shopify/customer data
        if item.order_number:
            customer = await db.customers.find_one(
                {"order_number": str(item.order_number)},
                {"_id": 0, "total_spent": 1}
            )
            if customer:
                new_item.sale_price = float(customer.get("total_spent", 0))
                new_item.profit = new_item.sale_price - new_item.cost
        
        # Insert into database
        item_dict = new_item.model_dump()
        await db.inventory_v2.insert_one(item_dict)
        
        # Invalidate inventory cache after adding new item
        asyncio.create_task(invalidate_inventory_cache())
        
        # Fetch the inserted item without _id
        inserted_item = await db.inventory_v2.find_one({"id": new_item.id}, {"_id": 0})
        
        return {"success": True, "message": "Inventory item added", "item": inserted_item}
    except Exception as e:
        logger.error(f"Error adding inventory item: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.put("/inventory/v2/{item_id}")
async def update_inventory_item(item_id: str, update: InventoryItemUpdate):
    """Update inventory item"""
    try:
        update_data = {k: v for k, v in update.model_dump().items() if v is not None}
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        
        # Recalculate profit if cost or sale_price updated
        item = await db.inventory_v2.find_one({"id": item_id}, {"_id": 0})
        if item:
            cost = update_data.get("cost", item.get("cost", 0))
            sale_price = update_data.get("sale_price", item.get("sale_price", 0))
            
            # If order number is being updated, fetch sale price
            if "order_number" in update_data and update_data["order_number"]:
                customer = await db.customers.find_one(
                    {"order_number": str(update_data["order_number"])},
                    {"_id": 0, "total_spent": 1}
                )
                if customer:
                    sale_price = float(customer.get("total_spent", 0))
                    update_data["sale_price"] = sale_price
            
            update_data["profit"] = sale_price - cost
        
        result = await db.inventory_v2.update_one(
            {"id": item_id},
            {"$set": update_data}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Inventory item not found")
        
        # Invalidate inventory cache after updating
        asyncio.create_task(invalidate_inventory_cache())
        
        return {"success": True, "message": "Inventory item updated"}
    except Exception as e:
        logger.error(f"Error updating inventory item: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/inventory/v2/{item_id}/delivery-status")
async def add_delivery_status(item_id: str, status_update: DeliveryStatusUpdate):
    """Add manual delivery status entry to timeline"""
    try:
        # Create status entry
        timestamp = status_update.timestamp or datetime.now(timezone.utc).isoformat()
        entry = DeliveryStatusEntry(
            timestamp=timestamp,
            status=status_update.status
        )
        
        # Add to timeline
        result = await db.inventory_v2.update_one(
            {"id": item_id},
            {
                "$push": {"delivery_timeline": entry.model_dump()},
                "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}
            }
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Inventory item not found")
        
        return {"success": True, "message": "Delivery status added"}
    except Exception as e:
        logger.error(f"Error adding delivery status: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.delete("/inventory/v2/{item_id}")
async def delete_inventory_item(item_id: str):
    """Delete inventory item"""
    try:
        result = await db.inventory_v2.delete_one({"id": item_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Inventory item not found")
        
        return {"success": True, "message": "Inventory item deleted"}
    except Exception as e:
        logger.error(f"Error deleting inventory item: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/inventory/v2/upload")
async def upload_inventory_excel(file: UploadFile = File(...), store_name: str = "tnvcollectionpk"):
    """Upload Excel file with inventory (Box No, SKU, Size, Color, Collection, Cost)"""
    try:
        import openpyxl
        from io import BytesIO
        
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        sheet = wb.active
        
        items_added = 0
        errors = []
        
        # Get all customers with their order_skus and line_items for matching
        customers = await db.customers.find(
            {"order_skus": {"$exists": True, "$ne": []}},
            {"_id": 0, "order_number": 1, "order_skus": 1, "total_spent": 1, "line_items": 1}
        ).to_list(50000)
        
        # Create a mapping: SKU -> list of (order_number, price_per_item)
        sku_to_orders = {}
        for customer in customers:
            order_skus = customer.get("order_skus", [])
            line_items = customer.get("line_items", [])
            order_number = customer.get("order_number")
            
            # Try to match SKU with line items to get individual item price
            for sku in order_skus:
                sku_upper = sku.upper().strip()
                if sku_upper not in sku_to_orders:
                    sku_to_orders[sku_upper] = []
                
                # Find matching line item for this SKU to get the actual product price
                item_price = 0
                for line_item in line_items:
                    line_sku = line_item.get("sku", "").upper().strip()
                    if line_sku == sku_upper:
                        item_price = float(line_item.get("price", 0))
                        break
                
                # Fallback to total_spent / number of items if price not found
                if item_price == 0 and len(order_skus) > 0:
                    item_price = float(customer.get("total_spent", 0)) / len(order_skus)
                
                sku_to_orders[sku_upper].append({
                    "order_number": order_number,
                    "sale_price": item_price
                })
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Skip if SKU is empty (column 2, index 1)
                if not row[1]:
                    continue
                
                box_no = str(row[0]).strip() if row[0] else None
                sku = str(row[1]).strip()
                size = str(row[2]).strip() if row[2] else None
                color = str(row[3]).strip() if row[3] else None
                collection = str(row[4]).strip() if row[4] else None
                cost = float(row[5]) if row[5] else 0.0
                
                # Build product name from SKU, size, color
                product_name = f"{sku}"
                if size:
                    product_name += f" - Size {size}"
                if color:
                    product_name += f" - {color}"
                
                # Check if this SKU matches any orders
                sku_upper = sku.upper().strip()
                order_number = None
                sale_price = 0.0
                
                if sku_upper in sku_to_orders:
                    # Get the first matching order with the product price
                    first_order = sku_to_orders[sku_upper][0]
                    order_number = first_order["order_number"]
                    sale_price = float(first_order["sale_price"])
                
                # Create inventory item
                new_item = InventoryItem(
                    sku=sku,
                    product_name=product_name,
                    collection=collection,
                    cost=cost,
                    order_number=order_number,
                    sale_price=sale_price,
                    profit=sale_price - cost if sale_price > 0 else 0.0,
                    store_name=store_name
                )
                
                item_dict = new_item.model_dump()
                await db.inventory_v2.insert_one(item_dict)
                items_added += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
                continue
        
        # Auto-sync inventory with store orders after upload
        sync_result = await sync_inventory_with_store(store_name)
        
        return {
            "success": True,
            "message": f"Processed {items_added} inventory items and synced with {store_name}",
            "items_added": items_added,
            "errors": errors if errors else None,
            "sync_status": sync_result
        }
        
    except Exception as e:
        logger.error(f"Error uploading inventory: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/inventory/sync/{store_name}")
async def sync_inventory_with_store(store_name: str):
    """
    Sync uploaded inventory with store orders to update:
    - Confirmation Tracker (show stock availability)
    - Dispatch Tracker (show stock availability)  
    - Inventory Overview (update stats)
    - Dynamic Pricing (update pricing based on stock)
    - Inventory Health (update health metrics)
    """
    try:
        # Get all inventory items for this store
        inventory_items = await db.inventory_v2.find(
            {"store_name": store_name},
            {"_id": 0, "sku": 1, "cost": 1, "sale_price": 1, "profit": 1}
        ).to_list(50000)
        
        if not inventory_items:
            return {
                "success": False,
                "message": f"No inventory found for store: {store_name}"
            }
        
        # Create SKU lookup map
        sku_map = {item["sku"].upper().strip(): item for item in inventory_items}
        
        # Update customers collection with inventory data
        updated_count = 0
        customers = await db.customers.find(
            {"store_name": store_name, "order_skus": {"$exists": True, "$ne": []}},
            {"_id": 0, "customer_id": 1, "order_skus": 1, "line_items": 1}
        ).to_list(50000)
        
        for customer in customers:
            order_skus = customer.get("order_skus", [])
            line_items = customer.get("line_items", [])
            
            # Check if any SKU in this order is in inventory
            in_stock_skus = []
            out_of_stock_skus = []
            
            for sku in order_skus:
                sku_upper = sku.upper().strip()
                if sku_upper in sku_map:
                    in_stock_skus.append(sku)
                else:
                    out_of_stock_skus.append(sku)
            
            # Update customer record with inventory status
            await db.customers.update_one(
                {"customer_id": customer["customer_id"]},
                {
                    "$set": {
                        "in_stock_skus": in_stock_skus,
                        "out_of_stock_skus": out_of_stock_skus,
                        "inventory_synced": True,
                        "inventory_sync_date": datetime.now(timezone.utc)
                    }
                }
            )
            updated_count += 1
        
        return {
            "success": True,
            "message": f"Synced inventory for {store_name}",
            "inventory_items": len(inventory_items),
            "orders_updated": updated_count,
            "store_name": store_name
        }
        
    except Exception as e:
        logger.error(f"Error syncing inventory: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/inventory/v2/sync-shopify-prices")
async def sync_shopify_prices_to_inventory(store_name: str = None):
    """
    Sync sale prices from Shopify orders (line_items) to inventory.
    Updates sale_price and calculates profit for each inventory item.
    """
    try:
        query = {}
        if store_name and store_name != 'all':
            query['store_name'] = store_name
            
        # Get all inventory items
        inventory_items = await db.inventory_v2.find(query, {"_id": 0}).to_list(50000)
        
        if not inventory_items:
            return {"success": False, "message": "No inventory items found"}
        
        updated_count = 0
        skipped_count = 0
        
        for item in inventory_items:
            sku = item.get('sku', '').strip()
            if not sku:
                skipped_count += 1
                continue
            
            # Find orders with this SKU to get the sale price
            customer_query = {
                "line_items.sku": {"$regex": f"^{sku}$", "$options": "i"}
            }
            if store_name and store_name != 'all':
                customer_query["store_name"] = store_name
                
            customer = await db.customers.find_one(
                customer_query,
                {"_id": 0, "line_items": 1}
            )
            
            sale_price = item.get('sale_price', 0) or 0
            
            if customer and customer.get('line_items'):
                # Find the matching line item
                for line_item in customer['line_items']:
                    item_sku = line_item.get('sku', '').strip()
                    if item_sku.lower() == sku.lower():
                        sale_price = float(line_item.get('price', 0) or 0)
                        break
            
            # Calculate profit
            cost = float(item.get('cost', 0) or 0)
            profit = sale_price - cost
            
            # Update inventory item using SKU as identifier
            await db.inventory_v2.update_one(
                {"sku": sku, "store_name": item.get('store_name')},
                {"$set": {
                    "sale_price": sale_price,
                    "profit": profit,
                    "shopify_synced": True,
                    "last_price_sync": datetime.now(timezone.utc).isoformat()
                }}
            )
            updated_count += 1
        
        return {
            "success": True,
            "message": f"Synced prices for {updated_count} items",
            "updated_count": updated_count,
            "skipped_count": skipped_count,
            "total_items": len(inventory_items)
        }
        
    except Exception as e:
        logger.error(f"Error syncing Shopify prices: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/inventory/v2/import-shopify-products")
async def import_shopify_product_prices(store_name: str = None):
    """
    Import sale prices directly from Shopify Products API.
    Fetches all products from Shopify and updates inventory sale prices.
    """
    try:
        import httpx
        
        # Get store credentials
        if not store_name or store_name == 'all':
            return {"success": False, "message": "Please select a specific store"}
        
        store = await db.stores.find_one({"store_name": store_name}, {"_id": 0})
        if not store:
            return {"success": False, "message": f"Store '{store_name}' not found"}
        
        # Get Shopify credentials - check multiple possible field names
        shop_domain = store.get('shopify_domain') or store.get('shop_url') or store.get('store_url', '').replace('https://', '').replace('http://', '').rstrip('/')
        access_token = store.get('shopify_token') or store.get('shopify_access_token') or store.get('access_token')
        
        if not shop_domain or not access_token:
            return {"success": False, "message": "Shopify credentials not configured for this store"}
        
        logger.info(f"Fetching products from Shopify: {shop_domain}")
        
        # Fetch products from Shopify
        headers = {
            "X-Shopify-Access-Token": access_token,
            "Content-Type": "application/json"
        }
        
        all_products = []
        url = f"https://{shop_domain}/admin/api/2024-01/products.json?limit=250"
        
        async with httpx.AsyncClient(timeout=60.0) as client:
            while url:
                response = await client.get(url, headers=headers)
                if response.status_code != 200:
                    logger.error(f"Shopify API error: {response.text}")
                    return {"success": False, "message": f"Shopify API error: {response.status_code}"}
                
                data = response.json()
                all_products.extend(data.get('products', []))
                
                # Handle pagination
                link_header = response.headers.get('Link', '')
                url = None
                if 'rel="next"' in link_header:
                    for link in link_header.split(','):
                        if 'rel="next"' in link:
                            url = link.split(';')[0].strip('<> ')
                            break
        
        logger.info(f"Fetched {len(all_products)} products from Shopify for {store_name}")
        
        # Create SKU to price mapping from Shopify variants
        sku_price_map = {}
        for product in all_products:
            for variant in product.get('variants', []):
                sku = (variant.get('sku') or '').strip()
                if sku:
                    price = float(variant.get('price', 0) or 0)
                    compare_at_price = float(variant.get('compare_at_price', 0) or 0)
                    # Use price as sale price
                    sku_price_map[sku.lower()] = {
                        'sale_price': price,
                        'original_price': compare_at_price if compare_at_price > 0 else price,
                        'product_title': product.get('title', ''),
                        'variant_title': variant.get('title', '')
                    }
        
        logger.info(f"Built price map with {len(sku_price_map)} SKUs")
        
        # Update inventory items
        inventory_items = await db.inventory_v2.find(
            {"store_name": store_name},
            {"_id": 0}
        ).to_list(50000)
        
        updated_count = 0
        not_found_count = 0
        
        for item in inventory_items:
            sku = (item.get('sku') or '').strip().lower()
            if not sku:
                continue
            
            price_info = sku_price_map.get(sku)
            if price_info:
                cost = float(item.get('cost', 0) or 0)
                sale_price = price_info['sale_price']
                profit = sale_price - cost
                
                await db.inventory_v2.update_one(
                    {"id": item.get('id')},
                    {"$set": {
                        "sale_price": sale_price,
                        "original_price": price_info['original_price'],
                        "profit": profit,
                        "shopify_synced": True,
                        "shopify_product_title": price_info['product_title'],
                        "last_price_sync": datetime.now(timezone.utc).isoformat()
                    }}
                )
                updated_count += 1
            else:
                not_found_count += 1
        
        return {
            "success": True,
            "message": f"Imported prices for {updated_count} items from Shopify",
            "updated_count": updated_count,
            "not_found_in_shopify": not_found_count,
            "total_shopify_products": len(all_products),
            "total_shopify_skus": len(sku_price_map),
            "total_inventory_items": len(inventory_items)
        }
        
    except Exception as e:
        logger.error(f"Error importing Shopify product prices: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))



@api_router.post("/inventory/v2/sync-orders")
async def sync_inventory_orders(store_name: str = None):
    """
    Sync Order# column by matching inventory SKUs with orders.
    Finds orders containing each SKU and updates the order_number field.
    """
    try:
        if not store_name or store_name == 'all':
            return {"success": False, "message": "Please select a specific store"}
        
        # Get all inventory items
        inventory_items = await db.inventory_v2.find(
            {"store_name": store_name},
            {"_id": 0}
        ).to_list(50000)
        
        # Get all orders with line items
        orders = await db.customers.find(
            {"store_name": store_name, "line_items": {"$exists": True}},
            {"_id": 0, "name": 1, "line_items.sku": 1, "fulfillment_status": 1, "financial_status": 1}
        ).to_list(100000)
        
        # Build SKU to order mapping
        sku_order_map = {}
        for order in orders:
            order_name = order.get("name", "")
            for li in order.get("line_items", []):
                sku = (li.get("sku") or "").strip().lower()
                if sku:
                    if sku not in sku_order_map:
                        sku_order_map[sku] = []
                    sku_order_map[sku].append({
                        "order_number": order_name,
                        "fulfillment_status": order.get("fulfillment_status"),
                        "financial_status": order.get("financial_status")
                    })
        
        logger.info(f"Built order map with {len(sku_order_map)} SKUs from {len(orders)} orders")
        
        # Update inventory items with order numbers
        updated_count = 0
        for item in inventory_items:
            sku = (item.get("sku") or "").strip().lower()
            if not sku:
                continue
            
            order_info = sku_order_map.get(sku)
            if order_info:
                # Get the most recent order
                latest_order = order_info[-1]
                await db.inventory_v2.update_one(
                    {"id": item.get("id")},
                    {"$set": {
                        "order_number": latest_order["order_number"],
                        "order_fulfillment_status": latest_order["fulfillment_status"],
                        "order_financial_status": latest_order["financial_status"],
                        "total_orders": len(order_info),
                        "last_order_sync": datetime.now(timezone.utc).isoformat()
                    }}
                )
                updated_count += 1
        
        return {
            "success": True,
            "message": f"Synced orders for {updated_count} inventory items",
            "updated_count": updated_count,
            "total_orders": len(orders),
            "total_inventory_items": len(inventory_items)
        }
        
    except Exception as e:
        logger.error(f"Error syncing inventory orders: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/inventory/v2/sync-shopify-stock")
async def sync_shopify_inventory_stock(store_name: str = None):
    """
    Sync inventory stock data from Shopify.
    Fetches all products and variants from Shopify and imports/updates inventory.
    Includes: SKU, Title, Price, Stock Quantity, Variant details.
    """
    import httpx
    
    try:
        if not store_name or store_name == 'all':
            return {"success": False, "message": "Please select a specific store"}
        
        store = await db.stores.find_one({"store_name": store_name}, {"_id": 0})
        if not store:
            return {"success": False, "message": f"Store '{store_name}' not found"}
        
        # Get Shopify credentials
        shop_domain = store.get('shopify_domain') or store.get('shop_url') or store.get('store_url', '').replace('https://', '').replace('http://', '').rstrip('/')
        access_token = store.get('shopify_token') or store.get('shopify_access_token') or store.get('access_token')
        
        if not shop_domain or not access_token:
            return {"success": False, "message": "Shopify credentials not configured for this store"}
        
        logger.info(f"🔄 Syncing inventory from Shopify: {shop_domain}")
        
        headers = {
            "X-Shopify-Access-Token": access_token,
            "Content-Type": "application/json"
        }
        
        # Fetch all products from Shopify
        all_products = []
        url = f"https://{shop_domain}/admin/api/2024-01/products.json?limit=250"
        
        async with httpx.AsyncClient(timeout=60.0) as client:
            while url:
                response = await client.get(url, headers=headers)
                if response.status_code != 200:
                    logger.error(f"Shopify API error: {response.text}")
                    return {"success": False, "message": f"Shopify API error: {response.status_code}"}
                
                data = response.json()
                all_products.extend(data.get('products', []))
                
                # Handle pagination
                link_header = response.headers.get('Link', '')
                url = None
                if 'rel="next"' in link_header:
                    for link in link_header.split(','):
                        if 'rel="next"' in link:
                            url = link.split(';')[0].strip('<> ')
                            break
        
        logger.info(f"📦 Fetched {len(all_products)} products from Shopify")
        
        # Process products and variants into inventory items
        created_count = 0
        updated_count = 0
        skipped_count = 0
        
        for product in all_products:
            product_title = product.get('title', '')
            product_type = product.get('product_type', '')
            vendor = product.get('vendor', '')
            tags = product.get('tags', '')
            
            # Get image URL
            images = product.get('images', [])
            image_url = images[0].get('src') if images else None
            
            for variant in product.get('variants', []):
                sku = (variant.get('sku') or '').strip()
                if not sku:
                    skipped_count += 1
                    continue
                
                variant_title = variant.get('title', '')
                price = float(variant.get('price', 0) or 0)
                compare_at_price = float(variant.get('compare_at_price', 0) or 0)
                inventory_quantity = int(variant.get('inventory_quantity', 0) or 0)
                barcode = variant.get('barcode', '')
                weight = variant.get('weight', 0)
                
                # Extract size/color from variant title or options
                size = ''
                color = ''
                if variant_title and variant_title != 'Default Title':
                    parts = variant_title.split(' / ')
                    if len(parts) >= 1:
                        size = parts[0]
                    if len(parts) >= 2:
                        color = parts[1]
                
                # Check if item exists
                existing = await db.inventory_v2.find_one(
                    {"sku": sku, "store_name": store_name},
                    {"_id": 0, "id": 1, "cost": 1}
                )
                
                inventory_item = {
                    "sku": sku,
                    "product_name": product_title,
                    "variant_title": variant_title,
                    "size": size,
                    "color": color,
                    "sale_price": price,
                    "original_price": compare_at_price if compare_at_price > 0 else price,
                    "quantity": inventory_quantity,
                    "stock": inventory_quantity,
                    "barcode": barcode,
                    "weight": weight,
                    "product_type": product_type,
                    "vendor": vendor,
                    "tags": tags,
                    "image_url": image_url,
                    "store_name": store_name,
                    "shopify_synced": True,
                    "last_shopify_sync": datetime.now(timezone.utc).isoformat()
                }
                
                if existing:
                    # Preserve cost if it exists
                    if existing.get('cost'):
                        inventory_item['cost'] = existing['cost']
                        inventory_item['profit'] = price - existing['cost']
                    
                    await db.inventory_v2.update_one(
                        {"sku": sku, "store_name": store_name},
                        {"$set": inventory_item}
                    )
                    updated_count += 1
                else:
                    # New item - generate ID
                    inventory_item['id'] = f"{store_name}_{sku}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
                    inventory_item['cost'] = 0
                    inventory_item['profit'] = price
                    inventory_item['created_at'] = datetime.now(timezone.utc).isoformat()
                    
                    await db.inventory_v2.insert_one(inventory_item)
                    created_count += 1
        
        logger.info(f"✅ Inventory sync complete: {created_count} created, {updated_count} updated, {skipped_count} skipped")
        
        return {
            "success": True,
            "message": f"Synced {created_count + updated_count} inventory items from Shopify",
            "created_count": created_count,
            "updated_count": updated_count,
            "skipped_count": skipped_count,
            "total_products": len(all_products)
        }
        
    except Exception as e:
        logger.error(f"Error syncing Shopify inventory: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ========================================
# OLD INVENTORY ENDPOINTS (KEPT FOR BACKWARD COMPATIBILITY)
# ========================================

@api_router.get("/inventory")
async def get_inventory():
    """
    Get all inventory items with stats
    """
    try:
        inventory_items = await db.inventory.find({}, {"_id": 0}).sort("sku", 1).to_list(10000)
        stats = await inventory_manager.get_inventory_stats()
        
        return {
            "success": True,
            "inventory": inventory_items,
            "stats": stats
        }
    except Exception as e:
        logger.error(f"Error fetching inventory: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/inventory/upload")
async def upload_inventory_excel(file: UploadFile = File(...)):
    """
    Upload Excel file with inventory (SKU, Product Name, Opening Stock, Reorder Level)
    """
    try:
        import openpyxl
        from io import BytesIO
        
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        sheet = wb.active
        
        items_added = 0
        errors = []
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row[0]:  # Skip if SKU is empty
                    continue
                
                sku = str(row[0]).strip()
                product_name = str(row[1]).strip() if row[1] else "Unknown Product"
                opening_stock = int(row[2]) if row[2] else 0
                reorder_level = int(row[3]) if row[3] else 5
                
                result = await inventory_manager.set_opening_stock(sku, product_name, opening_stock, reorder_level)
                
                if result['success']:
                    items_added += 1
                else:
                    errors.append(f"Row {row_idx}: {result.get('error')}")
                    
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
                continue
        
        return {
            "success": True,
            "message": f"Processed {items_added} inventory items",
            "items_added": items_added,
            "errors": errors if errors else None
        }
        
    except Exception as e:
        logger.error(f"Error uploading inventory: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/inventory/upload/{store_name}")
async def upload_inventory_for_store(store_name: str, file: UploadFile = File(...)):
    """
    Upload Excel file with inventory for a specific store
    Expected columns: SKU, Size, color, cost, BOX NO
    Each row counts as 1 unit of quantity
    """
    try:
        import openpyxl
        from io import BytesIO
        
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        sheet = wb.active
        
        items_added = 0
        items_updated = 0
        errors = []
        
        # Track SKU counts (each row = 1 unit)
        sku_counts = {}
        sku_details = {}
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Expected columns: BOX NO, SKU, Size, color, cost
                if not row or len(row) < 2:
                    continue
                
                box_no = str(row[0]).strip() if row[0] else ""
                sku = str(row[1]).strip() if row[1] else ""
                size = str(row[2]).strip() if row[2] and len(row) > 2 else ""
                color = str(row[3]).strip() if row[3] and len(row) > 3 else ""
                cost = str(row[4]).strip() if row[4] and len(row) > 4 else ""
                
                if not sku:
                    continue
                
                sku_upper = sku.upper()
                
                # Count occurrences
                if sku_upper not in sku_counts:
                    sku_counts[sku_upper] = 0
                    sku_details[sku_upper] = {
                        "size": size,
                        "color": color,
                        "cost": cost,
                        "box_no": box_no
                    }
                
                sku_counts[sku_upper] += 1
                    
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
                continue
        
        # Now save to database
        for sku_upper, quantity in sku_counts.items():
            try:
                details = sku_details[sku_upper]
                
                existing = await db.inventory_items.find_one(
                    {"sku": sku_upper, "store_name": store_name},
                    {"_id": 0}
                )
                
                if existing:
                    # Update existing
                    await db.inventory_items.update_one(
                        {"sku": sku_upper, "store_name": store_name},
                        {"$inc": {"quantity": quantity},
                         "$set": {
                            "size": details["size"],
                            "color": details["color"],
                            "cost": details["cost"],
                            "box_no": details["box_no"],
                            "updated_at": datetime.now(timezone.utc).isoformat()
                         }}
                    )
                    items_updated += 1
                else:
                    # Create new
                    inventory_item = {
                        "sku": sku_upper,
                        "store_name": store_name,
                        "quantity": quantity,
                        "size": details["size"],
                        "color": details["color"],
                        "cost": details["cost"],
                        "box_no": details["box_no"],
                        "created_at": datetime.now(timezone.utc).isoformat(),
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }
                    await db.inventory_items.insert_one(inventory_item)
                    items_added += 1
                    
            except Exception as e:
                errors.append(f"SKU {sku_upper}: {str(e)}")
                continue
        
        return {
            "success": True,
            "message": f"Processed {items_added + items_updated} items for {store_name}",
            "items_added": items_added,
            "items_updated": items_updated,
            "errors": errors if errors else None
        }
        
    except Exception as e:
        logger.error(f"Error uploading inventory: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/inventory/items")
async def get_inventory_items(store_name: Optional[str] = None):
    """
    Get inventory items, optionally filtered by store
    """
    try:
        query = {}
        if store_name:
            query["store_name"] = store_name
        
        items = await db.inventory_items.find(query, {"_id": 0}).sort("sku", 1).to_list(10000)
        
        return {
            "success": True,
            "items": items
        }
    except Exception as e:
        logger.error(f"Error fetching inventory items: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/inventory/add")
async def add_inventory_item(
    sku: str,
    product_name: str,
    opening_stock: int = 0,
    reorder_level: int = 5
):
    """
    Add single inventory item manually
    """
    try:
        result = await inventory_manager.set_opening_stock(sku, product_name, opening_stock, reorder_level)
        
        if result['success']:
            return result
        else:
            raise HTTPException(status_code=400, detail=result.get('error'))
            
    except Exception as e:
        logger.error(f"Error adding inventory item: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/inventory/adjust")
async def adjust_inventory(
    sku: str,
    adjustment: int,
    reason: str = "Manual adjustment",
    user: str = "admin"
):
    """
    Manually adjust stock quantity
    """
    try:
        result = await inventory_manager.adjust_stock(sku, adjustment, reason, user)
        
        if result['success']:
            return result
        else:
            raise HTTPException(status_code=400, detail=result.get('error'))
            
    except Exception as e:
        logger.error(f"Error adjusting inventory: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/inventory/history/{sku}")
async def get_inventory_history(sku: str, limit: int = 50):
    """
    Get stock transaction history for a SKU
    """
    try:
        history = await inventory_manager.get_stock_history(sku, limit)
        
        return {
            "success": True,
            "sku": sku,
            "history": history
        }
    except Exception as e:
        logger.error(f"Error getting inventory history: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/inventory/deduct-on-delivery")
async def manual_stock_deduction(
    customer_id: str,
    store_name: str
):
    """
    Manually trigger stock deduction for a delivered order
    Used for testing or corrections
    """
    try:
        # Get customer order
        customer = await db.customers.find_one(
            {"customer_id": customer_id, "store_name": store_name},
            {"_id": 0}
        )
        
        if not customer:
            raise HTTPException(status_code=404, detail="Customer not found")
        
        if customer.get('delivery_status') != 'DELIVERED':
            raise HTTPException(status_code=400, detail="Order not yet delivered")
        
        # Check if already deducted
        if customer.get('stock_deducted'):
            return {
                "success": False,
                "message": "Stock already deducted for this order"
            }
        
        # Deduct stock
        result = await inventory_manager.deduct_stock_on_delivery(
            customer.get('order_skus', []),
            customer_id,
            customer.get('order_number', 'N/A')
        )
        
        # Mark as deducted
        if result['success']:
            await db.customers.update_one(
                {"customer_id": customer_id, "store_name": store_name},
                {"$set": {"stock_deducted": True, "stock_deducted_at": datetime.now(timezone.utc).isoformat()}}
            )
        
        return result
        
    except Exception as e:
        logger.error(f"Error in manual stock deduction: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/upload-orders-csv")
async def upload_shopify_orders_csv(file: UploadFile = File(...), store_name: str = "tnvcollectionpk"):
    """
    Upload Shopify orders CSV export - COMPLETE ORDER DATA including cancelled orders
    This properly imports: order_number, fulfillment_status, tracking_number, cancelled_at
    """
    try:
        from csv_orders_import import parse_shopify_orders_csv_full
        
        # Read CSV content
        content = await file.read()
        csv_text = content.decode('utf-8')
        
        logger.info(f"📄 Received orders CSV: {file.filename}, size: {len(csv_text)} bytes")
        
        # Parse CSV with full order details
        orders_list = parse_shopify_orders_csv_full(csv_text, store_name)
        
        logger.info(f"✅ Parsed {len(orders_list)} orders from CSV")
        
        # Upsert orders into database
        updated_count = 0
        inserted_count = 0
        
        for order in orders_list:
            # Find existing order by order_number
            existing = await db.customers.find_one({
                "order_number": order['order_number'],
                "store_name": store_name
            })
            
            if existing:
                # Update existing order
                await db.customers.update_one(
                    {"_id": existing["_id"]},
                    {"$set": order}
                )
                updated_count += 1
            else:
                # Insert new order
                await db.customers.insert_one(order)
                inserted_count += 1
        
        logger.info(f"✅ CSV Import: {inserted_count} new, {updated_count} updated")
        
        return {
            "success": True,
            "message": "Orders imported successfully",
            "inserted": inserted_count,
            "updated": updated_count,
            "total": len(orders_list)
        }
    
    except Exception as e:
        logger.error(f"❌ CSV upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/upload-csv")
async def upload_shopify_csv(file: UploadFile = File(...), store_name: str = "Default Store", shop_url: str = ""):
    """
    Upload Shopify orders CSV export and extract customer data (OLD - customer-level only)
    """
    try:
        # Read CSV content
        content = await file.read()
        csv_text = content.decode('utf-8')
        
        logger.info(f"Received CSV file: {file.filename}, size: {len(csv_text)} bytes")
        
        # Parse CSV
        customers_list = parse_shopify_orders_csv(csv_text)
        
        logger.info(f"Parsed {len(customers_list)} customers from CSV")
        
        # Add store name to each customer
        for customer in customers_list:
            customer['store_name'] = store_name
        
        # Create or update store record
        existing_store = await db.stores.find_one({"store_name": store_name})
        if not existing_store:
            store_obj = Store(
                store_name=store_name,
                shop_url=shop_url or f"{store_name.lower().replace(' ', '-')}.myshopify.com"
            )
            await db.stores.insert_one(store_obj.model_dump())
            logger.info(f"Created store record: {store_name}")
        
        # Merge strategy: Update existing customers or insert new ones
        updated_count = 0
        inserted_count = 0
        
        for customer in customers_list:
            # Try to find existing customer by customer_id or phone (if phone exists)
            query = {"store_name": store_name}
            if customer.get('phone'):
                # Use phone as unique identifier if available
                query["phone"] = customer['phone']
            else:
                # Use customer_id as fallback
                query["customer_id"] = customer['customer_id']
            
            existing = await db.customers.find_one(query)
            
            if existing:
                # Update existing customer - merge sizes and SKUs
                existing_sizes = set(existing.get('shoe_sizes', []))
                new_sizes = set(customer.get('shoe_sizes', []))
                merged_sizes = list(existing_sizes.union(new_sizes))
                
                # Remove "Unknown" if we have real sizes
                if len(merged_sizes) > 1 and "Unknown" in merged_sizes:
                    merged_sizes.remove("Unknown")
                
                # Merge SKUs
                existing_skus = set(existing.get('order_skus', []))
                new_skus = set(customer.get('order_skus', []))
                merged_skus = list(existing_skus.union(new_skus))
                
                update_data = {
                    "first_name": customer.get('first_name') or existing.get('first_name'),
                    "last_name": customer.get('last_name') or existing.get('last_name'),
                    "email": customer.get('email') or existing.get('email'),
                    "phone": customer.get('phone') or existing.get('phone'),
                    "country_code": customer.get('country_code') or existing.get('country_code'),
                    "shoe_sizes": merged_sizes,
                    "order_skus": merged_skus,
                    "order_count": existing.get('order_count', 0) + customer.get('order_count', 0),
                    "total_spent": existing.get('total_spent', 0) + customer.get('total_spent', 0),
                    "last_order_date": max(
                        customer.get('last_order_date') or '',
                        existing.get('last_order_date') or ''
                    ) or None
                }
                
                await db.customers.update_one(
                    {"_id": existing["_id"]},
                    {"$set": update_data}
                )
                updated_count += 1
            else:
                # Insert new customer
                await db.customers.insert_one(customer)
                inserted_count += 1
        
        logger.info(f"Updated {updated_count} existing customers, inserted {inserted_count} new customers")
        
        # Get total count for this store
        total_customers = await db.customers.count_documents({"store_name": store_name})
        
        return {
            "success": True,
            "message": f"Added {inserted_count} new, updated {updated_count} existing. Total: {total_customers} customers in {store_name}",
            "customers_imported": len(customers_list),
            "inserted": inserted_count,
            "updated": updated_count,
            "total": total_customers,
            "store_name": store_name
        }
        
    except Exception as e:
        logger.error(f"Error uploading CSV: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to process CSV: {str(e)}")


@api_router.post("/cache/refresh-inventory")
async def refresh_inventory_cache():
    """Manually refresh the inventory cache"""
    try:
        cache = await invalidate_inventory_cache()
        return {
            "success": True,
            "message": "Inventory cache refreshed",
            "items_cached": len(cache["data"]),
            "stock_skus_cached": len(cache["stock_skus"]),
            "last_updated": cache["last_updated"].isoformat() if cache["last_updated"] else None
        }
    except Exception as e:
        logger.error(f"Error refreshing inventory cache: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/cache/status")
async def get_cache_status():
    """Get current cache status"""
    global _inventory_cache
    return {
        "inventory_cache": {
            "items_count": len(_inventory_cache["data"]),
            "stock_skus_count": len(_inventory_cache["stock_skus"]),
            "last_updated": _inventory_cache["last_updated"].isoformat() if _inventory_cache["last_updated"] else None,
            "ttl_seconds": _inventory_cache["ttl_seconds"],
            "is_valid": _inventory_cache["last_updated"] is not None and 
                       (datetime.now(timezone.utc) - _inventory_cache["last_updated"]).total_seconds() < _inventory_cache["ttl_seconds"]
        }
    }


@api_router.get("/marketing/stats")
async def get_marketing_stats():
    """Get marketing dashboard statistics"""
    try:
        from datetime import datetime, timezone, timedelta
        
        now = datetime.now(timezone.utc)
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        week_start = now - timedelta(days=7)
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        # Get all customers
        all_customers = await db.customers.find({}, {"_id": 0}).to_list(100000)
        
        # Helper function to parse date strings
        def parse_date(date_str):
            if not date_str:
                return None
            try:
                if isinstance(date_str, datetime):
                    if date_str.tzinfo is None:
                        return date_str.replace(tzinfo=timezone.utc)
                    return date_str
                dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                return dt
            except:
                return None
        
        # Calculate today's revenue
        today_revenue = sum(
            c.get('total_spent', 0) 
            for c in all_customers 
            if parse_date(c.get('created_at')) and parse_date(c.get('created_at')) >= today_start
        )
        
        # Calculate week revenue
        week_revenue = sum(
            c.get('total_spent', 0)
            for c in all_customers
            if parse_date(c.get('created_at')) and parse_date(c.get('created_at')) >= week_start
        )
        
        # Calculate month revenue
        month_revenue = sum(
            c.get('total_spent', 0)
            for c in all_customers
            if parse_date(c.get('created_at')) and parse_date(c.get('created_at')) >= month_start
        )
        
        # Count orders
        total_orders = len(all_customers)
        pending_orders = len([c for c in all_customers if c.get('fulfillment_status', '').lower() in ['unfulfilled', 'pending']])
        
        # WhatsApp sent count
        whatsapp_sent = len([c for c in all_customers if c.get('messaged') == True])
        
        # Get inventory value
        inventory_items = await db.inventory_v2.find({}, {"_id": 0, "sale_price": 1, "cost": 1}).to_list(10000)
        inventory_value = sum(item.get('sale_price', item.get('cost', 0)) for item in inventory_items)
        
        # Calculate conversion rate (orders / unique visitors - simplified)
        conversion_rate = round((total_orders / max(total_orders * 3, 1)) * 100, 1) if total_orders > 0 else 0
        
        return {
            "success": True,
            "stats": {
                "todayRevenue": round(today_revenue, 2),
                "weekRevenue": round(week_revenue, 2),
                "monthRevenue": round(month_revenue, 2),
                "totalRevenue": round(sum(c.get('total_spent', 0) for c in all_customers), 2),
                "totalOrders": total_orders,
                "pendingOrders": pending_orders,
                "inventoryValue": round(inventory_value, 2),
                "whatsappSent": whatsapp_sent,
                "conversionRate": conversion_rate
            }
        }
    
    except Exception as e:
        logger.error(f"Error fetching marketing stats: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/campaigns")
async def get_campaigns():
    """Get all campaigns"""
    try:
        campaigns = await db.campaigns.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
        return {
            "success": True,
            "campaigns": campaigns
        }
    except Exception as e:
        logger.error(f"Error fetching campaigns: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/campaigns/create")
async def create_campaign(campaign_data: dict):
    """Create a new marketing campaign"""
    try:
        from uuid import uuid4
        
        campaign = {
            "id": str(uuid4()),
            "name": campaign_data.get("name"),
            "type": campaign_data.get("type", "discount"),
            "target": campaign_data.get("target", "all"),
            "discount_percentage": campaign_data.get("discount_percentage", 0),
            "start_date": campaign_data.get("start_date"),
            "end_date": campaign_data.get("end_date"),
            "status": campaign_data.get("status", "draft"),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.campaigns.insert_one(campaign)
        
        return {
            "success": True,
            "message": "Campaign created successfully",
            "campaign_id": campaign["id"]
        }
    except Exception as e:
        logger.error(f"Error creating campaign: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.put("/campaigns/{campaign_id}/status")
async def update_campaign_status(campaign_id: str, status_data: dict):
    """Update campaign status"""
    try:
        result = await db.campaigns.update_one(
            {"id": campaign_id},
            {"$set": {
                "status": status_data.get("status"),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Campaign not found")
        
        return {"success": True, "message": "Campaign status updated"}
    except Exception as e:
        logger.error(f"Error updating campaign status: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.delete("/campaigns/{campaign_id}")
async def delete_campaign(campaign_id: str):
    """Delete a campaign"""
    try:
        result = await db.campaigns.delete_one({"id": campaign_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Campaign not found")
        
        return {"success": True, "message": "Campaign deleted"}
    except Exception as e:
        logger.error(f"Error deleting campaign: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/marketing/campaigns")
async def get_marketing_campaigns():
    """Get active marketing campaigns"""
    try:
        # Placeholder - will be implemented with campaign feature
        campaigns = []
        
        return {
            "success": True,
            "campaigns": campaigns
        }
    
    except Exception as e:
        logger.error(f"Error fetching campaigns: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/inventory/v2/campaign-items")
async def get_campaign_items(tag: str = None):
    """Get inventory items for campaign management"""
    try:
        query = {}
        if tag and tag != 'all':
            query['campaign_tag'] = tag
        
        items = await db.inventory_v2.find(query, {"_id": 0}).to_list(10000)
        
        return {
            "success": True,
            "items": items,
            "total": len(items)
        }
    
    except Exception as e:
        logger.error(f"Error fetching campaign items: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/inventory/v2/bulk-tag")
async def bulk_tag_items(request: dict):
    """Apply tags to multiple inventory items"""
    try:
        skus = request.get('skus', [])
        tag = request.get('tag')
        
        if not skus or not tag:
            raise HTTPException(status_code=400, detail="SKUs and tag are required")
        
        # Update all items with the tag
        result = await db.inventory_v2.update_many(
            {"sku": {"$in": skus}},
            {"$set": {"campaign_tag": tag}}
        )
        
        return {
            "success": True,
            "updated_count": result.modified_count,
            "message": f"Tagged {result.modified_count} items as {tag}"
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error bulk tagging: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/inventory/v2/bulk-pricing")
async def bulk_pricing_update(request: dict):
    """Update pricing for multiple items"""
    try:
        skus = request.get('skus', [])
        discount_type = request.get('discount_type', 'percentage')
        discount_value = float(request.get('discount_value', 0))
        
        if not skus:
            raise HTTPException(status_code=400, detail="No SKUs provided")
        
        # Get current items
        items = await db.inventory_v2.find({"sku": {"$in": skus}}, {"_id": 0}).to_list(10000)
        
        updated_count = 0
        for item in items:
            current_price = item.get('sale_price', item.get('cost', 0))
            
            if discount_type == 'percentage':
                new_price = current_price * (1 - discount_value / 100)
                discount_pct = discount_value
            else:  # fixed
                new_price = max(0, current_price - discount_value)
                discount_pct = round((discount_value / current_price) * 100, 2) if current_price > 0 else 0
            
            # Update the item
            await db.inventory_v2.update_one(
                {"sku": item['sku']},
                {"$set": {
                    "sale_price": round(new_price, 2),
                    "discount_percentage": discount_pct,
                    "original_price": current_price
                }}
            )
            updated_count += 1
        
        return {
            "success": True,
            "updated_count": updated_count,
            "message": f"Updated pricing for {updated_count} items"
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error bulk pricing: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# Flash Sales Collection Helper
async def init_flash_sales_collection():
    """Initialize flash sales collection if it doesn't exist"""
    collections = await db.list_collection_names()
    if 'flash_sales' not in collections:
        await db.create_collection('flash_sales')

@api_router.get("/flash-sales")
async def get_flash_sales():
    """Get all flash sales"""
    try:
        sales = await db.flash_sales.find({}, {"_id": 0}).to_list(1000)
        
        return {
            "success": True,
            "sales": sales
        }
    
    except Exception as e:
        logger.error(f"Error fetching flash sales: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/flash-sales/create")
async def create_flash_sale(request: dict):
    """Create a new flash sale"""
    try:
        from datetime import datetime, timezone, timedelta
        import uuid
        
        name = request.get('name')
        collection = request.get('collection')
        discount = float(request.get('discount', 0))
        duration_hours = int(request.get('duration_hours', 24))
        start_time_str = request.get('start_time')
        
        # Parse start time
        if start_time_str:
            start_time = datetime.fromisoformat(start_time_str.replace('Z', '+00:00'))
        else:
            start_time = datetime.now(timezone.utc)
        
        end_time = start_time + timedelta(hours=duration_hours)
        
        flash_sale = {
            "id": str(uuid.uuid4()),
            "name": name,
            "collection": collection,
            "discount": discount,
            "duration_hours": duration_hours,
            "start_time": start_time.isoformat(),
            "end_time": end_time.isoformat(),
            "active": start_time <= datetime.now(timezone.utc),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "orders_count": 0,
            "revenue": 0,
            "conversion_rate": 0
        }
        
        await db.flash_sales.insert_one(flash_sale)
        
        # Update inventory items with flash sale tag
        if collection and collection != 'all':
            await db.inventory_v2.update_many(
                {"collection": {"$regex": collection, "$options": "i"}},
                {"$set": {"campaign_tag": "flash_sale", "flash_sale_discount": discount}}
            )
        
        return {
            "success": True,
            "message": "Flash sale created successfully",
            "sale_id": flash_sale['id']
        }
    
    except Exception as e:
        logger.error(f"Error creating flash sale: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/flash-sales/{sale_id}/toggle")
async def toggle_flash_sale(sale_id: str, request: dict):
    """Activate or pause a flash sale"""
    try:
        active = request.get('active', False)
        
        result = await db.flash_sales.update_one(
            {"id": sale_id},
            {"$set": {"active": active}}
        )
        
        if result.modified_count > 0:
            return {
                "success": True,
                "message": f"Flash sale {'activated' if active else 'paused'}"
            }
        else:
            raise HTTPException(status_code=404, detail="Flash sale not found")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error toggling flash sale: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.delete("/flash-sales/{sale_id}")
async def delete_flash_sale(sale_id: str):
    """Delete a flash sale"""
    try:
        result = await db.flash_sales.delete_one({"id": sale_id})
        
        if result.deleted_count > 0:
            return {
                "success": True,
                "message": "Flash sale deleted"
            }
        else:
            raise HTTPException(status_code=404, detail="Flash sale not found")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting flash sale: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting flash sale: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/bundles")
async def get_bundles():
    """Get all product bundles"""
    try:
        bundles = await db.bundles.find({}, {"_id": 0}).to_list(1000)
        
        return {
            "success": True,
            "bundles": bundles
        }
    
    except Exception as e:
        logger.error(f"Error fetching bundles: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/bundles/create")
async def create_bundle(request: dict):
    """Create a new product bundle"""
    try:
        import uuid
        
        bundle = {
            "id": str(uuid.uuid4()),
            "name": request.get('name'),
            "description": request.get('description'),
            "items": request.get('items', []),
            "discount": float(request.get('discount', 0)),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "orders": 0,
            "revenue": 0
        }
        
        await db.bundles.insert_one(bundle)
        
        # Tag items as bundle
        skus = [item['sku'] for item in bundle['items']]
        await db.inventory_v2.update_many(
            {"sku": {"$in": skus}},
            {"$set": {"campaign_tag": "bundle"}}
        )
        
        return {
            "success": True,
            "message": "Bundle created successfully",
            "bundle_id": bundle['id']
        }
    
    except Exception as e:
        logger.error(f"Error creating bundle: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.delete("/bundles/{bundle_id}")
async def delete_bundle(bundle_id: str):
    """Delete a bundle"""
    try:
        result = await db.bundles.delete_one({"id": bundle_id})
        
        if result.deleted_count > 0:
            return {
                "success": True,
                "message": "Bundle deleted"
            }
        else:
            raise HTTPException(status_code=404, detail="Bundle not found")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting bundle: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


        # Create CSV
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=['first_name', 'last_name', 'email', 'phone', 'total_spent', 'order_count'])
        writer.writeheader()
        
        for customer in customers:
            writer.writerow({
                'first_name': customer.get('first_name', ''),
                'last_name': customer.get('last_name', ''),
                'email': customer.get('email', ''),
                'phone': customer.get('phone', ''),
                'total_spent': customer.get('total_spent', 0),
                'order_count': customer.get('order_count', 1)
            })
        
        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=segment_{segment_type}.csv"}
        )
    
    except Exception as e:
        logger.error(f"Error exporting segment: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


        logger.error(f"Error bulk pricing: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/shoe-sizes")
async def get_shoe_sizes(store_name: Optional[str] = None):
    """
    Get all unique shoe sizes, optionally filtered by store
    """
    query = {}
    if store_name and store_name != "all":
        query['store_name'] = store_name
    
    customers = await db.customers.find(query, {"_id": 0, "shoe_sizes": 1}).to_list(20000)
    
    all_sizes = set()
    for customer in customers:
        for size in customer.get('shoe_sizes', []):
            all_sizes.add(size)
    
    return {"shoe_sizes": sorted(list(all_sizes))}


@api_router.get("/reports/agents")
async def get_agent_reports():
    """
    Get agent performance report with total sales
    """
    try:
        # Get all agents
        agents = await db.agents.find({}, {"_id": 0, "password": 0}).to_list(100)
        
        reports = []
        for agent in agents:
            username = agent["username"]
            
            # Count customers messaged by this agent
            messaged_count = await db.customers.count_documents({"messaged_by": username})
            
            # Count conversions
            converted_count = await db.customers.count_documents({
                "messaged_by": username,
                "converted": True
            })
            
            # Calculate total sales
            converted_customers = await db.customers.find({
                "messaged_by": username,
                "converted": True
            }, {"_id": 0, "sale_amount": 1}).to_list(10000)
            
            total_sales = sum(c.get('sale_amount', 0) or 0 for c in converted_customers)
            
            # Conversion rate
            conversion_rate = (converted_count / messaged_count * 100) if messaged_count > 0 else 0
            
            reports.append({
                "agent_username": username,
                "agent_name": agent["full_name"],
                "messages_sent": messaged_count,
                "conversions": converted_count,
                "conversion_rate": round(conversion_rate, 2),
                "total_sales": round(total_sales, 2)
            })
        
        return {"reports": reports}
    except Exception as e:
        logger.error(f"Error generating reports: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to generate reports")


@api_router.get("/reports/daily")
async def get_daily_reports(agent_username: Optional[str] = None):
    """
    Get day-wise reporting with messages, conversions, and sales
    """
    try:
        from datetime import datetime, timedelta
        from collections import defaultdict
        
        # Build query
        query = {"messaged": True}
        if agent_username and agent_username != "all":
            query["messaged_by"] = agent_username
        
        # Get all messaged customers
        customers = await db.customers.find(query, {
            "_id": 0,
            "last_messaged_at": 1,
            "converted": 1,
            "sale_amount": 1,
            "messaged_by": 1
        }).to_list(50000)
        
        # Group by date
        daily_stats = defaultdict(lambda: {
            "messages": 0,
            "conversions": 0,
            "sales": 0.0
        })
        
        for customer in customers:
            if customer.get('last_messaged_at'):
                # Parse date (YYYY-MM-DD)
                msg_date = customer['last_messaged_at'][:10]
                
                daily_stats[msg_date]["messages"] += 1
                
                if customer.get('converted'):
                    daily_stats[msg_date]["conversions"] += 1
                    daily_stats[msg_date]["sales"] += customer.get('sale_amount', 0) or 0
        
        # Convert to list and sort by date
        reports = []
        for date, stats in sorted(daily_stats.items(), reverse=True):
            conversion_rate = (stats["conversions"] / stats["messages"] * 100) if stats["messages"] > 0 else 0
            reports.append({
                "date": date,
                "messages_sent": stats["messages"],
                "conversions": stats["conversions"],
                "conversion_rate": round(conversion_rate, 2),
                "total_sales": round(stats["sales"], 2)
            })
        
        return {"daily_reports": reports}
    except Exception as e:
        logger.error(f"Error generating daily reports: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to generate daily reports")


@api_router.get("/countries")
async def get_countries(store_name: Optional[str] = None):
    """
    Get all unique countries, optionally filtered by store
    """
    query = {}
    if store_name and store_name != "all":
        query['store_name'] = store_name
    
    customers = await db.customers.find(query, {"_id": 0, "country_code": 1}).to_list(20000)
    
    countries = set()
    for customer in customers:
        country = customer.get('country_code')
        if country:
            countries.add(country)
    
    return {"countries": sorted(list(countries))}


@api_router.post("/stores", response_model=Store)
async def create_store(store: StoreCreate):
    """
    Add a new store
    """
    try:
        # Check if store already exists
        existing = await db.stores.find_one({"store_name": store.store_name})
        if existing:
            raise HTTPException(status_code=400, detail="Store with this name already exists")
        
        store_data = store.model_dump()
        # Auto-set shopify_domain from shop_url if not provided
        if not store_data.get('shopify_domain') and store_data.get('shop_url'):
            store_data['shopify_domain'] = store_data['shop_url']
        
        store_obj = Store(**store_data)
        doc = store_obj.model_dump()
        await db.stores.insert_one(doc)
        logger.info(f"Created store: {store.store_name}")
        return store_obj
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating store: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to create store: {str(e)}")


@api_router.get("/stores", response_model=List[Store])
async def get_stores():
    """
    Get all stores
    """
    stores = await db.stores.find({}, {"_id": 0}).to_list(100)
    return stores


@api_router.put("/stores/{store_name}/token")
async def update_store_token(store_name: str, token: str):
    """
    Update store's Shopify API token
    """
    try:
        result = await db.stores.update_one(
            {"store_name": store_name},
            {"$set": {"shopify_token": token}}
        )
        if result.modified_count == 0:
            raise HTTPException(status_code=404, detail="Store not found")
        logger.info(f"Updated token for store: {store_name}")
        return {"success": True, "message": f"Token updated for {store_name}"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating store token: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.put("/stores/{store_name}/domain")
async def update_store_domain(store_name: str, domain: str):
    """
    Update store's Shopify domain
    """
    try:
        result = await db.stores.update_one(
            {"store_name": store_name},
            {"$set": {"shopify_domain": domain}}
        )
        if result.modified_count == 0:
            raise HTTPException(status_code=404, detail="Store not found")
        logger.info(f"Updated domain for store: {store_name}")
        return {"success": True, "message": f"Domain updated for {store_name}"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating store domain: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.delete("/stores/{store_id}")
async def delete_store(store_id: str):
    """
    Delete a store and all its customers
    """
    try:
        # Get store first
        store = await db.stores.find_one({"id": store_id})
        if not store:
            raise HTTPException(status_code=404, detail="Store not found")
        
        store_name = store["store_name"]
        
        # Delete all customers from this store
        customers_result = await db.customers.delete_many({"store_name": store_name})
        
        # Delete the store
        await db.stores.delete_one({"id": store_id})
        
        logger.info(f"Deleted store '{store_name}' and {customers_result.deleted_count} customers")
        
        return {
            "success": True,
            "message": f"Deleted store '{store_name}' and {customers_result.deleted_count} customers"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting store: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to delete store: {str(e)}")


@api_router.get("/message-greetings")
async def get_random_greeting():
    """
    Get a random greeting message to avoid repetition
    """
    greetings = [
        "Hello",
        "Hi",
        "Hey",
        "Hello there",
        "Hi there",
        "Hey there",
        "Greetings",
        "Good day",
        "Hello!",
        "Hi!",
        "Hey!",
        "Salaam",
        "Assalam o Alaikum",
        "Hope you're doing well",
        "How are you",
        "Dear Customer"
    ]
    
    import random
    return {"greeting": random.choice(greetings)}


@api_router.post("/whatsapp-link")
async def generate_whatsapp_link(request: WhatsAppRequest):
    """
    Generate WhatsApp link for a phone number
    """
    # Clean phone number
    cleaned_phone = ''.join(filter(str.isdigit, request.phone))
    
    # If no country code and phone doesn't start with +, you might need to add it
    if request.country_code and not cleaned_phone.startswith(request.country_code.replace('+', '')):
        # Get country dial code mapping
        country_dial_codes = {
            'US': '1', 'IN': '91', 'GB': '44', 'AE': '971', 'SA': '966',
            'CA': '1', 'AU': '61', 'PK': '92', 'BD': '880'
        }
        dial_code = country_dial_codes.get(request.country_code, '')
        if dial_code:
            cleaned_phone = dial_code + cleaned_phone
    
    return {
        "whatsapp_link": f"https://wa.me/{cleaned_phone}",
        "whatsapp_web": f"https://api.whatsapp.com/send?phone={cleaned_phone}"
    }


# Include WhatsApp CRM routes (NEW comprehensive WhatsApp system)
try:
    from whatsapp_crm_routes import whatsapp_router as whatsapp_crm_router
    app.include_router(whatsapp_crm_router)
    print("✅ WhatsApp CRM routes loaded")
except Exception as e:
    print(f"⚠️ Could not load WhatsApp CRM routes: {str(e)}")

# Include legacy WhatsApp routes (OLD system - for backward compatibility)
try:
    from routes.whatsapp import router as whatsapp_legacy_router
    api_router.include_router(whatsapp_legacy_router)
    print("✅ Legacy WhatsApp routes loaded")
except Exception as e:
    print(f"⚠️ Could not load legacy WhatsApp routes: {str(e)}")


@api_router.get("/dashboard/stats")
async def get_dashboard_stats(store_name: str = None):
    """Get dashboard statistics (optimized)"""
    try:
        query = {}
        if store_name and store_name != "all":
            query["store_name"] = store_name
        
        # Get fulfillment status counts
        fulfillment_pipeline = [
            {"$match": query},
            {"$group": {
                "_id": "$fulfillment_status",
                "count": {"$sum": 1}
            }}
        ]
        
        fulfillment_results = await db.customers.aggregate(fulfillment_pipeline).to_list(100)
        fulfillment_status = {item["_id"]: item["count"] for item in fulfillment_results if item["_id"]}
        
        # Get payment status counts
        payment_pipeline = [
            {"$match": query},
            {"$group": {
                "_id": "$payment_status",
                "count": {"$sum": 1}
            }}
        ]
        
        payment_results = await db.customers.aggregate(payment_pipeline).to_list(100)
        payment_status = {item["_id"]: item["count"] for item in payment_results if item["_id"]}
        
        # Get delivery status counts  
        delivery_pipeline = [
            {"$match": query},
            {"$group": {
                "_id": "$delivery_status",
                "count": {"$sum": 1}
            }}
        ]
        
        delivery_results = await db.customers.aggregate(delivery_pipeline).to_list(100)
        delivery_status = {item["_id"]: item["count"] for item in delivery_results if item["_id"]}
        
        return {
            "fulfillmentStatus": fulfillment_status,
            "paymentStatus": payment_status,
            "deliveryStatus": delivery_status
        }
        
    except Exception as e:
        logger.error(f"Error fetching dashboard stats: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))



# ============= WhatsApp Business API Endpoints =============

class WhatsAppMessage(BaseModel):
    phone: str
    message: str

class BulkWhatsAppMessage(BaseModel):
    recipients: List[Dict]



@api_router.get("/whatsapp/conversations")
async def get_whatsapp_conversations():
    """
    Get all WhatsApp conversations with customers
    Returns list of conversations with last message and unread count
    """
    try:
        # Get all customers who have WhatsApp messages
        pipeline = [
            {
                '$match': {
                    'whatsapp_messages': {'$exists': True, '$ne': []}
                }
            },
            {
                '$project': {
                    '_id': 0,
                    'phone': '$phone',
                    'customer_name': {'$concat': ['$first_name', ' ', '$last_name']},
                    'first_name': 1,
                    'last_name': 1,
                    'last_message': {'$arrayElemAt': ['$whatsapp_messages.message', -1]},
                    'last_message_time': {'$arrayElemAt': ['$whatsapp_messages.timestamp', -1]},
                    'messages': '$whatsapp_messages'
                }
            },
            {
                '$addFields': {
                    'unread_count': 0  # Can be enhanced with actual unread tracking
                }
            },
            {
                '$sort': {'last_message_time': -1}
            }
        ]
        
        conversations = await db.customers.aggregate(pipeline).to_list(1000)
        
        return {
            'success': True,
            'conversations': conversations,
            'total': len(conversations)
        }
        
    except Exception as e:
        logger.error(f"Error fetching WhatsApp conversations: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/whatsapp/send")
async def send_whatsapp_message(data: WhatsAppMessage):
    """Send a WhatsApp message to a customer"""
    try:
        result = await whatsapp_service.send_text_message(data.phone, data.message)
        return result
    except Exception as e:
        logger.error(f"Error sending WhatsApp message: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/whatsapp/send-bulk")
async def send_bulk_whatsapp(data: BulkWhatsAppMessage):
    """Send WhatsApp messages to multiple recipients"""
    try:
        result = await whatsapp_service.send_bulk_messages(data.recipients)
        return result
    except Exception as e:
        logger.error(f"Error sending bulk WhatsApp: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/whatsapp/templates")
async def get_whatsapp_templates():
    """Get list of approved message templates"""
    try:
        result = await whatsapp_service.get_message_templates()
        return result
    except Exception as e:
        logger.error(f"Error fetching templates: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/whatsapp/messages/{customer_id}")
async def get_customer_whatsapp_messages(customer_id: str):
    """Get WhatsApp message history for a customer"""
    try:
        messages = await db.whatsapp_messages.find(
            {"$or": [{"from": customer_id}, {"to": customer_id}]}
        ).sort("timestamp", -1).limit(50).to_list(50)
        
        return {
            "success": True,
            "messages": messages
        }
    except Exception as e:
        logger.error(f"Error fetching all messages: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/whatsapp/messages-by-phone/{phone}")
async def get_whatsapp_messages_by_phone(phone: str):
    """
    Get WhatsApp message history for a customer by phone number
    """
    try:
        # Clean phone number
        phone_clean = phone.replace('+', '').replace(' ', '').replace('-', '')
        
        # Get customer messages from customer record
        customer = await db.customers.find_one(
            {'phone': {'$regex': phone_clean}},
            {'_id': 0, 'whatsapp_messages': 1, 'first_name': 1, 'last_name': 1}
        )
        
        if not customer:
            return {
                'success': True,
                'messages': [],
                'customer_name': 'Unknown'
            }
        
        messages = customer.get('whatsapp_messages', [])
        
        # Sort by timestamp
        messages.sort(key=lambda x: x.get('timestamp', ''), reverse=False)
        
        return {
            'success': True,
            'messages': messages,
            'customer_name': f"{customer.get('first_name', '')} {customer.get('last_name', '')}".strip()
        }
        
    except Exception as e:
        logger.error(f"Error fetching messages by phone: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))



@api_router.post("/whatsapp/templates/create")
async def create_template(data: dict):
    """
    Create a new WhatsApp message template and submit for approval
    """
    try:
        name = data.get("name")
        category = data.get("category")
        language = data.get("language", "en")
        body_text = data.get("body")
        header_text = data.get("header")
        footer_text = data.get("footer")
        buttons = data.get("buttons")
        
        # Validate required fields
        if not name or not category or not body_text:
            raise HTTPException(status_code=400, detail="Name, category, and body are required")
        
        # Validate template name format (lowercase, underscores only)
        if not name.replace("_", "").isalnum() or name != name.lower():
            raise HTTPException(
                status_code=400, 
                detail="Template name must be lowercase with underscores only (e.g., order_confirmation)"
            )
        
        # Create template via WhatsApp API
        result = await whatsapp_service.create_message_template(
            name=name,
            category=category,
            language=language,
            body_text=body_text,
            header_text=header_text,
            footer_text=footer_text,
            buttons=buttons
        )
        
        # Store template creation in database for tracking
        if result.get("success"):
            await db.whatsapp_template_submissions.insert_one({
                "template_id": result.get("template_id"),
                "name": name,
                "category": category,
                "language": language,
                "status": result.get("status", "PENDING"),
                "body": body_text,
                "header": header_text,
                "footer": footer_text,
                "submitted_at": datetime.now(timezone.utc).isoformat(),
                "submitted_by": "dashboard"
            })
        
        return result
        
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Error creating template: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/whatsapp/templates/{template_name}")
async def delete_template(template_name: str):
    """Delete a WhatsApp message template"""
    try:
        result = await whatsapp_service.delete_message_template(template_name)
        return result
    except Exception as e:
        logger.error(f"Error deleting template: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

        logger.error(f"Error fetching messages: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/whatsapp/messages/all")
async def get_all_whatsapp_messages(limit: int = 100):
    """Get all WhatsApp message history"""
    try:
        messages = await db.whatsapp_messages.find({}, {"_id": 0}).sort("timestamp", -1).limit(limit).to_list(limit)
        
        return {
            "success": True,
            "messages": messages
        }
    except Exception as e:
        logger.error(f"Error fetching all messages: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/whatsapp/send-template")
async def send_template_whatsapp(data: dict):
    """Send a WhatsApp template message"""
    try:
        phone = data.get("phone")
        template_name = data.get("template_name")
        language = data.get("language", "en")
        variables = data.get("variables", [])
        
        # Build components for template
        components = []
        if variables:
            components.append({
                "type": "body",
                "parameters": [{"type": "text", "text": str(var)} for var in variables]
            })
        
        result = await whatsapp_service.send_template_message(
            to=phone,
            template_name=template_name,
            language_code=language,
            components=components if components else None
        )
        return result
    except Exception as e:
        logger.error(f"Error sending template: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/whatsapp/notify-order-status")
async def notify_order_status_change(data: dict):
    """Send automatic WhatsApp notification when order status changes"""
    try:
        customer_id = data.get("customer_id")
        store_name = data.get("store_name")
        new_status = data.get("new_status")
        
        # Get customer details
        customer = await db.customers.find_one(
            {"customer_id": customer_id, "store_name": store_name},
            {"_id": 0}
        )
        
        if not customer or not customer.get("phone"):
            return {"success": False, "error": "Customer not found or no phone number"}
        
        # Create status-specific message
        message = generate_status_message(customer, new_status)
        
        # Send WhatsApp
        result = await whatsapp_service.send_text_message(customer["phone"], message)
        
        # Log notification
        if result.get("success"):
            await db.whatsapp_messages.insert_one({
                "customer_id": customer_id,
                "to": customer["phone"],
                "message": message,
                "message_id": result.get("message_id"),
                "type": "status_notification",
                "status": new_status,
                "direction": "outgoing",
                "timestamp": datetime.now(timezone.utc).isoformat()
            })
        
        return result
    except Exception as e:
        logger.error(f"Error sending notification: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# ============= WhatsApp Marketing API Endpoints =============

@api_router.post("/whatsapp/marketing/send")
async def send_marketing_message(data: dict):
    """
    Send marketing template message (promotional content)
    Uses AI-optimized delivery for better engagement
    """
    try:
        phone = data.get("phone")
        template_name = data.get("template_name")
        language = data.get("language", "en")
        variables = data.get("variables", [])
        campaign_id = data.get("campaign_id")
        
        # Build components
        components = []
        if variables:
            components.append({
                "type": "body",
                "parameters": [{"type": "text", "text": str(var)} for var in variables]
            })
        
        result = await whatsapp_marketing.send_marketing_template(
            to=phone,
            template_name=template_name,
            language_code=language,
            components=components if components else None,
            campaign_id=campaign_id
        )
        
        # Log marketing message
        if result.get("success"):
            await db.whatsapp_messages.insert_one({
                "message_id": result.get("message_id"),
                "to": phone,
                "type": "marketing",
                "template_name": template_name,
                "campaign_id": campaign_id,
                "direction": "outgoing",
                "timestamp": datetime.now(timezone.utc).isoformat()
            })
        
        return result
    except Exception as e:
        logger.error(f"Error sending marketing message: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/whatsapp/marketing/campaign")
async def send_marketing_campaign(data: dict):
    """
    Send marketing campaign to multiple recipients
    Features:
    - AI-optimized delivery
    - Smart targeting
    - Performance tracking
    """
    try:
        recipient_list = data.get("recipients", [])
        template_name = data.get("template_name")
        language = data.get("language", "en")
        variables = data.get("variables", [])
        campaign_name = data.get("campaign_name")
        
        if not recipient_list:
            raise HTTPException(status_code=400, detail="Recipients list is required")
        
        # Build components
        components = []
        if variables:
            components.append({
                "type": "body",
                "parameters": [{"type": "text", "text": str(var)} for var in variables]
            })
        
        result = await whatsapp_marketing.send_marketing_campaign(
            recipient_list=recipient_list,
            template_name=template_name,
            language_code=language,
            components=components if components else None,
            campaign_name=campaign_name
        )
        
        # Store campaign results
        await db.whatsapp_campaigns.insert_one({
            "campaign_id": result.get("campaign_id"),
            "campaign_name": campaign_name,
            "template_name": template_name,
            "total_recipients": result.get("total_recipients"),
            "sent": result.get("sent"),
            "failed": result.get("failed"),
            "success_rate": result.get("success_rate"),
            "started_at": result.get("started_at"),
            "completed_at": result.get("completed_at"),
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        
        return result
    except Exception as e:
        logger.error(f"Error sending campaign: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/whatsapp/marketing/templates")
async def get_marketing_templates():
    """Get recommended marketing templates"""
    try:
        return {
            "success": True,
            "templates": MARKETING_TEMPLATES
        }
    except Exception as e:
        logger.error(f"Error fetching marketing templates: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/whatsapp/campaigns")
async def get_campaigns(limit: int = 50):
    """Get marketing campaign history"""
    try:
        campaigns = await db.whatsapp_campaigns.find({}, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
        return {
            "success": True,
            "campaigns": campaigns
        }
    except Exception as e:
        logger.error(f"Error fetching campaigns: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/whatsapp/campaigns/{campaign_id}/analytics")
async def get_campaign_analytics(campaign_id: str):
    """Get campaign performance analytics"""
    try:
        result = await whatsapp_marketing.get_campaign_analytics(campaign_id)
        return result
    except Exception as e:
        logger.error(f"Error fetching analytics: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

        
        if not customer or not customer.get("phone"):
            return {"success": False, "error": "Customer not found or no phone number"}
        
        # Create status-specific message
        message = generate_status_message(customer, new_status)
        
        # Send WhatsApp
        result = await whatsapp_service.send_text_message(customer["phone"], message)
        
        # Log notification
        if result.get("success"):
            await db.whatsapp_messages.insert_one({
                "customer_id": customer_id,
                "to": customer["phone"],
                "message": message,
                "message_id": result.get("message_id"),
                "type": "status_notification",
                "status": new_status,
                "direction": "outgoing",
                "timestamp": datetime.now(timezone.utc).isoformat()
            })
        
        return result
    except Exception as e:
        logger.error(f"Error sending notification: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

def generate_status_message(customer: dict, status: str) -> str:
    """Generate appropriate message based on order status"""
    name = customer.get("first_name", "Customer")
    order_num = customer.get("order_number", "N/A")
    tracking = customer.get("tracking_number", "N/A")
    
    messages = {
        "DELIVERED": f"Hi {name}! 🎉\n\nYour order #{order_num} has been delivered successfully!\n\nThank you for shopping with TNV Collection. We hope you love your purchase!\n\nHave questions? Feel free to reply!",
        
        "IN_TRANSIT": f"Hello {name}! 📦\n\nGood news! Your order #{order_num} is now in transit.\n\nTracking: {tracking}\n\nExpected delivery soon. We'll keep you updated!\n\nTNV Collection",
        
        "OUT_FOR_DELIVERY": f"Hi {name}! 🚚\n\nYour order #{order_num} is out for delivery today!\n\nTracking: {tracking}\n\nPlease keep your phone handy. Our delivery person will contact you shortly.\n\nTNV Collection",
        
        "RETURN_IN_PROCESS": f"Hello {name},\n\nYour order #{order_num} is being returned.\n\nTracking: {tracking}\n\nIf you have any concerns, please contact us.\n\nTNV Collection",
        
        "WAITING_TO_BE_PICKED_UP": f"Hi {name}! 📋\n\nYour order #{order_num} is ready and waiting to be picked up by our courier.\n\nTracking: {tracking}\n\nWe'll notify you once it's dispatched!\n\nTNV Collection"
    }
    
    return messages.get(status, f"Hi {name},\n\nOrder #{order_num} status update:\nStatus: {status}\nTracking: {tracking}\n\nTNV Collection")

# ========================================
# AUTO-SYNC SETTINGS ENDPOINTS
# ========================================

class AutoSyncSettings(BaseModel):
    enabled: bool = False
    interval_minutes: int = 60
    sync_shopify: bool = True
    sync_tcs: bool = True
    sync_dtdc: bool = True
    sync_inventory: bool = True
    last_sync: Optional[str] = None

@api_router.get("/settings/auto-sync")
async def get_auto_sync_settings():
    """Get auto-sync configuration"""
    try:
        settings = await db.system_settings.find_one(
            {"setting_type": "auto_sync"},
            {"_id": 0}
        )
        
        if not settings:
            # Return default settings
            return {
                "enabled": False,
                "interval_minutes": 60,
                "sync_shopify": True,
                "sync_tcs": True,
                "sync_dtdc": True,
                "sync_inventory": True,
                "last_sync": None
            }
        
        return settings
        
    except Exception as e:
        logger.error(f"Error fetching auto-sync settings: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/settings/auto-sync")
async def update_auto_sync_settings(settings: AutoSyncSettings):
    """Update auto-sync configuration"""
    try:
        settings_dict = settings.model_dump()
        settings_dict["setting_type"] = "auto_sync"
        settings_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
        
        await db.system_settings.update_one(
            {"setting_type": "auto_sync"},
            {"$set": settings_dict},
            upsert=True
        )
        
        # If enabled, trigger an immediate sync
        if settings.enabled:
            # Trigger background sync
            logger.info(f"Auto-sync enabled with {settings.interval_minutes} minute interval")
        
        return {
            "success": True,
            "message": "Auto-sync settings updated",
            "settings": settings_dict
        }
        
    except Exception as e:
        logger.error(f"Error updating auto-sync settings: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/settings/auto-sync/trigger")
async def trigger_manual_sync():
    """Manually trigger a sync of all stores"""
    try:
        settings = await db.system_settings.find_one({"setting_type": "auto_sync"})
        
        if not settings:
            raise HTTPException(status_code=400, detail="Auto-sync not configured")
        
        # Get all stores
        stores = await db.stores.find({}, {"_id": 0, "store_name": 1}).to_list(100)
        
        sync_results = []
        
        for store in stores:
            store_name = store["store_name"]
            
            try:
                # Sync Shopify orders
                if settings.get("sync_shopify", True):
                    logger.info(f"Syncing Shopify orders for {store_name}...")
                    # Call Shopify sync endpoint internally
                    # This will be done via the existing sync endpoint
                
                # Sync inventory
                if settings.get("sync_inventory", True):
                    logger.info(f"Syncing inventory for {store_name}...")
                    # Call inventory sync function
                    sync_result = await sync_inventory_with_store(store_name)
                    sync_results.append(sync_result)
                
            except Exception as e:
                logger.error(f"Error syncing {store_name}: {str(e)}")
                sync_results.append({
                    "store_name": store_name,
                    "success": False,
                    "error": str(e)
                })
        
        # Update last sync time
        await db.system_settings.update_one(
            {"setting_type": "auto_sync"},
            {"$set": {"last_sync": datetime.now(timezone.utc).isoformat()}}
        )
        
        return {
            "success": True,
            "message": "Manual sync completed",
            "stores_synced": len(sync_results),
            "results": sync_results
        }
        
    except Exception as e:
        logger.error(f"Error triggering manual sync: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))



# Include the new modular routers BEFORE adding api_router to app
api_router.include_router(facebook_router)
api_router.include_router(finance_router)
api_router.include_router(pricing_router)
api_router.include_router(tcs_router)
api_router.include_router(customers_router)
api_router.include_router(clearance_router)
api_router.include_router(users_router)
api_router.include_router(tenants_router)
api_router.include_router(meta_ads_router)
api_router.include_router(whatsapp_embedded_router)

# Include the router in the main app
app.include_router(api_router)
app.include_router(whatsapp_webhook_router)
app.include_router(tracking_router)
app.include_router(subscriptions_router, prefix="/api")
app.include_router(shopify_webhooks_router, prefix="/api")
app.include_router(api_keys_router)
app.include_router(shopify_oauth_router)
app.include_router(lead_ads_router)
app.include_router(super_admin_router)
app.include_router(dwz56_router)
app.include_router(alibaba_1688_router)
app.include_router(fulfillment_router)
app.include_router(product_scraper_router)
app.include_router(tmapi_buyer_router)
app.include_router(fulfillment_webhooks_router)
app.include_router(whatsapp_notifications_router)
app.include_router(fulfillment_pipeline_router)
app.include_router(shopify_fulfillment_sync_router)
app.include_router(dwz56_sync_router)
app.include_router(email_notification_router)
app.include_router(storefront_router)
app.include_router(storefront_cms_router)
app.include_router(warehouse_router)
app.include_router(sync_service_router)
app.include_router(shopify_sync_router)
app.include_router(whatsapp_api_router)
app.include_router(marketing_router)
app.include_router(settings_router)
# Note: whatsapp_crm_router is now included via api_router (line 5667)

# ==================== Meta WhatsApp Embedded Signup Webhook (Root Level) ====================
from fastapi.responses import PlainTextResponse

@app.get("/webhook/whatsapp-business")
async def meta_whatsapp_webhook_verify(request: Request):
    """Root-level webhook verification for Meta WhatsApp Embedded Signup"""
    mode = request.query_params.get("hub.mode")
    token = request.query_params.get("hub.verify_token")
    challenge = request.query_params.get("hub.challenge")
    
    verify_token = "omnisales_whatsapp_webhook"
    
    print(f"[WEBHOOK VERIFY] mode={mode}, token={token}, challenge={challenge}")
    
    if mode == "subscribe" and token == verify_token:
        print("[WEBHOOK VERIFY] Success!")
        return PlainTextResponse(content=challenge, status_code=200)
    
    print(f"[WEBHOOK VERIFY] Failed - expected token: {verify_token}")
    return PlainTextResponse(content="Verification failed", status_code=403)

@app.post("/webhook/whatsapp-business")
async def meta_whatsapp_webhook_handler(request: Request):
    """Root-level webhook handler for Meta WhatsApp events"""
    try:
        data = await request.json()
        print(f"[WEBHOOK EVENT] Received: {data}")
        return {"status": "ok"}
    except Exception as e:
        print(f"[WEBHOOK ERROR] {str(e)}")
        return {"status": "error"}

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Add middleware to prevent caching
class NoCacheMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        return response

app.add_middleware(NoCacheMiddleware)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup_auto_sync():
    """Start automatic TCS sync on server startup"""
    try:
        from auto_tcs_sync import AutoTCSSync
        
        # Get TCS config
        config = await db.tcs_config.find_one({'service': 'tcs_pakistan'}, {'_id': 0})
        
        if config:
            logger.info("🚀 Starting automatic TCS sync service...")
            auto_sync = AutoTCSSync(db, config)
            
            # Run in background
            asyncio.create_task(auto_sync.run_continuous_sync(
                batch_size=20,           # Process 20 orders at a time
                delay_between_orders=3,   # 3 seconds between each order
                delay_between_batches=60  # 1 minute between batches
            ))
            
            logger.info("✅ Automatic TCS sync service started successfully")
        else:
            logger.warning("⚠️ TCS not configured - automatic sync disabled")
            
    except Exception as e:
        logger.error(f"❌ Failed to start auto sync: {str(e)}")


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

# Browser Extension Download
import zipfile

@app.get("/api/extension/download")
async def download_extension():
    """Download the WaMerce 1688 browser extension"""
    extension_path = Path("/app/wamerce-1688-extension.zip")
    if not extension_path.exists():
        raise HTTPException(status_code=404, detail="Extension not found")
    
    return FileResponse(
        extension_path,
        media_type="application/zip",
        filename="wamerce-1688-extension.zip"
    )

@app.get("/api/extension/info")
async def get_extension_info():
    """Get extension installation instructions"""
    return {
        "name": "WaMerce 1688 Importer",
        "version": "1.0.0",
        "description": "Import products from 1688.com with one click. Auto-translate Chinese to English.",
        "features": [
            "One-click import from any 1688 product page",
            "Bulk import from store/collection pages",
            "Auto-translate Chinese to English",
            "Floating button on all 1688 pages",
            "Import buttons on product cards"
        ],
        "installation": {
            "step1": "Download the extension ZIP file",
            "step2": "Open Chrome and go to chrome://extensions/",
            "step3": "Enable 'Developer mode' in the top right",
            "step4": "Click 'Load unpacked' and select the extracted folder",
            "step5": "Click the extension icon and enter your WaMerce server URL"
        },
        "download_url": "/api/extension/download"
    }


@app.get("/api/download/chrome-extension")
async def download_extension_alt():
    """Alternative download path for the WaMerce 1688 browser extension"""
    extension_path = Path("/app/wamerce-1688-extension.zip")
    if not extension_path.exists():
        raise HTTPException(status_code=404, detail="Extension not found")
    
    return FileResponse(
        extension_path,
        media_type="application/zip",
        filename="wamerce-1688-extension.zip"
    )
