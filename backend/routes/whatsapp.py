from fastapi import APIRouter, HTTPException, UploadFile, File
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime, timezone
from core import db, logger
import uuid

router = APIRouter(prefix="/whatsapp", tags=["WhatsApp Messaging"])


class WhatsAppContact(BaseModel):
    id: str
    name: str
    phone: str
    email: Optional[str] = None
    country_code: Optional[str] = "PK"
    order_number: Optional[str] = None
    sizes: Optional[List[str]] = []
    store_name: Optional[str] = None
    whatsapp_messaged: bool = False
    whatsapp_messaged_by: Optional[str] = None
    whatsapp_last_messaged_at: Optional[str] = None
    whatsapp_message_count: int = 0


class WhatsAppLinkRequest(BaseModel):
    phone: str
    country_code: Optional[str] = "PK"


class MarkMessagedRequest(BaseModel):
    agent_username: str


@router.get("/contacts")
async def get_whatsapp_contacts():
    """Get all WhatsApp contacts"""
    try:
        contacts = await db.whatsapp_contacts.find({}, {"_id": 0}).to_list(10000)
        return {
            "success": True,
            "contacts": contacts,
            "total": len(contacts)
        }
    except Exception as e:
        logger.error(f"Error fetching WhatsApp contacts: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/upload-contacts")
async def upload_whatsapp_contacts(file: UploadFile = File(...)):
    """
    Upload Excel file with WhatsApp contacts
    Expected columns: Name, Phone, Email, Country Code (optional)
    """
    try:
        import openpyxl
        from io import BytesIO
        
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        sheet = wb.active
        
        contacts_added = 0
        contacts_updated = 0
        errors = []
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row or not row[0]:  # Skip empty rows
                    continue
                
                name = str(row[0]).strip() if row[0] else ""
                phone = str(row[1]).strip() if row[1] else ""
                email = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                country_code = str(row[3]).strip().upper() if len(row) > 3 and row[3] else "PK"
                
                if not name or not phone:
                    errors.append(f"Row {row_idx}: Name and Phone are required")
                    continue
                
                # Clean phone number
                phone_clean = phone.replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
                
                # Check if contact exists
                contact_id = f"wa_{phone_clean}"
                existing = await db.whatsapp_contacts.find_one({"id": contact_id})
                
                contact_data = {
                    "id": contact_id,
                    "name": name,
                    "phone": phone_clean,
                    "email": email,
                    "country_code": country_code,
                    "source": "csv_upload",  # Mark source
                    "order_number": existing.get("order_number") if existing else None,  # Preserve order# from dashboard import
                    "sizes": existing.get("sizes", []) if existing else [],  # Preserve sizes from dashboard import
                    "store_name": existing.get("store_name") if existing else None,  # Preserve store from dashboard import
                    "whatsapp_messaged": existing.get("whatsapp_messaged", False) if existing else False,
                    "whatsapp_messaged_by": existing.get("whatsapp_messaged_by") if existing else None,
                    "whatsapp_last_messaged_at": existing.get("whatsapp_last_messaged_at") if existing else None,
                    "whatsapp_message_count": existing.get("whatsapp_message_count", 0) if existing else 0,
                    "created_at": existing.get("created_at") if existing else datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
                
                if existing:
                    await db.whatsapp_contacts.update_one(
                        {"id": contact_id},
                        {"$set": contact_data}
                    )
                    contacts_updated += 1
                else:
                    await db.whatsapp_contacts.insert_one(contact_data)
                    contacts_added += 1
                    
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
                continue
        
        return {
            "success": True,
            "message": f"Processed {contacts_added + contacts_updated} contacts",
            "contacts_added": contacts_added,
            "contacts_updated": contacts_updated,
            "errors": errors if errors else None
        }
        
    except Exception as e:
        logger.error(f"Error uploading WhatsApp contacts: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/generate-link")
async def generate_whatsapp_link(request: WhatsAppLinkRequest):
    """Generate WhatsApp link for a phone number with random greeting"""
    try:
        import random
        import urllib.parse
        
        phone = request.phone.replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
        
        # Remove leading zeros and add country code if needed
        if phone.startswith("0"):
            phone = phone[1:]
        
        # Add country code based on country_code
        country_codes = {
            "PK": "92",
            "US": "1",
            "UK": "44",
            "IN": "91",
            "UAE": "971"
        }
        
        if not phone.startswith("+"):
            country_prefix = country_codes.get(request.country_code, "92")
            if not phone.startswith(country_prefix):
                phone = f"{country_prefix}{phone}"
        
        # Random greeting templates (to avoid spam detection)
        greetings = [
            "Hi! Hope you're doing well.",
            "Hello! How are you today?",
            "Hey there! Hope you're having a great day.",
            "Hi! Trust you're doing great.",
            "Hello! Hope this message finds you well.",
            "Hey! How's everything going?",
            "Hi there! Hope you're doing fine.",
            "Hello! Greetings from our team.",
            "Hey! Hope you're having a wonderful day.",
            "Hi! Just wanted to reach out.",
            "Hello! Hope you're doing amazing.",
            "Hey there! How have you been?",
            "Hi! Hope all is well with you.",
            "Hello! Sending you warm greetings.",
            "Hey! Hope your day is going great."
        ]
        
        # Select random greeting
        greeting = random.choice(greetings)
        encoded_message = urllib.parse.quote(greeting)
        
        whatsapp_url = f"https://wa.me/{phone}?text={encoded_message}"
        
        return {
            "success": True,
            "whatsapp_url": whatsapp_url,
            "phone": phone,
            "greeting": greeting
        }
    except Exception as e:
        logger.error(f"Error generating WhatsApp link: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/mark-messaged/{contact_id}")
async def mark_contact_messaged(contact_id: str, request: MarkMessagedRequest):
    """Mark a contact as messaged by an agent"""
    try:
        contact = await db.whatsapp_contacts.find_one({"id": contact_id})
        
        if not contact:
            raise HTTPException(status_code=404, detail="Contact not found")
        
        current_count = contact.get("whatsapp_message_count", 0)
        
        await db.whatsapp_contacts.update_one(
            {"id": contact_id},
            {"$set": {
                "whatsapp_messaged": True,
                "whatsapp_messaged_by": request.agent_username,
                "whatsapp_last_messaged_at": datetime.now(timezone.utc).isoformat(),
                "whatsapp_message_count": current_count + 1
            }}
        )
        
        logger.info(f"Contact {contact_id} marked as messaged by {request.agent_username}")
        
        return {
            "success": True,
            "message": "Contact marked as messaged"
        }
    except Exception as e:
        logger.error(f"Error marking contact as messaged: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export-report")
async def export_whatsapp_report():
    """Export WhatsApp messaging report"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        from fastapi.responses import StreamingResponse
        
        # Get all contacts
        contacts = await db.whatsapp_contacts.find({}, {"_id": 0}).to_list(10000)
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Sheet 1: All Contacts
        ws1 = wb.active
        ws1.title = "All Contacts"
        
        # Headers
        headers = ["Name", "Phone", "Email", "Status", "Messaged By", "Last Messaged", "Message Count"]
        ws1.append(headers)
        
        # Style headers
        for cell in ws1[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for contact in contacts:
            ws1.append([
                contact.get("name", ""),
                contact.get("phone", ""),
                contact.get("email", ""),
                "Messaged" if contact.get("whatsapp_messaged") else "Not Messaged",
                contact.get("whatsapp_messaged_by", ""),
                contact.get("whatsapp_last_messaged_at", ""),
                contact.get("whatsapp_message_count", 0)
            ])
        
        # Sheet 2: Agent Summary
        ws2 = wb.create_sheet("Agent Summary")
        
        # Get agents
        agents = await db.agents.find({}, {"_id": 0}).to_list(100)
        
        # Headers
        ws2.append(["Agent Name", "Username", "Messages Sent"])
        
        # Style headers
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Calculate agent stats
        for agent in agents:
            username = agent.get("username")
            messaged_count = len([c for c in contacts if c.get("whatsapp_messaged_by") == username])
            ws2.append([
                agent.get("full_name", ""),
                username,
                messaged_count
            ])
        
        # Auto-size columns
        for ws in [ws1, ws2]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=whatsapp_report_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )
        
    except Exception as e:
        logger.error(f"Error exporting WhatsApp report: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/contacts/{contact_id}")
async def delete_contact(contact_id: str):
    """Delete a WhatsApp contact"""
    try:
        result = await db.whatsapp_contacts.delete_one({"id": contact_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Contact not found")
        
        return {
            "success": True,
            "message": "Contact deleted successfully"
        }
    except Exception as e:
        logger.error(f"Error deleting contact: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/contacts")
async def delete_all_contacts():
    """Delete all WhatsApp contacts"""
    try:
        result = await db.whatsapp_contacts.delete_many({})
        
        return {
            "success": True,
            "message": f"Deleted {result.deleted_count} contacts",
            "deleted_count": result.deleted_count
        }
    except Exception as e:
        logger.error(f"Error deleting all contacts: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/import-from-customers")
async def import_from_customers(
    store_name: Optional[str] = None,
    fulfillment_status: Optional[str] = None,
    limit: int = 1000
):
    """
    Import customers from dashboard/orders with order numbers and sizes
    Does NOT import colors, only sizes
    """
    try:
        # Build query
        query = {}
        
        if store_name and store_name != "all":
            query["store_name"] = store_name
        
        if fulfillment_status and fulfillment_status != "all":
            query["fulfillment_status"] = fulfillment_status
        
        # Get customers from database
        customers = await db.customers.find(query, {"_id": 0}).limit(limit).to_list(limit)
        
        imported_count = 0
        updated_count = 0
        skipped_count = 0
        
        for customer in customers:
            phone = customer.get("phone", "").strip()
            
            if not phone:
                skipped_count += 1
                continue
            
            # Clean phone number
            phone_clean = phone.replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
            
            # Extract sizes - prioritize shoe_sizes field first
            sizes = []
            
            # First, check shoe_sizes field (most reliable)
            if customer.get("shoe_sizes"):
                for size in customer.get("shoe_sizes", []):
                    size_str = str(size).strip()
                    if size_str and size_str not in sizes:
                        sizes.append(size_str)
            
            # If no shoe_sizes, try extracting from order SKUs
            if not sizes:
                order_skus = customer.get("order_skus", [])
                
                for sku in order_skus:
                    if not sku:
                        continue
                    
                    sku_str = str(sku)
                    
                    # Try different patterns:
                    # Pattern 1: "FG328B-BROWN-40" -> look for last part
                    # Pattern 2: "JN1523US-D826 BLACK-43" -> look for last part after dash
                    # Pattern 3: "78-BITE STEP OFF BLSJ-BLACK BLACK BACKGROUND-39" -> last part
                    
                    parts = sku_str.split("-")
                    
                    # Check each part from end to beginning for a size (numeric value between 35-50)
                    for part in reversed(parts):
                        part_clean = part.strip()
                        # Check if it's a numeric size (shoe sizes are typically 35-50)
                        if part_clean.replace(".", "").isdigit():
                            size_num = float(part_clean) if "." in part_clean else int(part_clean)
                            # Validate it's a reasonable shoe size
                            if 20 <= size_num <= 60:  # Wider range to catch all sizes
                                if part_clean not in sizes:
                                    sizes.append(part_clean)
                                break  # Found size for this SKU, move to next
            
            contact_id = f"wa_{phone_clean}"
            existing = await db.whatsapp_contacts.find_one({"id": contact_id})
            
            contact_data = {
                "id": contact_id,
                "name": existing.get("name") if (existing and existing.get("source") == "csv_upload") else f"{customer.get('first_name', '')} {customer.get('last_name', '')}".strip(),  # Preserve CSV name
                "phone": phone_clean,
                "email": existing.get("email") if (existing and existing.get("source") == "csv_upload") else customer.get("email", ""),  # Preserve CSV email
                "country_code": customer.get("country_code", "PK"),
                "source": existing.get("source", "dashboard_import"),  # Preserve original source
                "order_number": customer.get("order_number", ""),
                "sizes": sizes,  # Array of sizes (no colors)
                "store_name": customer.get("store_name", ""),
                "whatsapp_messaged": existing.get("whatsapp_messaged", False) if existing else False,
                "whatsapp_messaged_by": existing.get("whatsapp_messaged_by") if existing else None,
                "whatsapp_last_messaged_at": existing.get("whatsapp_last_messaged_at") if existing else None,
                "whatsapp_message_count": existing.get("whatsapp_message_count", 0) if existing else 0,
                "created_at": existing.get("created_at") if existing else datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            
            if existing:
                await db.whatsapp_contacts.update_one(
                    {"id": contact_id},
                    {"$set": contact_data}
                )
                updated_count += 1
            else:
                await db.whatsapp_contacts.insert_one(contact_data)
                imported_count += 1
        
        return {
            "success": True,
            "message": f"Imported {imported_count + updated_count} contacts from customers",
            "imported": imported_count,
            "updated": updated_count,
            "skipped": skipped_count,
            "total_processed": len(customers)
        }
        
    except Exception as e:
        logger.error(f"Error importing from customers: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
